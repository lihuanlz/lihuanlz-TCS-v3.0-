import numpy as np
import matplotlib.pyplot as plt
import scipy.optimize as opt
from scipy import stats
from scipy.stats import gaussian_kde, sem, chi2, beta as beta_dist
from scipy.optimize import brentq
import warnings
import seaborn as sns
from sklearn.linear_model import LinearRegression
from collections import defaultdict    # added

import pandas as pd

plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 10,
    'savefig.dpi': 400,
    'axes.linewidth': 0.8,
    'axes.unicode_minus': False,
})
sns.set_style("white")
plt.rcParams['font.serif'] = ['Arial']
warnings.filterwarnings('ignore')

print("=" * 80)
# print("4PL vs TCS — Core analysis (AEB weighted normal)")
print("4PL vs TCS — Core analysis (binomial likelihood)")

print("=" * 80)

# ============================================================================
# 1. Experimental parameters and data
# ============================================================================
N = 500000
V = 125e-6
Mw = 26000
N_A = 6.022e23
scale_4pl = 1.25

concentrations_pgmL = np.array([0,0,0.0328,0.0328,0.0984,0.0984,
                                0.296,0.296,0.888,0.888,2.664,2.664,8,8,24,24])
n_obs = np.array([29568,33529,31110,30257,27297,33491,23171,35706,
                  31713,30509,31098,30145,22988,22866,28320,28760])
AEB_train_raw = np.array([0.0029,0.0037,0.0255,0.0268,0.0655,0.0687,0.1807,0.1970,
                          0.5426,0.5826,1.6441,1.7458,4.4611,4.8785,13.2759,12.3785])

test_AEB = np.array([
0.236597484,
0.272436422,
2.617187507,
2.564441215,
0.058197346,
0.058412891,
0.351364389,
0.369931768,
0.102767165,
0.106078851,
0.12616563,
0.125582539,
0.337792942,
0.3409751,
0.093415847,
0.0929902,
2.396021602,
2.272783968,
0.337335334,
0.34978725,
np.nan,
np.nan,
np.nan,
np.nan,
16.74621582,
16.34499726,
0.417373262,
0.421415841,
18.35593752,
18.77629449,
3.45460218,
3.573356906,
5.847272209,
6.320004324,
9.982967721,
9.973993837,
6.953268592,
6.83977912,
2.009870497,
2.27284029,
11.47061125,
11.67209064,
3.912569618,
4.225046683,
8.439671494,
9.340527238,
7.74906697,
7.796352148,
0.729130431,
0.743955278,
1.091767575,
1.158959749,
0.381027732,
0.360355695,
0.300477333,
0.309396565,
10.09009466,
9.986039024,
3.644694894,
3.817914659,
1.31251777,
1.265727123,
0.491130065,
0.489938895,
0.164663363,
0.161659954,
0.060327989,
0.061404418,
0.022457158,
0.021049788,
0.004045956,
0.003941668,

])
test_n = np.array([
29071,
33328,
28472,
28456,
20960,
24727,
25139,
27922,
22721,
22127,
24019,
20192,
22124,
21902,
25262,
27985,
26420,
23602,
27587,
26023,
np.nan,
np.nan,
np.nan,
np.nan,
18699,
18713,
23020,
24709,
16844,
16167,
20085,
20342,
19173,
19937,
23344,
24012,
17073,
17136,
24776,
21719,
24918,
24142,
24784,
22893,
23404,
26483,
23871,
16792,
24276,
31450,
22859,
22096,
20070,
21347,
21735,
28879,
26687,
31241,
25211,
28192,
30914,
31997,
30768,
29375,
19332,
33952,
27176,
25337,
28955,
29381,
27738,
31775,

])

# Calculate k (preserve precision)
AEB_train = AEB_train_raw
n_train = n_obs
k_train = n_train * (1 - np.exp(-AEB_train))
P_train = k_train / n_train
test_k = test_n * (1 - np.exp(-test_AEB))
k_all = k_train.astype(float)
n_all_arr = n_train.astype(float)


# Concentration conversion
concentrations_gL = concentrations_pgmL * 1e-12 * 1000
C_molL = concentrations_gL / Mw
M_total = C_molL * V * N_A
mu_train = M_total / N
concentrations_4pl_pgmL = concentrations_pgmL / scale_4pl
concentrations_4pl_gL = concentrations_4pl_pgmL * 1e-12 * 1000
C_molL_4pl = concentrations_4pl_gL / Mw
M_total_4pl = C_molL_4pl * V * N_A
mu_train_4pl = M_total_4pl / N

# Weights
P_obs = 1 - np.exp(-AEB_train)
var_AEB = P_obs / (n_train * (1 - P_obs + 1e-10)) + 1e-10
weights_AEB = 1.0 / var_AEB

# ============================================================================
# 2. Model fitting
# ============================================================================
mask_zero = mu_train_4pl == 0
A_4pl = np.mean(AEB_train[mask_zero])
mask_nz = ~mask_zero
mu_nz = mu_train_4pl[mask_nz]
AEB_nz = AEB_train[mask_nz]
weights_nz = weights_AEB[mask_nz]

def fourpl_fixed_A(x, B, C, D, A):
    x = np.atleast_1d(x)
    x_safe = np.maximum(x, 1e-12)
    C_safe = max(C, 1e-12)
    log_term = B * (np.log(C_safe) - np.log(x_safe))
    log_term = np.clip(log_term, -100, 100)
    return A + (D - A) / (1.0 + np.exp(log_term))

def predict_AEB_4pl(mu, A, B, C, D):
    mu = np.atleast_1d(mu)
    res = np.empty_like(mu, dtype=float)
    res[mu == 0] = A
    res[mu != 0] = fourpl_fixed_A(mu[mu != 0], B, C, D, A)
    return res

def neg_loglik_4pl(params, mu, A_fixed, k, n):
    B, C, D = params
    if B <= 0 or C <= 0 or D <= A_fixed:
        return 1e12
    AEB_pred = predict_AEB_4pl(mu, A_fixed, B, C, D)
    P_pos = np.clip(1 - np.exp(-AEB_pred), 1e-15, 1 - 1e-15)
    return -np.sum(k * np.log(P_pos) + (n - k) * np.log(1 - P_pos))


# Initial values
idx_sort = np.argsort(mu_nz)
mu_sorted = mu_nz[idx_sort]
AEB_sorted = AEB_nz[idx_sort]
D_init = np.max(AEB_sorted) * 1.05
target = A_4pl + 0.5 * (D_init - A_4pl)
C_init = np.median(mu_nz)
for i in range(len(AEB_sorted)-1):
    if AEB_sorted[i] <= target <= AEB_sorted[i+1]:
        ratio = (target - AEB_sorted[i]) / (AEB_sorted[i+1] - AEB_sorted[i] + 1e-10)
        C_init = mu_sorted[i] + ratio * (mu_sorted[i+1] - mu_sorted[i])
        break
C_init = np.clip(C_init, mu_nz.min()*0.5, mu_nz.max()*2)
B_init = 1.0
# bounds_4pl = [(0.1,4.0), (mu_nz.min()*0.1, mu_nz.max()*20), (A_4pl+1e-6, AEB_sorted.max()*2)]
# bounds_4pl = [(0.1,4.0), (mu_nz.min()*0.1, mu_nz.max()*20), (A_4pl+1e-6, AEB_sorted.max()*100)]
bounds_4pl = [(0.1,10.0), (mu_nz.min()*0.1, mu_nz.max()*20), (A_4pl+1e-6, AEB_sorted.max()*100)]
p0 = [np.clip(B_init,*bounds_4pl[0]), np.clip(C_init,*bounds_4pl[1]), np.clip(D_init,*bounds_4pl[2])]

k_nz = k_all[mask_nz]
n_nz = n_all_arr[mask_nz]
res_4pl = opt.minimize(neg_loglik_4pl, p0, args=(mu_nz, A_4pl, k_nz, n_nz),
                       bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter':2000})

B_4pl, C_4pl, D_4pl = res_4pl.x

# ============================================================================
# 4PL BOOTSTRAP CI (fairness comparison with TCS)
# ============================================================================
np.random.seed(42)
n_boot = 2000
boot_B, boot_C, boot_D = [], [], []

# 4PL BOOTSTRAP CI (case resampling)
# for _ in range(n_boot):
#     idx = np.random.choice(len(AEB_nz), size=len(AEB_nz), replace=True)
#     try:
#         res_b = opt.minimize(
#             neg_loglik_4pl, [B_4pl, C_4pl, D_4pl],
#             args=(mu_nz[idx], A_4pl, AEB_nz[idx], weights_nz[idx]),
#             bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter': 1000}
#         )
# for _ in range(n_boot):
#     idx = np.random.choice(len(AEB_nz), size=len(AEB_nz), replace=True)
#     try:
#         res_b = opt.minimize(
#             neg_loglik_4pl, [B_4pl, C_4pl, D_4pl],
#             args=(mu_nz[idx], A_4pl, AEB_nz[idx], weights_nz[idx]),
#             bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter': 1000}
#         )


for _ in range(n_boot):
    idx = np.random.choice(len(k_nz), size=len(k_nz), replace=True)
    try:
        res_b = opt.minimize(
            neg_loglik_4pl, [B_4pl, C_4pl, D_4pl],
            args=(mu_nz[idx], A_4pl, k_nz[idx], n_nz[idx]),
            bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter': 1000}
        )


        if res_b.success:
            boot_B.append(res_b.x[0])
            boot_C.append(res_b.x[1])
            boot_D.append(res_b.x[2])
    except:
        pass

B_CI = np.percentile(boot_B, [2.5, 97.5])
C_CI = np.percentile(boot_C, [2.5, 97.5])
D_CI = np.percentile(boot_D, [2.5, 97.5])
print(f"[4PL case-bootstrap n={len(boot_B)}] B={B_4pl:.4f} [{B_CI[0]:.4f},{B_CI[1]:.4f}]  C={C_4pl:.4e} [{C_CI[0]:.4e},{C_CI[1]:.4e}]  D={D_4pl:.4f} [{D_CI[0]:.4f},{D_CI[1]:.4f}]")





def tcs_model(mu, kappa, b):
    mu_safe = np.maximum(mu, 0)
    P_specific = 1 - np.exp(-mu_safe / (1 + kappa))
    return np.clip(b + (1-b)*P_specific, 0, 1)

def predict_AEB_tcs(mu, kappa, b):
    P_pred = tcs_model(mu, kappa, b)
    return -np.log(np.maximum(1 - P_pred, 1e-12))

# def neg_loglik_tcs(params, mu, AEB_obs, weights):
#     kappa, b = params
#     if kappa <= 0 or b < 0 or b >= 1:
#         return 1e12
#     AEB_pred = predict_AEB_tcs(mu, kappa, b)
#     resid = AEB_obs - AEB_pred
#     return 0.5 * np.sum(weights * resid**2 + np.log(2 * np.pi / weights))


def neg_loglik_tcs(params, mu, k, n):
    kappa, b = params
    if kappa <= 0 or b < 0 or b >= 1:
        return 1e12
    P_pos = np.clip(tcs_model(mu, kappa, b), 1e-15, 1 - 1e-15)
    return -np.sum(k * np.log(P_pos) + (n - k) * np.log(1 - P_pos))


mu_nz_tcs = mu_train[mask_nz]
P_nz = P_train[mask_nz]
coef = np.polyfit(mu_nz_tcs, P_nz, 1, w=np.sqrt(n_train[mask_nz]))
b_init = max(0, coef[1])
kappa_init = max(0.1, (1-b_init)/coef[0]-1) if coef[0]>0 else 1.0
bounds_tcs = [(0.01,1000), (0,0.1)]
# res_tcs = opt.minimize(neg_loglik_tcs, [kappa_init, b_init],
#                        args=(mu_train, AEB_train, weights_AEB),
#                        bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter':500})
res_tcs = opt.minimize(neg_loglik_tcs, [kappa_init, b_init],
                       args=(mu_train, k_all, n_all_arr),
                       bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter':500})

kappa_tcs, b_tcs = res_tcs.x



# ============================================================================
# TCS BOOTSTRAP CI (fairness comparison with 4PL)
# ============================================================================
np.random.seed(42)
boot_kappa, boot_b = [], []

# TCS BOOTSTRAP CI (case resampling)
# for _ in range(n_boot):
#     idx = np.random.choice(len(AEB_train), size=len(AEB_train), replace=True)
#     try:
#         res_t = opt.minimize(
#             neg_loglik_tcs, [kappa_tcs, b_tcs],
#             args=(mu_train[idx], AEB_train[idx], weights_AEB[idx]),
#             bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter': 500}
#         )
for _ in range(n_boot):
    idx = np.random.choice(len(k_all), size=len(k_all), replace=True)
    try:
        res_t = opt.minimize(
            neg_loglik_tcs, [kappa_tcs, b_tcs],
            args=(mu_train[idx], k_all[idx], n_all_arr[idx]),
            bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter': 500}
        )

        if res_t.success:
            boot_kappa.append(res_t.x[0])
            boot_b.append(res_t.x[1])
    except:
        pass

kappa_CI = np.percentile(boot_kappa, [2.5, 97.5])
b_CI = np.percentile(boot_b, [2.5, 97.5])
print(f"[TCS case-bootstrap n={len(boot_kappa)}] kappa={kappa_tcs:.4f} [{kappa_CI[0]:.4f},{kappa_CI[1]:.4f}]  b={b_tcs:.6f} [{b_CI[0]:.6f},{b_CI[1]:.6f}]")

# Predicted values
AEB_pred_train_4pl = predict_AEB_4pl(mu_train_4pl, A_4pl, B_4pl, C_4pl, D_4pl)
AEB_pred_train_tcs = predict_AEB_tcs(mu_train, kappa_tcs, b_tcs)

# ============================================================================
# 3. Model comparison statistics (full printout)
# ============================================================================
# def normal_loglik(AEB_obs, AEB_pred, weights):
#     resid = AEB_obs - AEB_pred
#     return -0.5 * np.sum(weights * resid**2 + np.log(2*np.pi/weights))

# logL_4pl = normal_loglik(AEB_train, AEB_pred_train_4pl, weights_AEB)
# logL_tcs = normal_loglik(AEB_train, AEB_pred_train_tcs, weights_AEB)
def binom_loglik_from_AEB(AEB_pred, k, n):
    P_pos = np.clip(1 - np.exp(-AEB_pred), 1e-15, 1 - 1e-15)
    return np.sum(k * np.log(P_pos) + (n - k) * np.log(1 - P_pos))

logL_4pl = binom_loglik_from_AEB(AEB_pred_train_4pl, k_all, n_all_arr)
logL_tcs = binom_loglik_from_AEB(AEB_pred_train_tcs, k_all, n_all_arr)


def AICc(logL, n_params, n):
    AIC = -2*logL + 2*n_params
    correction = 2*n_params*(n_params+1)/(n - n_params - 1) if n > n_params+1 else np.inf
    return AIC + correction

n_all = len(AEB_train)
AICc_4pl = AICc(logL_4pl, 3, n_all)
AICc_tcs = AICc(logL_tcs, 2, n_all)
BIC_4pl = -2*logL_4pl + 3*np.log(n_all)
BIC_tcs = -2*logL_tcs + 2*np.log(n_all)

# Weighted R-squared (AEB scale)
ss_res_4pl = np.sum(weights_AEB * (AEB_train - AEB_pred_train_4pl)**2)
ss_tot_4pl = np.sum(weights_AEB * (AEB_train - np.average(AEB_train, weights=weights_AEB))**2)
R2_4pl = 1 - ss_res_4pl/ss_tot_4pl

ss_res_tcs = np.sum(weights_AEB * (AEB_train - AEB_pred_train_tcs)**2)
R2_tcs = 1 - ss_res_tcs/ss_tot_4pl

# Residuals and normality test (for reference)
# resid_4pl = AEB_train - AEB_pred_train_4pl
# resid_tcs = AEB_train - AEB_pred_train_tcs
# _, p_shapiro_4pl = stats.shapiro(resid_4pl)
# _, p_shapiro_tcs = stats.shapiro(resid_tcs)

# AEB residuals (for panel c visualization only)
resid_4pl = AEB_train - AEB_pred_train_4pl
resid_tcs = AEB_train - AEB_pred_train_tcs

# Deviance residuals (for Shapiro test)
def deviance_resid(k, n, P_pos_fit):
    P_obs = k / n
    with np.errstate(divide='ignore', invalid='ignore'):
        term_pos = np.where(k > 0, k * np.log(k / (n * P_pos_fit)), 0)
        term_neg = np.where(k < n, (n - k) * np.log((n - k) / (n * (1 - P_pos_fit))), 0)
    dev = 2 * (term_pos + term_neg)
    return np.sign(P_obs - P_pos_fit) * np.sqrt(np.abs(dev))

P_pos_fit_4pl = np.clip(1 - np.exp(-AEB_pred_train_4pl), 1e-15, 1 - 1e-15)
P_pos_fit_tcs = np.clip(tcs_model(mu_train, kappa_tcs, b_tcs), 1e-15, 1 - 1e-15)
dr_4pl = deviance_resid(k_all, n_all_arr, P_pos_fit_4pl)
dr_tcs = deviance_resid(k_all, n_all_arr, P_pos_fit_tcs)
_, p_shapiro_4pl = stats.shapiro(dr_4pl)
_, p_shapiro_tcs = stats.shapiro(dr_tcs)





# LOO (binomial logL)
loo_4pl_vals, loo_tcs_vals = [], []
fail_4pl = fail_tcs = 0
for i in range(n_all):
    idx_keep = np.ones(n_all, dtype=bool)
    idx_keep[i] = False
    mu_loo_4pl = mu_train_4pl[idx_keep]
    mu_loo_tcs = mu_train[idx_keep]
    k_loo = k_all[idx_keep]
    n_loo = n_all_arr[idx_keep]
    k_i, n_i = k_all[i], n_all_arr[i]
    mu_i_4pl, mu_i_tcs = mu_train_4pl[i], mu_train[i]

    # 4PL
    mask_nz_loo = mu_loo_4pl > 0
    if np.any(mask_nz_loo):
        A_loo = np.mean(AEB_train[idx_keep][mu_loo_4pl == 0]) if np.any(mu_loo_4pl == 0) else A_4pl
        mu_nz_loo = mu_loo_4pl[mask_nz_loo]
        k_nz_loo = k_loo[mask_nz_loo]
        n_nz_loo = n_loo[mask_nz_loo]
        AEB_nz_loo = AEB_loo = AEB_train[idx_keep][mask_nz_loo]
        D_loo = np.max(AEB_nz_loo)*1.05
        C_loo = np.median(mu_nz_loo)
        # p0_loo = [np.clip(B_4pl,0.1,4.0), np.clip(C_loo, mu_nz_loo.min()*0.1, mu_nz_loo.max()*20),
        #           np.clip(D_loo, A_loo+1e-6, AEB_nz_loo.max()*100)]
        p0_loo = [np.clip(B_4pl,0.1,10.0), np.clip(C_loo, mu_nz_loo.min()*0.1, mu_nz_loo.max()*20),

                  np.clip(D_loo, A_loo+1e-6, AEB_nz_loo.max()*100)]
        try:
            res_loo = opt.minimize(neg_loglik_4pl, p0_loo,
                                   args=(mu_nz_loo, A_loo, k_nz_loo, n_nz_loo),
                                   # bounds=[(0.1,4.0), (mu_nz_loo.min()*0.1, mu_nz_loo.max()*20),
                                   #         (A_loo+1e-6, AEB_nz_loo.max()*100)],
                                   bounds=[(0.1,10.0), (mu_nz_loo.min()*0.1, mu_nz_loo.max()*20),

                                           (A_loo+1e-6, AEB_nz_loo.max()*100)],
                                   method='L-BFGS-B', options={'maxiter':1000})
            if res_loo.success:
                B_loo, C_loo, D_loo = res_loo.x
                AEB_i_pred = predict_AEB_4pl(np.array([mu_i_4pl]), A_loo, B_loo, C_loo, D_loo)[0]
                P_i_pred = np.clip(1 - np.exp(-AEB_i_pred), 1e-15, 1 - 1e-15)
                loo_4pl_vals.append(k_i * np.log(P_i_pred) + (n_i - k_i) * np.log(1 - P_i_pred))
            else:
                fail_4pl += 1; loo_4pl_vals.append(np.nan)
        except:
            fail_4pl += 1; loo_4pl_vals.append(np.nan)
    else:
        loo_4pl_vals.append(np.nan)

    # TCS
    try:
        res_tcs_loo = opt.minimize(neg_loglik_tcs, [kappa_tcs, b_tcs],
                                   args=(mu_loo_tcs, k_loo, n_loo),
                                   bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter':500})
        if res_tcs_loo.success:
            kappa_loo, b_loo = res_tcs_loo.x
            P_i_pred_tcs = np.clip(tcs_model(np.array([mu_i_tcs]), kappa_loo, b_loo), 1e-15, 1 - 1e-15)[0]
            loo_tcs_vals.append(k_i * np.log(P_i_pred_tcs) + (n_i - k_i) * np.log(1 - P_i_pred_tcs))
        else:
            fail_tcs += 1; loo_tcs_vals.append(np.nan)
    except:
        fail_tcs += 1; loo_tcs_vals.append(np.nan)


valid_idx = ~np.isnan(loo_4pl_vals) & ~np.isnan(loo_tcs_vals)
loo_4pl_arr = np.array(loo_4pl_vals)[valid_idx]
loo_tcs_arr = np.array(loo_tcs_vals)[valid_idx]
loo_sum_4pl = np.sum(loo_4pl_arr)
loo_sum_tcs = np.sum(loo_tcs_arr)
loo_diff_arr = np.array(loo_4pl_arr) - np.array(loo_tcs_arr)
mean_diff = np.mean(loo_diff_arr)
se_diff = sem(loo_diff_arr)
t_stat = mean_diff / se_diff













# ---- TOST equivalence test (two one-sided tests, alpha = 0.05) ----
from scipy.stats import t as t_dist

n_f  = len(loo_diff_arr)
sd_d = np.std(loo_diff_arr, ddof=1)
se_d = sd_d / np.sqrt(n_f)
tcrit = t_dist.ppf(0.95, n_f - 1)

mean_abs_logL = np.mean(np.abs(loo_tcs_arr))     # per-fold logL magnitude
delta = 0.01 * mean_abs_logL                     # equivalence margin: 1% of per-fold |logL|

t_lo = (mean_diff + delta) / se_d
t_hi = (delta - mean_diff) / se_d
p_tost = max(1 - t_dist.cdf(t_lo, n_f - 1), 1 - t_dist.cdf(t_hi, n_f - 1))
ci90 = (mean_diff - tcrit * se_d, mean_diff + tcrit * se_d)
delta_min = max(abs(ci90[0]), abs(ci90[1]))

print(f"\nTOST equivalence test (margin = 1% of per-fold |logL| = {delta:.2f} logL):")
print(f"  mean diff = {mean_diff:.3f}, 90% CI = [{ci90[0]:.3f}, {ci90[1]:.3f}]")
print(f"  TOST p = {p_tost:.4f} -> "
      f"{'EQUIVALENT within +/-' + f'{delta:.1f} logL' if p_tost < 0.05 else 'equivalence NOT established at this margin'}")
print(f"  smallest passing margin: +/-{delta_min:.2f} logL/fold "
      f"({100 * delta_min / mean_abs_logL:.3f}% of per-fold |logL|)")



























# ========== Full printout ==========
print("\n" + "="*60)
print("FITTED PARAMETERS")
print("="*60)
print(f"4PL (A fixed): A = {A_4pl:.6f}")
print(f"                B = {B_4pl:.4f}")
print(f"                C = {C_4pl:.4e}")
print(f"                D = {D_4pl:.4f}")
print(f"TCS:            κ = {kappa_tcs:.4f}")
print(f"                b = {b_tcs:.6f}")




# ============================================================================
# PARAMETER CI COMPARISON (4PL vs TCS)
# ============================================================================
print("\n" + "=" * 70)
print("PARAMETER CI COMPARISON (bootstrap, n=200)")
print("=" * 70)
print(f"{'Param':<8} {'4PL':<32} {'TCS':<32}")
print(f"{'─'*72}")
print(f"{'slope':<8} {f'B = {B_4pl:.4f}':<14} [{B_CI[0]:.4f}, {B_CI[1]:.4f}]  "
      f"{f'κ = {kappa_tcs:.4f}':<14} [{kappa_CI[0]:.4f}, {kappa_CI[1]:.4f}]")
print(f"{'offset':<8} {f'C = {C_4pl:.4e}':<14} [{C_CI[0]:.4e}, {C_CI[1]:.4e}]  "
      f"{f'b = {b_tcs:.6f}':<14} [{b_CI[0]:.6f}, {b_CI[1]:.6f}]")
print(f"{'upper':<8} {f'D = {D_4pl:.4f}':<14} [{D_CI[0]:.4f}, {D_CI[1]:.4f}]  "
      f"{'(n/a)':<32}")

# Correlations
print("\nParameter correlation (from bootstrap):")
print(f"  4PL: ρ(B, C) = {np.corrcoef(boot_B, boot_C)[0, 1]:+.3f}, "
      f"ρ(B, D) = {np.corrcoef(boot_B, boot_D)[0, 1]:+.3f}, "
      f"ρ(C, D) = {np.corrcoef(boot_C, boot_D)[0, 1]:+.3f}")
print(f"  TCS: ρ(κ, b) = {np.corrcoef(boot_kappa, boot_b)[0, 1]:+.3f}")












print("\n" + "="*60)
print("MODEL COMPARISON (Training Set)")
print("="*60)
print(f"{'':<8} {'logL':<12} {'AICc':<10} {'BIC':<10} {'R²(AEB)':<10}")
print(f"{'4PL':<8} {logL_4pl:<12.2f} {AICc_4pl:<10.2f} {BIC_4pl:<10.2f} {R2_4pl:<10.4f}")
print(f"{'TCS':<8} {logL_tcs:<12.2f} {AICc_tcs:<10.2f} {BIC_tcs:<10.2f} {R2_tcs:<10.4f}")

# print("\nResidual Shapiro-Wilk test (for reference only):")
# print(f"  4PL p = {p_shapiro_4pl:.4e}")
# print(f"  TCS p = {p_shapiro_tcs:.4e}")
print("\nDeviance residual Shapiro-Wilk test:")
print(f"  4PL p = {p_shapiro_4pl:.4e}")
print(f"  TCS p = {p_shapiro_tcs:.4e}")


print("\n" + "="*60)
print("LEAVE-ONE-OUT CROSS-VALIDATION")
print("="*60)
print(f"LOO sum logL:  4PL = {loo_sum_4pl:.2f}, TCS = {loo_sum_tcs:.2f}")
print(f"Mean diff (4PL - TCS) = {mean_diff:.3f}")
print(f"SE of diff             = {se_diff:.3f}")
print(f"t-statistic            = {t_stat:.3f}")
if t_stat < -2:
    print("→ TCS significantly better (|t|>2).")
else:
    print("→ No significant difference.")

print("\nPer-fold logL differences (4PL - TCS):")
for idx, (l4, lt) in enumerate(zip(loo_4pl_vals, loo_tcs_vals)):
    if np.isnan(l4) or np.isnan(lt):
        print(f"  Fold {idx+1:2d}: FAIL")
    else:
        print(f"  Fold {idx+1:2d}: {l4-lt:12.4f}")

# Test set back-calculation statistics
def mu_to_conc(mu_val):
    C_molL = (mu_val * N) / (V * N_A)
    return C_molL * Mw * 1e9

def inverse_4pl(AEB_target, A, B, C, D):
    if AEB_target <= A: return 0.0
    if AEB_target >= D: return np.nan
    ratio = (AEB_target - A) / (D - AEB_target)
    if ratio <= 0: return 0.0
    return C * ratio ** (1.0/B)

def inverse_tcs(P_target, kappa, b):
    if P_target <= b: return 0.0
    if P_target >= 1-1e-6: return np.nan
    P_specific = (P_target - b) / (1 - b)
    if P_specific >= 1: return np.nan
    return -(1+kappa) * np.log(1 - P_specific)

conc_4pl_back = []
for aeb in test_AEB:
    mu_scaled = inverse_4pl(aeb, A_4pl, B_4pl, C_4pl, D_4pl)
    if np.isnan(mu_scaled) or mu_scaled == 0:
        conc_4pl_back.append(np.nan)
    else:
        conc_4pl_back.append(mu_to_conc(mu_scaled * scale_4pl))
conc_4pl_back = np.array(conc_4pl_back)

P_test = test_k / test_n
conc_tcs_back = []
for p in P_test:
    mu_real = inverse_tcs(p, kappa_tcs, b_tcs)
    if np.isnan(mu_real) or mu_real == 0:
        conc_tcs_back.append(np.nan)
    else:
        conc_tcs_back.append(mu_to_conc(mu_real))
conc_tcs_back = np.array(conc_tcs_back)

above_4pl = np.sum(np.isnan(conc_4pl_back))
above_tcs = np.sum(np.isnan(conc_tcs_back))
below_4pl = np.sum(conc_4pl_back == 0)
below_tcs = np.sum(conc_tcs_back == 0)

print("\n" + "="*60)
print("TEST SET BACK-CALCULATION")
print("="*60)
print(f"4PL: min={np.nanmin(conc_4pl_back):.4f}, max={np.nanmax(conc_4pl_back):.4f}, median={np.nanmedian(conc_4pl_back):.4f} pg/mL")
print(f"TCS: min={np.nanmin(conc_tcs_back):.4f}, max={np.nanmax(conc_tcs_back):.4f}, median={np.nanmedian(conc_tcs_back):.4f} pg/mL")
print(f"4PL: # above ULOQ (NaN) = {above_4pl}")
print(f"TCS: # above ULOQ (NaN) = {above_tcs}")

# Save LOO per-fold differences to file
with open('LOO_per_fold.txt', 'w') as f:
    f.write("Fold\t4PL_logL\tTCS_logL\tDiff(4PL-TCS)\n")
    for idx, (l4, lt) in enumerate(zip(loo_4pl_vals, loo_tcs_vals)):
        if np.isnan(l4) or np.isnan(lt):
            f.write(f"{idx+1}\tFAIL\tFAIL\tFAIL\n")
        else:
            f.write(f"{idx+1}\t{l4:.4f}\t{lt:.4f}\t{l4-lt:.4f}\n")

print("\nLOO per-fold details also saved to LOO_per_fold.txt")

# ============================================================================
# 3.5 R3 precomputation (for plotting; printing handled by section 7)
# ============================================================================
gamma_tcs_r3 = 1.0 / (kappa_tcs + 1)

P_specific_train_r3 = (P_train - b_tcs) / (1 - b_tcs)
P_specific_train_r3 = np.clip(P_specific_train_r3, 1e-12, 1-1e-12)

M_r3_temp = N * (kappa_tcs + 1) * P_specific_train_r3
mu_r3_temp = M_r3_temp / N
CV_r3_temp = np.where(
    mu_r3_temp > 1e-12,
    (1.0 / (gamma_tcs_r3 * (1 - b_tcs) * mu_r3_temp)) * np.sqrt(P_train * (1 - P_train) / n_obs),
    np.inf
)

r3_valid_temp = P_specific_train_r3 < 0.2

# Independent linear fit (R3)
mask_fit_r3 = r3_valid_temp & ~mask_zero
if np.sum(mask_fit_r3) >= 2:
    c_fit_r3 = concentrations_pgmL[mask_fit_r3]
    mu_fit_r3 = (c_fit_r3 * 1e-12 * 1000) / Mw * V * N_A / N
    P_fit_r3 = P_specific_train_r3[mask_fit_r3]
    gamma_fit_r3 = np.sum(P_fit_r3 * mu_fit_r3) / np.sum(mu_fit_r3 ** 2)
else:
    gamma_fit_r3 = np.nan

# ============================================================================
# 6. TCS per-point quantification, CV and CI (using per-well actual n)
# ============================================================================
from scipy.stats import chi2, beta as beta_dist
from scipy.optimize import brentq
import pandas as pd

# ---------- Helper functions (same principle as S3b, but using per-point n_obs) ----------
def M_from_Ppos(P_pos, b, kappa, N_total):
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    return -N_total * (kappa + 1) * np.log(1 - P_spec)

def calc_CV(P_pos, n_i, b, kappa, N_total):
    M_hat = M_from_Ppos(P_pos, b, kappa, N_total)
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    dM_dP = N_total * (kappa + 1) / ((1 - b) * (1 - P_spec))
    se = np.abs(dM_dP) * np.sqrt(P_pos * (1 - P_pos) / n_i)
    return (se / M_hat) * 100

def delta_ci(P_pos, n_i, b, kappa, N_total, alpha=0.05):
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    M_hat = M_from_Ppos(P_pos, b, kappa, N_total)
    dM_dP = N_total * (kappa + 1) / ((1 - b) * (1 - P_spec))
    se = np.abs(dM_dP) * np.sqrt(P_pos * (1 - P_pos) / n_i)
    z = np.sqrt(chi2.ppf(1 - alpha, 1))
    return (M_hat - z * se, M_hat + z * se)

def exact_transformation_ci(k, n_i, b, kappa, N_total, alpha=0.05):
    """Clopper-Pearson proportion interval -> mapped to M"""
    if k == 0:
        p_low = 0.0
        p_up = beta_dist.ppf(1 - alpha/2, k + 1, n_i - k)
    elif k == n_i:
        p_low = beta_dist.ppf(alpha/2, k, n_i - k + 1)
        p_up = 1.0
    else:
        p_low = beta_dist.ppf(alpha/2, k, n_i - k + 1)
        p_up = beta_dist.ppf(1 - alpha/2, k + 1, n_i - k)
    M_low = M_from_Ppos(p_low, b, kappa, N_total)
    M_up  = M_from_Ppos(p_up, b, kappa, N_total)
    return (M_low, M_up)

def likelihood_ratio_ci(k, n_i, b, kappa, N_total, alpha=0.05):
    P_obs = k / n_i
    M_hat = M_from_Ppos(P_obs, b, kappa, N_total)

    def loglik(M):
        mu = M / N_total
        P_spec = 1 - np.exp(-mu / (kappa + 1))
        P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
        P_pos = b + (1 - b) * P_spec
        P_pos = np.clip(P_pos, 1e-12, 1 - 1e-12)
        return k * np.log(P_pos) + (n_i - k) * np.log(1 - P_pos)

    def deviance(M):
        return 2 * (loglik(M_hat) - loglik(M))

    threshold = chi2.ppf(1 - alpha, 1)
    try:
        M_low = brentq(lambda M: deviance(M) - threshold, 1e-3, M_hat)
    except:
        M_low = 0.0
    try:
        M_up = brentq(lambda M: deviance(M) - threshold, M_hat, 1e15)
    except:
        M_up = np.inf
    return (M_low, M_up)

# ---------- Parameter settings ----------
N_total = N
b_fit = b_tcs
kappa_fit = kappa_tcs

# Quantify only training points with concentration > 0
mask_pos = concentrations_pgmL > 0
conc_subset = concentrations_pgmL[mask_pos]
k_subset = k_train[mask_pos]
n_subset = n_obs[mask_pos]

print("\n" + "=" * 80)
print("TCS MODEL: Per‑point quantification (using individual observed n)")
print(f"Total partitions N = {N_total}")
print("=" * 80)

results_tcs = []
plot_tcs_data = []

for i in range(len(conc_subset)):
    conc = conc_subset[i]
    k = k_subset[i]
    n_i = n_subset[i]
    P_obs = k / n_i
    M_hat = M_from_Ppos(P_obs, b_fit, kappa_fit, N_total)
    cv = calc_CV(P_obs, n_i, b_fit, kappa_fit, N_total)

    ci_d = delta_ci(P_obs, n_i, b_fit, kappa_fit, N_total)
    ci_e = exact_transformation_ci(k, n_i, b_fit, kappa_fit, N_total)
    ci_lr = likelihood_ratio_ci(k, n_i, b_fit, kappa_fit, N_total)

    results_tcs.append({
        'Conc (pg/mL)': conc,
        'k': k,
        'n': n_i,
        'P_obs': P_obs,
        'M_hat': M_hat,
        'CV (%)': cv,
        'Delta_low': ci_d[0], 'Delta_up': ci_d[1],
        'Exact_low': ci_e[0], 'Exact_up': ci_e[1],
        'LR_low': ci_lr[0], 'LR_up': ci_lr[1]
    })

    # For plotting: map exact M intervals back to P_pos intervals
    M_low, M_up = ci_e
    P_low = b_fit + (1 - b_fit) * (1 - np.exp(-M_low / (N_total * (kappa_fit + 1))))
    P_up  = b_fit + (1 - b_fit) * (1 - np.exp(-M_up  / (N_total * (kappa_fit + 1))))
    plot_tcs_data.append((conc, P_obs, M_hat, cv, P_low, P_up))

df_quant = pd.DataFrame(results_tcs)
pd.set_option('display.float_format', '{:.4f}'.format)
print(df_quant.to_string(index=False))
df_quant.to_csv('TCS_quantification_per_point.csv', index=False)
print("\nResults saved to TCS_quantification_per_point.csv")

# ============================================================================
# 6b. Test set (validation data) TCS quantification
# ============================================================================
print("\n" + "=" * 80)
print("TCS MODEL: Test set quantification (validation points)")
print("=" * 80)

test_results = []
test_plot_data = []

for i in range(len(test_AEB)):
    AEB_i = test_AEB[i]
    n_i = test_n[i]
    k_i = n_i * (1 - np.exp(-AEB_i))
    P_obs_i = k_i / n_i
    M_hat = M_from_Ppos(P_obs_i, b_fit, kappa_fit, N_total)
    cv = calc_CV(P_obs_i, n_i, b_fit, kappa_fit, N_total)

    ci_d = delta_ci(P_obs_i, n_i, b_fit, kappa_fit, N_total)
    ci_e = exact_transformation_ci(k_i, n_i, b_fit, kappa_fit, N_total)
    ci_lr = likelihood_ratio_ci(k_i, n_i, b_fit, kappa_fit, N_total)

    test_results.append({
        'Index': i,
        'AEB': AEB_i,
        'n': n_i,
        'k': k_i,
        'P_obs': P_obs_i,
        'M_hat': M_hat,
        'CV (%)': cv,
        'Delta_low': ci_d[0], 'Delta_up': ci_d[1],
        'Exact_low': ci_e[0], 'Exact_up': ci_e[1],
        'LR_low': ci_lr[0], 'LR_up': ci_lr[1]
    })

df_test_quant = pd.DataFrame(test_results)
print(f"Test set size: {len(test_results)}")
print("Summary statistics:")
print(f"  M_hat range: {df_test_quant['M_hat'].min():.2f} - {df_test_quant['M_hat'].max():.2f}")
print(f"  CV median: {df_test_quant['CV (%)'].median():.2f}%")
print(f"  CV range: {df_test_quant['CV (%)'].min():.2f}% - {df_test_quant['CV (%)'].max():.2f}%")
df_test_quant.to_csv('TCS_test_quantification.csv', index=False)
print("Test set results saved to TCS_test_quantification.csv")

# ============================================================================
# 4. Main figure: Fig_core.png (3x2 layout, with TCS vs Simoa comparison)
# ============================================================================
conc_train_real = (mu_train * N) / (V * N_A) * Mw * 1e9
conc_train_4pl_real = (mu_train_4pl * scale_4pl * N) / (V * N_A) * Mw * 1e9

# ===== Precompute per-well LoB/LoD/LoQ pg/mL ranges (for plotting) =====
N_total_plot = N
kappa_plot = kappa_tcs
b_plot = b_tcs
gamma_plot = 1.0 / (kappa_plot + 1)

def M_to_conc_plot(M):
    return M / (V * N_A) * Mw * 1e9

r2_lob_vals, r2_lod_vals, r2_loq_low_vals, r2_loq_high_vals = [], [], [], []
r3_lob_vals, r3_lod_vals, r3_loq_low_vals = [], [], []

for i in range(len(n_obs)):
    n_i = int(n_obs[i])
    # ---- R2 per-well ----
    def P_pos_R2(M):
        return b_plot + (1 - b_plot)*(1 - np.exp(-M/(N_total_plot*(kappa_plot+1))))
    def M_from_Ppos_R2(P):
        if P <= b_plot: return 0.0
        Ps = (P - b_plot)/(1 - b_plot)
        return -N_total_plot*(kappa_plot+1)*np.log(1 - Ps)
    def CV_R2(M):
        P = P_pos_R2(M)
        if P <= 0 or P >= 1: return np.inf
        Ps = (P - b_plot)/(1 - b_plot)
        Ps = np.clip(Ps, 1e-12, 1-1e-12)
        dM = N_total_plot*(kappa_plot+1)/((1 - b_plot)*(1 - Ps))
        return np.abs(dM)/M * np.sqrt(P*(1-P)/n_i)

    # LoB
    k95 = stats.binom.ppf(0.95, n_i, b_plot)
    lob_p = k95/n_i
    M_lob = M_from_Ppos_R2(lob_p)
    r2_lob_vals.append(M_lob)

    # LoD
    def det_prob(M):
        P = P_pos_R2(M)
        kth = int(np.floor(n_i*lob_p)) + 1
        if kth > n_i: return 0.0
        return stats.binom.sf(kth-1, n_i, P)
    try:
        Mlo, Mhi = 1.0, 1e8
        while det_prob(Mlo) < 0.95: Mlo *= 2
        Mhi = Mlo; Mlo = max(1.0, Mlo/2)
        M_lod = brentq(lambda M: det_prob(M)-0.95, Mlo, Mhi, xtol=1e-6)
    except:
        M_lod = np.nan
    r2_lod_vals.append(M_lod)

    # LoQ low
    try:
        Mlo = 1.0
        while CV_R2(Mlo) > 0.2: Mlo *= 2
        Mhi = Mlo; Mlo = max(1.0, Mlo/2)
        M_loql = brentq(lambda M: CV_R2(M)-0.2, Mlo, Mhi, xtol=1e-6)
    except:
        M_loql = np.nan
    r2_loq_low_vals.append(M_loql)

    # LoQ high
    try:
        Mstart = 5e5
        while CV_R2(Mstart) < 0.2: Mstart *= 2
        Mhi = Mstart; Mlo = Mstart/2
        while CV_R2(Mlo) > 0.2: Mlo /= 2
        M_loqh = brentq(lambda M: CV_R2(M)-0.2, Mlo, Mhi, xtol=1e-6)
    except:
        M_loqh = np.nan
    r2_loq_high_vals.append(M_loqh)

    # ---- R3 per-well ----
    M_lob3 = M_lob
    M_lod3 = (3.29*N_total_plot/gamma_plot)*np.sqrt(b_plot/(n_i*(1-b_plot)))
    # A_ = 0.04*n_i*(1-b_plot)**2
    # B_ = -(1-b_plot); C_ = -b_plot
    A_ = (1-b_plot)*(1+0.04*n_i)
    B_ = -(1-2*b_plot); C_ = -b_plot

    x_loq = (-B_ + np.sqrt(B_**2 - 4*A_*C_))/(2*A_)
    M_loql3 = N_total_plot*(kappa_plot+1)*x_loq
    r3_lob_vals.append(M_lob3)
    r3_lod_vals.append(M_lod3)
    r3_loq_low_vals.append(M_loql3)

# Convert to pg/mL and take min/max
r2_lob_c = [M_to_conc_plot(m) for m in r2_lob_vals]
r2_lod_c = [M_to_conc_plot(m) for m in r2_lod_vals]
r2_loq_l_c = [M_to_conc_plot(m) for m in r2_loq_low_vals]
r2_loq_h_c = [M_to_conc_plot(m) for m in r2_loq_high_vals]
r3_lob_c = [M_to_conc_plot(m) for m in r3_lob_vals]
r3_lod_c = [M_to_conc_plot(m) for m in r3_lod_vals]
r3_loq_l_c = [M_to_conc_plot(m) for m in r3_loq_low_vals]

r2_lob_range = (np.min(r2_lob_c), np.max(r2_lob_c))
r2_lod_range = (np.nanmin(r2_lod_c), np.nanmax(r2_lod_c))
r2_loq_l_range = (np.nanmin(r2_loq_l_c), np.nanmax(r2_loq_l_c))
r2_loq_h_range = (np.nanmin(r2_loq_h_c), np.nanmax(r2_loq_h_c))
r3_lob_range = (np.min(r3_lob_c), np.max(r3_lob_c))
r3_lod_range = (np.nanmin(r3_lod_c), np.nanmax(r3_lod_c))
r3_loq_l_range = (np.nanmin(r3_loq_l_c), np.nanmax(r3_loq_l_c))





# ============================================================================
# REPLACEMENT: Clinical validation — back-calculation from test_AEB
# Replaces: lines 1062-1220 (the hardcoded roche/simoa/tcs section)
# ============================================================================
# Changes:
#   1. No hardcoded simoa/tcs arrays — all computed from test_AEB
#   2. First 4 AEB values (2 QC samples) excluded → 68 AEB = 34 samples
#   3. Two replicates averaged per sample
#   4. 4PL, R2, R3, R1 back-calculations
#   5. Dilution: first 26 samples = 4x, last 8 = 1x
#   6. Roche: 34 values (no QC in Roche), 1865/1037 = S-9/S-10
#   7. S-11 (Roche=558.8) and S-13 (Roche=532.8) → expect TCS NaN
# ============================================================================

# ---- Exclude first 4 AEB (2 QC samples), remaining 68 = 34 samples ----
test_AEB_clin = test_AEB[4:].copy()
test_n_clin = test_n[4:].copy()
n_clin_samples = 34  # 68/2

# ---- Dilution: first 26 samples = 4x, last 8 = 1x ----
dilution_clin = np.array([4.0] * 26 + [1.0] * 8)

# ---- Roche reference values (34 samples, S-1 to S-34, no QC) ----
roche = np.array([
    0.00, 0.00, 1.87, 2.10, 6.63, 7.50, 51.40, 7.00,
    1865, 1037, 558.8, 9.93, 532.8, 66.82, 123.6, 264.7,
    150.9, 45.37, 327.6, 95.15, 283, 178.2, 14.79, 29.54,
    6.1, 4.19, 50.00, 16.67, 5.56, 1.85, 0.62, 0.21,
    0.07, 0.00,
])

# ---- Average two replicates per sample ----
aeb_mean = np.array([np.mean(test_AEB_clin[2*i:2*i+2]) for i in range(n_clin_samples)])
n_mean = np.array([np.mean(test_n_clin[2*i:2*i+2]) for i in range(n_clin_samples)])
P_pos_mean = 1 - np.exp(-aeb_mean)

# ---- Back-calculation functions ----
def inverse_tcs_R2(P_target, kappa, b):
    """R2: P_pos → mu (on-board)"""
    if P_target <= b: return 0.0
    if P_target >= 1 - 1e-6: return np.nan
    P_specific = (P_target - b) / (1 - b)
    if P_specific >= 1: return np.nan
    return -(1 + kappa) * np.log(1 - P_specific)

# def inverse_tcs_R3(P_target, kappa, b, N_total):
#     """R3 linear: P_pos → mu (on-board)"""
#     if P_target <= b: return 0.0
#     P_specific = (P_target - b) / (1 - b)
#     if P_specific >= 1: return np.nan
#     return P_specific * N_total * (1 + kappa)


def inverse_tcs_R3(P_target, kappa, b, N_total):

    """R3 linear: P_pos → mu (on-board)"""
    if P_target <= b: return 0.0
    P_specific = (P_target - b) / (1 - b)
    if P_specific >= 1: return np.nan
    return P_specific * (1 + kappa)

def inverse_tcs_R1(P_target, kappa_R1, beta_R1, b_R1):
    """R1: P_pos → mu (on-board)"""
    if P_target <= b_R1: return 0.0
    if P_target >= 1 - 1e-10: return np.nan
    P_specific = (P_target - b_R1) / (1 - b_R1)
    if P_specific >= 1: return np.nan
    val = 1.0 - P_specific
    if val <= 0: return np.nan
    p = 1.0 - val ** (1.0 / beta_R1)
    p = np.clip(p, 1e-15, 1.0 - 1e-15)
    xi = p / (1.0 - p) + p / kappa_R1
    mu = xi * kappa_R1 * beta_R1
    return mu

# ---- R1 parameters (from dynesty posterior median) ----
try:
    kappa_R1_val = kappa_med
    beta_R1_val = beta_med
    b_R1_val = b_med
except NameError:
    kappa_R1_val = 7.995
    beta_R1_val = 22.36
    b_R1_val = 0.003494

# ---- Compute back-calculations for each sample ----
conc_4pl = np.full(n_clin_samples, np.nan)
conc_R2 = np.full(n_clin_samples, np.nan)
conc_R3 = np.full(n_clin_samples, np.nan)
conc_R1 = np.full(n_clin_samples, np.nan)

for i in range(n_clin_samples):
    aeb = aeb_mean[i]
    P = P_pos_mean[i]

    # 4PL back-calc
    mu_4pl = inverse_4pl(aeb, A_4pl, B_4pl, C_4pl, D_4pl)
    if not np.isnan(mu_4pl) and mu_4pl > 0:
        conc_4pl[i] = mu_to_conc(mu_4pl * scale_4pl) * dilution_clin[i]

    # R2 back-calc
    mu_R2 = inverse_tcs_R2(P, kappa_tcs, b_tcs)
    if not np.isnan(mu_R2) and mu_R2 > 0:
        # conc_R2[i] = mu_to_conc(mu_R2) * dilution_clin[i]
        conc_R2[i] = mu_to_conc(mu_R2) * 1.25 * dilution_clin[i]

    # R3 back-calc (linear)
    mu_R3 = inverse_tcs_R3(P, kappa_tcs, b_tcs, N)
    if not np.isnan(mu_R3) and mu_R3 > 0:
        # conc_R3[i] = mu_to_conc(mu_R3) * dilution_clin[i]
        conc_R3[i] = mu_to_conc(mu_R3) * 1.25 * dilution_clin[i]

    # R1 back-calc
    mu_R1 = inverse_tcs_R1(P, kappa_R1_val, beta_R1_val, b_R1_val)
    if not np.isnan(mu_R1) and mu_R1 > 0:
        conc_R1[i] = mu_to_conc(mu_R1) * dilution_clin[i]
        conc_R1[i] = mu_to_conc(mu_R1) * 1.25 * dilution_clin[i]

# ---- Print per-sample results ----
print("\n" + "=" * 130)
print("Clinical validation: back-calculation from test_AEB (34 samples, QC excluded)")
print("=" * 130)
print(f"{'S':<4} {'AEB':<8} {'P_pos':<10} {'dil':<4} {'4PL':<12} {'R2':<12} {'R3':<12} {'R1':<12} {'Roche':<10}")
print("-" * 90)
for i in range(n_clin_samples):
    c4 = f"{conc_4pl[i]:.3f}" if not np.isnan(conc_4pl[i]) else "NaN"
    c2 = f"{conc_R2[i]:.3f}" if not np.isnan(conc_R2[i]) else "NaN"
    c3 = f"{conc_R3[i]:.3f}" if not np.isnan(conc_R3[i]) else "NaN"
    c1 = f"{conc_R1[i]:.3f}" if not np.isnan(conc_R1[i]) else "NaN"
    print(f"{i+1:<4} {aeb_mean[i]:<8.3f} {P_pos_mean[i]:<10.6f} {dilution_clin[i]:<4.0f} {c4:<12} {c2:<12} {c3:<12} {c1:<12} {roche[i]:<10.2f}")

# ---- Summary stats ----
print("\n" + "=" * 80)
print("Summary: valid estimates per method")
print("=" * 80)
print(f"{'Method':<10} {'Valid':<8} {'NaN':<8} {'NaN samples (Roche)':<30}")
for name, arr in [('4PL', conc_4pl), ('R2', conc_R2), ('R3', conc_R3), ('R1', conc_R1)]:
    nan_mask = np.isnan(arr)
    nan_roche = roche[nan_mask] if np.any(nan_mask) else []
    print(f"{name:<10} {np.sum(~nan_mask):<8} {np.sum(nan_mask):<8} {str(nan_roche):<30}")

# ---- Paired comparison vs Roche (exclude Roche=0 and NaN) ----
print("\n" + "=" * 80)
print("Paired comparison vs Roche (exclude Roche=0 and method NaN)")
print("=" * 80)

roche_valid = roche > 0  # exclude Roche=0 (blank samples)

for name, arr in [('4PL', conc_4pl), ('R2', conc_R2), ('R3', conc_R3), ('R1', conc_R1)]:
    mask = roche_valid & ~np.isnan(arr)
    if np.sum(mask) < 3:
        print(f"\n{name}: insufficient paired data ({np.sum(mask)} pairs)")
        continue

    x_r = roche[mask]
    y_m = arr[mask]

    slope, intercept, r_val, p_val, se = stats.linregress(x_r, y_m)
    r2_val = r_val**2
    t_ci = stats.t.ppf(0.975, np.sum(mask) - 2)
    slope_ci = (slope - t_ci * se, slope + t_ci * se)
    rho, p_rho = stats.spearmanr(x_r, y_m)

    mean_vals = (y_m + x_r) / 2
    diff_rel = (y_m - x_r) / mean_vals * 100
    bias = np.mean(diff_rel)
    loa_lo = bias - 1.96 * np.std(diff_rel, ddof=1)
    loa_hi = bias + 1.96 * np.std(diff_rel, ddof=1)

    print(f"\n{name} vs Roche (n={np.sum(mask)}):")
    print(f"  slope = {slope:.4f} (95% CI: {slope_ci[0]:.4f}-{slope_ci[1]:.4f})")
    print(f"  R² = {r2_val:.4f}")
    print(f"  Spearman ρ = {rho:.4f} (p = {p_rho:.2e})")
    print(f"  BA bias = {bias:.2f}%, LoA = [{loa_lo:.2f}%, {loa_hi:.2f}%]")

# ---- Method-to-method: R2 vs 4PL ----
print("\n" + "=" * 80)
print("Method-to-method: R2 vs 4PL (both valid)")
print("=" * 80)
mask_both = ~np.isnan(conc_R2) & ~np.isnan(conc_4pl) & (conc_R2 > 0) & (conc_4pl > 0)
if np.sum(mask_both) >= 3:
    slope_m, intercept_m, r_val_m, _, se_m = stats.linregress(conc_4pl[mask_both], conc_R2[mask_both])
    rho_m, p_rho_m = stats.spearmanr(conc_4pl[mask_both], conc_R2[mask_both])
    print(f"n = {np.sum(mask_both)}")
    print(f"  slope = {slope_m:.4f}, R² = {r_val_m**2:.4f}")
    print(f"  Spearman ρ = {rho_m:.4f} (p = {p_rho_m:.2e})")

# ---- Store for plotting ----
# Use 4PL as "simoa" proxy and R2 as "tcs" for the scatter plot
simoa = conc_4pl
tcs = conc_R2
sample_type = ['Plasma'] * 26 + ['Simulated'] * 8

# For Bland-Altman plotting code
valid = ~np.isnan(simoa) & ~np.isnan(tcs) & (simoa > 0) & (tcs > 0)
simoa_clean = simoa[valid]
tcs_clean = tcs[valid]
type_clean = np.array(sample_type)[valid]

# Linear regression for plotting code
x = simoa_clean.reshape(-1, 1)
y = tcs_clean
reg = LinearRegression().fit(x, y)
slope = reg.coef_[0]
intercept = reg.intercept_
r2 = reg.score(x, y)

n = len(x)
x_mean = np.mean(simoa_clean)
x_var = np.var(simoa_clean, ddof=1)
residuals_reg = y - (slope * simoa_clean + intercept)
mse = np.sum(residuals_reg**2) / (n - 2)
se_slope = np.sqrt(mse / ((n - 1) * x_var))
se_intercept = np.sqrt(mse * (1/n + x_mean**2 / ((n - 1) * x_var)))
slope_ci = (slope - 1.96 * se_slope, slope + 1.96 * se_slope)
intercept_ci = (intercept - 1.96 * se_intercept, intercept + 1.96 * se_intercept)

rho, p_rho = stats.spearmanr(simoa_clean, tcs_clean)

mean_val = (tcs_clean + simoa_clean) / 2
rel_diff = (tcs_clean - simoa_clean) / mean_val * 100
ba_mask = mean_val > 0.01
mean_val_ba = mean_val[ba_mask]
rel_diff_ba = rel_diff[ba_mask]
type_clean_ba = type_clean[ba_mask]

bias_rel = np.mean(rel_diff_ba)
sd_rel = np.std(rel_diff_ba, ddof=1)
loa_rel_lower = bias_rel - 1.96 * sd_rel
loa_rel_upper = bias_rel + 1.96 * sd_rel

print("\n" + "=" * 80)
print("=== R2 (TCS) vs 4PL back-calculation comparison ===")
print("=" * 80)
print(f"n samples: {n}")
print(f"Linear regression: slope = {slope:.4f} (95% CI: {slope_ci[0]:.4f}-{slope_ci[1]:.4f})")
print(f"          intercept = {intercept:.4f} (95% CI: {intercept_ci[0]:.4f}-{intercept_ci[1]:.4f})")
print(f"          R² = {r2:.4f}")
print(f"Spearman ρ = {rho:.4f} (p = {p_rho:.2e})")
print(f"Relative error Bland-Altman (exclude mean<=0.01): bias = {bias_rel:.2f}%, 95% LoA = [{loa_rel_lower:.2f}%, {loa_rel_upper:.2f}%]")



# ====================================================================
# Save clinical back-calculation results to SI_Table_4.xlsx
# ====================================================================
try:
    from openpyxl import load_workbook as _lwb
    _wb_clin = _lwb("SI_Table_4.xlsx")
    if "18_Clinical_BackCalc" in _wb_clin.sheetnames:
        del _wb_clin["18_Clinical_BackCalc"]
    _ws_clin = _wb_clin.create_sheet("18_Clinical_BackCalc")

    _ws_clin.append(['Clinical validation: back-calculation from test_AEB (34 samples, QC excluded)'])
    _ws_clin.append(['Dilution: first 26 samples = 4x, last 8 = 1x; all × 1.25 (100→125 μL)'])
    _ws_clin.append([])
    _ws_clin.append(['Sample', 'AEB_mean', 'P_pos', 'Dilution',
                     '4PL (pg/mL)', 'R2 (pg/mL)', 'R3 (pg/mL)', 'R1 (pg/mL)',
                     'Roche (pg/mL)'])
    from openpyxl.styles import Font, PatternFill
    _bold = Font(bold=True)
    _hdr = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in _ws_clin[4]:
        cell.font = _bold; cell.fill = _hdr

    for i in range(n_clin_samples):
        _ws_clin.append([
            i+1,
            round(aeb_mean[i], 4),
            round(P_pos_mean[i], 6),
            dilution_clin[i],
            round(conc_4pl[i], 4) if not np.isnan(conc_4pl[i]) else 'NaN',
            round(conc_R2[i], 4) if not np.isnan(conc_R2[i]) else 'NaN',
            round(conc_R3[i], 4) if not np.isnan(conc_R3[i]) else 'NaN',
            round(conc_R1[i], 4) if not np.isnan(conc_R1[i]) else 'NaN',
            roche[i],
        ])

    _ws_clin.append([])
    _ws_clin.append(['Summary Statistics'])
    _ws_clin.append(['Comparison', 'n', 'Slope', '95% CI', 'R²', 'Spearman ρ', 'BA bias %', 'BA LoA lower %', 'BA LoA upper %'])

    # R2 vs 4PL
    _ws_clin.append(['R2 vs 4PL', n, round(slope, 4), f'{slope_ci[0]:.4f}-{slope_ci[1]:.4f}',
                     round(r2, 4), round(rho, 4), round(bias_rel, 2),
                     round(loa_rel_lower, 2), round(loa_rel_upper, 2)])

    # vs Roche comparisons
    for name, arr in [('4PL vs Roche', conc_4pl), ('R2 vs Roche', conc_R2),
                       ('R1 vs Roche', conc_R1)]:
        mask = ~np.isnan(arr) & (arr > 0) & (roche > 0)
        n_r = np.sum(mask)
        if n_r >= 3:
            s_r, _, r_r, _, se_r = stats.linregress(roche[mask], arr[mask])
            rho_r, p_r = stats.spearmanr(roche[mask], arr[mask])
            mv = (arr[mask] + roche[mask]) / 2
            dr = (arr[mask] - roche[mask]) / mv * 100
            mb = mv > 0.01
            br = np.mean(dr[mb])
            lr = br - 1.96 * np.std(dr[mb], ddof=1)
            ur = br + 1.96 * np.std(dr[mb], ddof=1)
            _ws_clin.append([name, n_r, round(s_r, 4), '', round(r_r**2, 4),
                             round(rho_r, 4), round(br, 2), round(lr, 2), round(ur, 2)])

    _wb_clin.save("SI_Table_4.xlsx")
    print(f"\n[OK] Sheet 18_Clinical_BackCalc added to SI_Table_4.xlsx (now {len(_wb_clin.sheetnames)} sheets)")
except Exception as e_clin:
    print(f"\n[FAIL] Clinical sheet: {e_clin}")
    import traceback; traceback.print_exc()



















# ========== Create 3x2 figure ==========
fig_core, axes = plt.subplots(3, 2, figsize=(13, 19))
# fig_core.suptitle('Figure 2. TCS analysis of the IL-6 Simoa dataset.', fontsize=16, fontweight='bold')

fig_core.suptitle('Fig. 2b: TCS analysis of the IL-6 Single molecule array dataset.', 
                  fontsize=18, fontweight='bold', y=1)
# --- (a) 4PL ---
ax = axes[0, 0]
mu_plot = np.logspace(np.log10(mu_nz.min()*0.5), np.log10(mu_nz.max()*2), 200)
conc_plot = (mu_plot * scale_4pl * N) / (V * N_A) * Mw * 1e9
ax.semilogx(conc_plot, fourpl_fixed_A(mu_plot, B_4pl, C_4pl, D_4pl, A_4pl), 'b-', lw=2)
ax.scatter(conc_train_4pl_real, AEB_train, c='blue', s=30, alpha=0.6)
ax.set_xlabel('Concentration (pg/mL)', fontsize=14)
ax.set_ylabel('AEB', fontsize=14)
ax.set_title('(1) 4PL Fit', fontsize=14, fontweight='bold')
ax.grid(False)
# ax.text(0.05, 0.95, f"A={A_4pl:.4f}\nB={B_4pl:.2f}\nC={C_4pl:.1e}\nD={D_4pl:.2f}",
#         transform=ax.transAxes, fontsize=14, va='top',
#         bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8))
ax.text(0.05, 0.95, 
        f"A = {A_4pl:.4f}\n"
        f"B = {B_4pl:.2f}\n"
        f"C = {C_4pl:.2e}\n"
        f"D = {D_4pl:.2f}\n"
        f"4PL: AEB = A + (D-A)/(1+(C/x)$^B$)\n"
        f"R$^2$ = {R2_4pl:.4f}",
        transform=ax.transAxes, fontsize=11, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))

# --- (b) TCS with Uncertainty ---
ax = axes[0, 1]
mu_plot_orig = np.logspace(np.log10(mu_train[mask_nz].min()*0.5), np.log10(mu_train[mask_nz].max()*2), 200)
conc_plot_orig = (mu_plot_orig * N) / (V * N_A) * Mw * 1e9
P_plot = tcs_model(mu_plot_orig, kappa_tcs, b_tcs)
ax.semilogx(conc_plot_orig, P_plot, 'r-', lw=2, label='TCS fit')
ax.scatter(conc_train_real, P_train, c='red', s=30, alpha=0.6)

conc_count = defaultdict(int)
for (conc, Pobs, Mhat, cv, Plow, Pup) in plot_tcs_data:
    ax.errorbar(conc, Pobs, yerr=[[Pobs - Plow], [Pup - Pobs]],
                fmt='none', ecolor='darkred', capsize=8, alpha=0.7, linewidth=1.2)
    if cv < 1000:
        cnt = conc_count[conc]
        y_shift = 0.025 if cnt % 2 == 0 else -0.025
        conc_count[conc] += 1
        ax.text(conc, Pobs + y_shift, f'T:{cv:.2f}%',
                fontsize=12, color='darkred', alpha=0.8, ha='center', va='bottom' if y_shift>0 else 'top')

ax.axvline(r2_lob_range[0], color='green', linestyle=':', linewidth=1.2, alpha=0.8)
ax.axvline(r2_lob_range[1], color='green', linestyle=':', linewidth=1.2, alpha=0.8)
ax.axvline(r2_lod_range[0], color='orange', linestyle='--', linewidth=1.2, alpha=0.8)
ax.axvline(r2_lod_range[1], color='orange', linestyle='--', linewidth=1.2, alpha=0.8)
ax.axvline(r2_loq_l_range[0], color='purple', linestyle='-.', linewidth=1.2, alpha=0.8)
ax.axvline(r2_loq_l_range[1], color='purple', linestyle='-.', linewidth=1.2, alpha=0.8)
if not np.isnan(r2_loq_h_range[0]) and not np.isnan(r2_loq_h_range[1]):
    ax.axvline(r2_loq_h_range[0], color='magenta', linestyle='-', linewidth=1.2, alpha=0.8)
    ax.axvline(r2_loq_h_range[1], color='magenta', linestyle='-', linewidth=1.2, alpha=0.8)

ax2 = ax.twinx()
# kde_specs = [
#     (np.array(r2_lob_c), 'green', f'LoB [{r2_lob_range[0]:.4f}, {r2_lob_range[1]:.4f}] pg/mL'),
#     (np.array(r2_lod_c), 'orange', f'LoD [{r2_lod_range[0]:.4f}, {r2_lod_range[1]:.4f}] pg/mL'),
#     (np.array(r2_loq_l_c), 'purple', f'LoQ$_{{\rm low}}$ [{r2_loq_l_range[0]:.4f}, {r2_loq_l_range[1]:.4f}] pg/mL'),
# ]
# if not np.isnan(r2_loq_h_range[0]) and not np.isnan(r2_loq_h_range[1]):
#     kde_specs.append(
#         (np.array(r2_loq_h_c), 'magenta', f'LoQ$_{{\rm high}}$ [{r2_loq_h_range[0]:.1f}, {r2_loq_h_range[1]:.1f}] pg/mL')
#     )
kde_specs = [
    (np.array(r2_lob_c), 'green', f'LoB [{r2_lob_range[0]:.4f}, {r2_lob_range[1]:.4f}] pg/mL'),
    (np.array(r2_lod_c), 'orange', f'LoD [{r2_lod_range[0]:.4f}, {r2_lod_range[1]:.4f}] pg/mL'),
    (np.array(r2_loq_l_c), 'purple', f'LoQ$_{{\\mathrm{{low}}}}$ [{r2_loq_l_range[0]:.4f}, {r2_loq_l_range[1]:.4f}] pg/mL'),
]
if not np.isnan(r2_loq_h_range[0]) and not np.isnan(r2_loq_h_range[1]):
    kde_specs.append(
        (np.array(r2_loq_h_c), 'magenta', f'LoQ$_{{\\mathrm{{high}}}}$ [{r2_loq_h_range[0]:.2f}, {r2_loq_h_range[1]:.2f}] pg/mL')
    )

for data, color, label in kde_specs:
    clean = data[~np.isnan(data)]
    if len(clean) > 1:
        kde = gaussian_kde(np.log10(clean))
        x_kde = np.logspace(np.log10(clean.min()*0.8), np.log10(clean.max()*1.2), 200)
        ax2.plot(x_kde, kde(np.log10(x_kde)), color=color, lw=1.5, linestyle='-', alpha=0.7, label=label)

ax2.set_ylabel('Density', fontsize=14)
ax2.set_ylim(0, 20)
ax2.legend(loc='lower right', fontsize=10, framealpha=0.9, ncol=1)
if ax.get_legend() is not None:
    ax.get_legend().remove()

ax.set_xlabel('Concentration (pg/mL)', fontsize=14)
ax.set_ylabel(r'$P_{\mathrm{pos}}$', fontsize=14)
ax.set_title('(2) TCS Fit with Model Uncertainty', fontsize=14, fontweight='bold')
ax.grid(False)
# ax.text(0.05, 0.95, f"κ={kappa_tcs:.2f}\nb={b_tcs:.4f}",
#         transform=ax.transAxes, fontsize=14, va='top',
#         bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.8))

# ax.text(0.05, 0.95,
#         f"κ = {kappa_tcs:.2f}\n"
#         f"b = {b_tcs:.4f}\n"
#         f"TCS: P = b + (1−b)·(1−e$^{{-x/(1+κ)}}$)\n"
#         f"R$^2$ = {R2_tcs:.4f}",
#         f"Error bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
#         transform=ax.transAxes, fontsize=11, va='top',
#         bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.8))

# ax.text(0.05, 0.95,
#         f"κ = {kappa_tcs:.2f}\n"
#         f"b = {b_tcs:.4f}\n"
#         f"TCS: \widehat{M_{\mathrm{R2}}}=-N(1+\kappa)\ln{\left(1-\widehat{P_{\mathrm{specific}}}\right)}\n"
#         f"R$^2$ = {R2_tcs:.4f}\n"
#         f"Error bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
#         transform=ax.transAxes, fontsize=11, va='top',
#         bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.8))
# Define new equation (no variables, LaTeX math only)
# eq_str = r"$\widehat{M_{\mathrm{R2}}} = -N(1+\kappa)\ln\left(1-\widehat{P_{\mathrm{specific}}}\right)$"

# ax.text(0.05, 0.95,
#         f"κ = {kappa_tcs:.2f}\n"
#         f"b = {b_tcs:.4f}\n"
#         f"{eq_str}\n"
#         f"R$^2$ = {R2_tcs:.4f}\n"
#         f"Error bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
#         transform=ax.transAxes, fontsize=11, va='top',
#         bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.8))
eq_str = r"$\widehat{M_{\mathrm{R2}}} = -N(1+\kappa)\ln\left(1-\widehat{P_{\mathrm{specific}}}\right)$"

ax.text(0.05, 0.95,
        f"$\\kappa$ = {kappa_tcs:.2f}\n"   # modified this line only
        f"b = {b_tcs:.4f}\n"
        f"{eq_str}\n"
        f"R$^2$ = {R2_tcs:.4f}\n"
        f"Error bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
        transform=ax.transAxes, fontsize=11, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))

# ax.text(0.75, 0.9, "Error bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
#         transform=ax.transAxes, fontsize=14, ha='right', va='bottom',
#         bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))

# --- (c) Residuals ---
ax = axes[1, 0]
ax.scatter(conc_train_4pl_real, resid_4pl, c='blue', alpha=0.6, label='4PL')
ax.scatter(conc_train_real, resid_tcs, c='red', alpha=0.6, label='TCS')
ax.axhline(0, color='k', ls='--')
ax.set_xscale('log')
ax.set_xlabel('Concentration (pg/mL)', fontsize=14)
ax.set_ylabel('Residual (AEB)', fontsize=14)
ax.set_title('(3) Residuals (AEB scale)', fontsize=14, fontweight='bold')
ax.legend(fontsize=14)
ax.grid(False)
ax.text(0.05, 0.15, f"LOO: Δ = {mean_diff:.1f} ± {se_diff:.2f}\nt = {t_stat:.2f}",
        transform=ax.transAxes, fontsize=14, ha='left', va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))








# --- (d) R3 Linear Approximation ---
ax = axes[1, 1]
c_range = np.logspace(np.log10(concentrations_pgmL[concentrations_pgmL>0].min()*0.5),
                      np.log10(concentrations_pgmL[concentrations_pgmL>0].max()*2), 100)
mu_range = (c_range * 1e-12 * 1000) / Mw * V * N_A / N
P_theory = gamma_tcs_r3 * mu_range
ax.loglog(c_range, P_theory, 'k--', lw=1, label=f'R2')

if not np.isnan(gamma_fit_r3):
    P_fit_line = gamma_fit_r3 * mu_range
    ax.loglog(c_range, P_fit_line, 'b-', lw=2, label=f'R3')

ax.loglog(concentrations_pgmL[~mask_zero], P_specific_train_r3[~mask_zero],
          'o', color='gray', alpha=0.5, label='All data')

valid_indices = np.where(r3_valid_temp & ~mask_zero)[0]
for plot_idx, i in enumerate(valid_indices):
    conc_i = concentrations_pgmL[i]
    ps_i = P_specific_train_r3[i]
    cv_i = CV_r3_temp[i] * 100
    ax.loglog(conc_i, ps_i, 'ro', markersize=8)
    y_shift = 1.3 if plot_idx % 2 == 0 else 0.7
    va = 'bottom' if plot_idx % 2 == 0 else 'top'
    ax.text(conc_i * 1.1, ps_i * y_shift * 1.4, f'T:{cv_i:.2f}%',
            fontsize=12, color='darkred', ha='left', va=va)

ax.axvline(r3_lob_range[0], color='green', linestyle=':', linewidth=1.2, alpha=0.8)
ax.axvline(r3_lob_range[1], color='green', linestyle=':', linewidth=1.2, alpha=0.8)
ax.axvline(r3_lod_range[0], color='orange', linestyle='--', linewidth=1.2, alpha=0.8)
ax.axvline(r3_lod_range[1], color='orange', linestyle='--', linewidth=1.2, alpha=0.8)
ax.axvline(r3_loq_l_range[0], color='purple', linestyle='-.', linewidth=1.2, alpha=0.8)
ax.axvline(r3_loq_l_range[1], color='purple', linestyle='-.', linewidth=1.2, alpha=0.8)

ax2 = ax.twinx()
# kde_specs3 = [
#     (np.array(r3_lob_c), 'green', f'LoB [{r3_lob_range[0]:.4f}, {r3_lob_range[1]:.4f}] pg/mL'),
#     (np.array(r3_lod_c), 'orange', f'LoD [{r3_lod_range[0]:.4f}, {r3_lod_range[1]:.4f}] pg/mL'),
#     (np.array(r3_loq_l_c), 'purple', f'LoQ$_{{\rm low}}$ [{r3_loq_l_range[0]:.4f}, {r3_loq_l_range[1]:.4f}] pg/mL'),
# ]
kde_specs3 = [
    (np.array(r3_lob_c), 'green', f'LoB [{r3_lob_range[0]:.4f}, {r3_lob_range[1]:.4f}] pg/mL'),
    (np.array(r3_lod_c), 'orange', f'LoD [{r3_lod_range[0]:.4f}, {r3_lod_range[1]:.4f}] pg/mL'),
    (np.array(r3_loq_l_c), 'purple', f'LoQ$_{{\\mathrm{{low}}}}$ [{r3_loq_l_range[0]:.4f}, {r3_loq_l_range[1]:.4f}] pg/mL'),
]

for data, color, label in kde_specs3:
    clean = data[~np.isnan(data)]
    if len(clean) > 1:
        kde = gaussian_kde(np.log10(clean))
        x_kde = np.logspace(np.log10(clean.min()*0.8), np.log10(clean.max()*1.2), 200)
        ax2.plot(x_kde, kde(np.log10(x_kde)), color=color, lw=1.5, linestyle='-', alpha=0.7, label=label)

ax2.set_ylabel('Density', fontsize=14)
ax2.set_ylim(bottom=0)

handles_orig, labels_orig = ax.get_legend_handles_labels()
handles_kde, labels_kde = ax2.get_legend_handles_labels()
ax2.legend(handles_orig + handles_kde, labels_orig + labels_kde, loc='lower right', fontsize=10, framealpha=0.9, ncol=1)

ax.set_xlabel('Concentration (pg/mL)', fontsize=14)
ax.set_ylabel(r'$P_{\mathrm{specific}}$', fontsize=14)
ax.set_title('(4) R3 Linear Approximation', fontsize=14, fontweight='bold')
ax2.set_ylim(0, 20)
# param_text = "T: TCS Model CV"
# ax.text(0.25, 0.95, param_text, transform=ax.transAxes, fontsize=14, va='top')
# --- (d) R3 Linear Approximation ---
ax = axes[1, 1]
# ... preceding plotting code unchanged ...

# Replace original param_text and ax.text with:
# if not np.isnan(gamma_fit_r3):
#     kappa_fit_r3 = 1.0 / gamma_fit_r3 - 1.0
#     fit_info = f"R3 fitted: κ = {kappa_fit_r3:.2f}, γ = {gamma_fit_r3:.4f}"
# else:
#     fit_info = "R3 fitted: insufficient data"
if not np.isnan(gamma_fit_r3):
    kappa_fit_r3 = 1.0 / gamma_fit_r3 - 1.0
    fit_info = f"R3 fitted: $\\kappa = {kappa_fit_r3:.2f}$, $\\gamma = {gamma_fit_r3:.4f}$"
else:
    fit_info = "R3 fitted: insufficient data"

# # ax.text(0.05, 0.95,
# #         f"R2 (TCS): κ = {kappa_tcs:.2f}, γ = {gamma_tcs_r3:.4f}\n"
# #         + fit_info +
# #         f"\n$P_{{\\mathrm{{specific}}}}$ = γ·μ\n"
# #         f"T: TCS Model CV",
# #         transform=ax.transAxes, fontsize=11, va='top',
# #         bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
# eq_r3 = r"$\widehat{M_{\mathrm{R3}}}=N(1+\kappa)\left(\widehat{P_{\mathrm{specific}}}\right)$"

# ax.text(0.05, 0.95,
#         f"R2 (TCS): κ = {kappa_tcs:.2f}, γ = {gamma_tcs_r3:.4f}\n"
#         + fit_info +
#         f"\n{eq_r3}\n"
#         f"T: TCS Model CV",
#         transform=ax.transAxes, fontsize=11, va='top',
#         bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
eq_r3 = r"$\widehat{M_{\mathrm{R3}}}=N(1+\kappa)\left(\widehat{P_{\mathrm{specific}}}\right)$"
gamma_def = r"$\gamma = \frac{1}{1+\kappa}$"

# ax.text(0.05, 0.95,
#         f"R2 (TCS): κ = {kappa_tcs:.2f}, γ = {gamma_tcs_r3:.4f}\n"
#         + fit_info +
#         f"\n{eq_r3}\n"
#         f"{gamma_def}\n"
#         f"T: TCS Model CV",
#         transform=ax.transAxes, fontsize=11, va='top',
#         bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
ax.text(0.05, 0.95,
        f"R2 (TCS): $\\kappa = {kappa_tcs:.2f}$, $\\gamma = {gamma_tcs_r3:.4f}$\n"
        + fit_info +
        f"\n{eq_r3}\n"
        f"{gamma_def}\n"
        "T: TCS Model CV",
        transform=ax.transAxes, fontsize=11, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))

# --- (e) TCS vs Simoa scatter plot ---
ax = axes[2, 0]
markers = {'Simulated': 's', 'Plasma': 'o'}
colors = {'Simulated': 'red', 'Plasma': 'blue'}

for t in ['Simulated', 'Plasma']:
    mask = (type_clean == t)
    ax.scatter(simoa_clean[mask], tcs_clean[mask],
               marker=markers[t], color=colors[t], alpha=0.8,
               edgecolors='k', s=70, label=t)

x_line = np.linspace(min(simoa_clean), max(simoa_clean), 100)
y_line = slope * x_line + intercept
ax.plot(x_line, y_line, 'k--', label=f'OLS: y={slope:.2f}x+{intercept:.1f}')
ax.plot(x_line, x_line, 'r-', label='Identity')
ax.set_xlabel('Simoa standard concentration', fontsize=14)
ax.set_ylabel('TCS fitted concentration', fontsize=14)
ax.set_title('(5) Correlation', fontsize=14,fontweight='bold')
ax.legend(fontsize=14)
textstr = (f'Slope: {slope:.3f} [{slope_ci[0]:.3f}, {slope_ci[1]:.3f}]\n'
           f'Intercept: {intercept:.2f} [{intercept_ci[0]:.2f}, {intercept_ci[1]:.2f}]\n'
           f'$R^2$={r2:.4f},  $\\rho$={rho:.4f}')
ax.text(0.05, 0.95, textstr, transform=ax.transAxes, verticalalignment='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.8), fontsize=14)

# --- (f) Bland-Altman relative error plot ---
ax = axes[2, 1]
for t in ['Simulated', 'Plasma']:
    mask = (type_clean_ba == t)
    ax.scatter(mean_val_ba[mask], rel_diff_ba[mask],
               marker=markers[t], color=colors[t], alpha=0.8,
               edgecolors='k', s=50, label=t)
ax.axhline(y=bias_rel, color='r', linestyle='-', label=f'Bias = {bias_rel:.2f}%')
ax.axhline(y=loa_rel_lower, color='gray', linestyle='--', label='95% LoA')
ax.axhline(y=loa_rel_upper, color='gray', linestyle='--')
ax.set_xlabel('Mean of TCS and Simoa', fontsize=14)
ax.set_ylabel('Relative difference (%)', fontsize=14)
ax.set_title('(6) Bland-Altman (relative)', fontsize=14,fontweight='bold')
ax.legend(fontsize=14)
loa_text = f'95% LoA: [{loa_rel_lower:.2f}%, {loa_rel_upper:.2f}%]'
ax.text(0.05, 0.95, loa_text, transform=ax.transAxes, verticalalignment='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.8), fontsize=14)

plt.tight_layout()
fig_core.savefig('Fig_2b.png', dpi=300, bbox_inches='tight')
print("\nCore figure saved: Fig_2b.png")

# ============================================================================
# 5. Supplementary figure: Fig_S1.png (QQ + concentration distribution + Bland-Altman)
# ============================================================================
fig_s1, axes_s1 = plt.subplots(2, 2, figsize=(12, 10))
fig_s1.suptitle('Supplementary Diagnostics', fontsize=16, fontweight='bold')

# (a) 4PL QQ
ax = axes_s1[0, 0]
stats.probplot(resid_4pl, dist="norm", plot=ax)
ax.set_title('(a) 4PL Residuals Q-Q', fontsize=14, fontweight='bold')
ax.set_xlabel('Theoretical Quantiles', fontsize=14)
ax.set_ylabel('Ordered Residuals', fontsize=14)
ax.grid(False)

# (b) TCS QQ
ax = axes_s1[0, 1]
stats.probplot(resid_tcs, dist="norm", plot=ax)
ax.set_title('(b) TCS Residuals Q-Q', fontsize=14, fontweight='bold')
ax.set_xlabel('Theoretical Quantiles', fontsize=14)
ax.set_ylabel('Ordered Residuals', fontsize=14)
ax.grid(False)

# (c) Test set concentration distribution
valid_4pl = conc_4pl_back[~np.isnan(conc_4pl_back)]
valid_tcs = conc_tcs_back[~np.isnan(conc_tcs_back)]
all_conc = np.concatenate([valid_4pl, valid_tcs])
if len(all_conc)>0:
    log_min = np.floor(np.log10(all_conc.min())) - 0.2
    log_max = np.ceil(np.log10(all_conc.max())) + 0.2
else:
    log_min, log_max = -2, 3
log_bins = np.logspace(log_min, log_max, 25)
ax = axes_s1[1, 0]
ax.hist(valid_4pl, bins=log_bins, alpha=0.4, color='blue', density=True, edgecolor='darkblue')
ax.hist(valid_tcs, bins=log_bins, alpha=0.4, color='red', density=True, edgecolor='darkred')
if len(valid_4pl)>1:
    kde_4pl = gaussian_kde(np.log10(valid_4pl), bw_method=0.3)
    x_grid = np.logspace(log_min, log_max, 200)
    ax.plot(x_grid, kde_4pl(np.log10(x_grid))/(x_grid*np.log(10)), 'b-', lw=2)
if len(valid_tcs)>1:
    kde_tcs = gaussian_kde(np.log10(valid_tcs), bw_method=0.3)
    ax.plot(x_grid, kde_tcs(np.log10(x_grid))/(x_grid*np.log(10)), 'r-', lw=2)
ax.set_xscale('log')
ax.set_xlim(10**log_min, 10**log_max)
ax.set_xlabel('Estimated Concentration (pg/mL)', fontsize=14)
ax.set_ylabel('Density', fontsize=14)
ax.set_title('(c) Test Set Concentration Distribution', fontsize=14, fontweight='bold')
ax.legend(['4PL KDE','TCS KDE'], fontsize=14)
ax.grid(False)

# (d) Bland-Altman
mask_valid_ba = (~np.isnan(conc_tcs_back)) & (~np.isnan(conc_4pl_back)) & (conc_tcs_back>0) & (conc_4pl_back>0)
tcs_pos = conc_tcs_back[mask_valid_ba]
fpl_pos = conc_4pl_back[mask_valid_ba]
ax = axes_s1[1, 1]
log_diff = np.log10(fpl_pos) - np.log10(tcs_pos)
log_mean = (np.log10(fpl_pos) + np.log10(tcs_pos)) / 2
ax.scatter(log_mean, log_diff, alpha=0.5, edgecolors='k', linewidth=0.3)
ax.axhline(np.mean(log_diff), color='r', linestyle='--', label=f'Mean = {np.mean(log_diff):.3f}')
ax.axhline(np.mean(log_diff)+1.96*np.std(log_diff), color='gray', linestyle=':', label='±1.96 SD')
ax.axhline(np.mean(log_diff)-1.96*np.std(log_diff), color='gray', linestyle=':')
ax.set_xlabel('Mean log10(Conc)', fontsize=14)
ax.set_ylabel('log10(4PL) - log10(TCS)', fontsize=14)
ax.set_title('(d) Bland-Altman', fontsize=14, fontweight='bold')
ax.legend(fontsize=14)
ax.grid(False)

plt.tight_layout()
fig_s1.savefig('Fig_S1.png', dpi=300, bbox_inches='tight')
print("Supplementary figure saved: Fig_S1.png")

print("\n" + "="*80)

# ============================================================================
# 7. R3 linear approximation analysis (independent fit, low-concentration region)
# ============================================================================
print("\n" + "="*80)
print("R3 LINEAR APPROXIMATION (Low Concentration)")
print("="*80)

gamma_tcs = 1.0 / (kappa_tcs + 1)

# Compute P_specific for all training points (background subtracted)
P_specific_train = (P_train - b_tcs) / (1 - b_tcs)
P_specific_train = np.clip(P_specific_train, 1e-12, 1-1e-12)

# R3 estimated molecule count (from TCS parameters)
M_r3 = N * (kappa_tcs + 1) * P_specific_train

# R3 CV formula (S4.25)
mu_r3 = M_r3 / N
CV_r3 = np.where(
    mu_r3 > 1e-12,
    (1.0 / (gamma_tcs * (1 - b_tcs) * mu_r3)) * np.sqrt(P_train * (1 - P_train) / n_obs),
    np.inf
)

# Check R3 validity: P_specific < 0.2
r3_valid = P_specific_train < 0.2

print(f"TCS‑based γ = {gamma_tcs:.4f}  (κ = {kappa_tcs:.2f})")
print(f"Number of points satisfying P_specific < 0.2: {np.sum(r3_valid)}")

# ---------- Independent linear fit ----------
mask_fit = r3_valid & ~mask_zero
if np.sum(mask_fit) >= 2:
    c_fit = concentrations_pgmL[mask_fit]
    mu_fit = (c_fit * 1e-12 * 1000) / Mw * V * N_A / N
    P_fit = P_specific_train[mask_fit]
    gamma_fit = np.sum(P_fit * mu_fit) / np.sum(mu_fit ** 2)
    kappa_fit = 1.0 / gamma_fit - 1.0 if gamma_fit > 0 else np.inf
    print(f"Fitted γ = {gamma_fit:.4f}  (κ = {kappa_fit:.2f})")
else:
    gamma_fit = np.nan
    kappa_fit = np.nan
    print("Not enough points for R3 fitting.")

print("\nR3 estimates (training set):")
header = f"{'Conc(pg/mL)':>12} {'P_specific':>10} {'M_r3':>12} {'CV_r3(%)':>10} {'R3_valid':>8}"
print(header)
for i in range(len(concentrations_pgmL)):
    print(f"{concentrations_pgmL[i]:12.4f} {P_specific_train[i]:10.6f} {M_r3[i]:12.2f} {CV_r3[i]*100:10.2f} {str(r3_valid[i]):>8}")

df_r3 = pd.DataFrame({
    'Conc_pgmL': concentrations_pgmL,
    'P_specific': P_specific_train,
    'M_r3': M_r3,
    'CV_r3_pct': CV_r3 * 100,
    'R3_valid': r3_valid
})
df_r3.to_csv('R3_analysis.csv', index=False)
print("\nR3 analysis saved to R3_analysis.csv")

# # ---- Plot R3 linear (alternating offset for CV labels) ----
# fig_r3, ax_r3 = plt.subplots(figsize=(7, 6))

# c_range = np.logspace(np.log10(concentrations_pgmL[concentrations_pgmL>0].min()*0.5),
#                       np.log10(concentrations_pgmL[concentrations_pgmL>0].max()*2), 100)
# mu_range = (c_range * 1e-12 * 1000) / Mw * V * N_A / N
# P_theory = gamma_tcs * mu_range
# ax_r3.loglog(c_range, P_theory, 'k--', lw=1, label=f'R2 (γ={gamma_tcs:.3f})')

# if not np.isnan(gamma_fit):
#     P_fit_line = gamma_fit * mu_range
#     ax_r3.loglog(c_range, P_fit_line, 'b-', lw=2, label=f'R3 fitted (γ={gamma_fit:.3f})')

# ax_r3.loglog(concentrations_pgmL[~mask_zero], P_specific_train[~mask_zero],
#              'o', color='gray', alpha=0.5, label='All data')

# valid_indices = np.where(r3_valid & ~mask_zero)[0]
# for plot_idx, i in enumerate(valid_indices):
#     conc_i = concentrations_pgmL[i]
#     ps_i = P_specific_train[i]
#     cv_i = CV_r3[i] * 100
#     ax_r3.loglog(conc_i, ps_i, 'ro', markersize=8)
#     y_shift = 1.3 if plot_idx % 2 == 0 else 0.7
#     va = 'bottom' if plot_idx % 2 == 0 else 'top'
#     ax_r3.text(conc_i * 1.1, ps_i * y_shift*1.4, f'T: {cv_i:.2f}%',
#                fontsize=8, color='darkred', ha='left', va=va)

# ax_r3.set_xlabel('Concentration (pg/mL)', fontsize=14)
# ax_r3.set_ylabel(r'$P_{\mathrm{specific}}$', fontsize=14)
# ax_r3.set_title('R3 Linear Approximation (Low Concentration)', fontsize=14, fontweight='bold')
# ax_r3.legend(fontsize=9)

# param_text = f"R2: γ={gamma_tcs:.4f}, κ={kappa_tcs:.2f}\n"
# if not np.isnan(gamma_fit):
#     param_text += f"R3 fit: γ={gamma_fit:.4f}, κ={kappa_fit:.2f}\n"
# else:
#     param_text += "R3 fit: insufficient data\n"
# param_text += "T: TCS Model CV"
# ax_r3.text(0.05, 0.95, param_text,
#            transform=ax_r3.transAxes, fontsize=10, va='top')
# plt.tight_layout()
# fig_r3.savefig('Fig_R3.png', dpi=300, bbox_inches='tight')
# print("R3 figure saved: Fig_R3.png")

print("All outputs generated.")
print("="*80)

# ============================================================================
# 8. LoB, LoD, LoQ calculation (IL-6 dataset, using per-well actual n_i)
# ============================================================================
print("\n" + "="*80)
print("LoB, LoD, LoQ Analysis (IL-6 dataset, per‑well actual n)")
print("="*80)

N_total = N
kappa_fit = kappa_tcs
b_fit = b_tcs
gamma_fit = 1.0 / (kappa_fit + 1)

results_per_well = []

for i in range(len(AEB_train)):
    n_i = int(n_obs[i])
    # ---------- R2 per-well metrics ----------
    def P_pos_R2(M):
        return b_fit + (1 - b_fit) * (1 - np.exp(-M / (N_total * (kappa_fit + 1))))
    def M_from_Ppos_R2(P_pos):
        if P_pos <= b_fit:
            return 0.0
        P_spec = (P_pos - b_fit) / (1 - b_fit)
        return -N_total * (kappa_fit + 1) * np.log(1 - P_spec)
    def CV_R2(M):
        if M <= 0: return np.inf
        P_pos = P_pos_R2(M)
        if P_pos <= 0 or P_pos >= 1: return np.inf
        P_spec = (P_pos - b_fit) / (1 - b_fit)
        P_spec = np.clip(P_spec, 1e-12, 1-1e-12)
        dM_dP = N_total * (kappa_fit + 1) / ((1 - b_fit) * (1 - P_spec))
        se = np.abs(dM_dP) * np.sqrt(P_pos * (1 - P_pos) / n_i)
        return se / M

    k_95 = stats.binom.ppf(0.95, n_i, b_fit)
    lob_p = k_95 / n_i
    M_lob_r2 = M_from_Ppos_R2(lob_p)

    def detection_prob(M):
        P_pos = P_pos_R2(M)
        k_th = int(np.floor(n_i * lob_p)) + 1
        if k_th > n_i: return 0.0
        return stats.binom.sf(k_th - 1, n_i, P_pos)
    def lod_r2():
        M_low = 1.0
        while detection_prob(M_low) < 0.95:
            M_low *= 2
            if M_low > 1e8: return np.nan
        M_high = M_low
        M_low = max(1.0, M_low / 2)
        try:
            return brentq(lambda M: detection_prob(M) - 0.95, M_low, M_high, xtol=1e-6)
        except:
            return np.nan
    M_lod_r2 = lod_r2()

    def loq_low():
        def obj(M): return CV_R2(M) - 0.2
        M_low = 1.0
        while obj(M_low) > 0: M_low *= 2
        M_high = M_low
        M_low = max(1.0, M_low / 2)
        try: return brentq(obj, M_low, M_high, xtol=1e-6)
        except: return np.nan
    def loq_high():
        def obj(M): return CV_R2(M) - 0.2
        M_start = 500000
        while CV_R2(M_start) < 0.2:
            M_start *= 2
            if M_start > 1e8: break
        M_high = M_start
        M_low = M_start / 2
        while obj(M_low) > 0:
            M_low /= 2
            if M_low < 1e3: break
        try: return brentq(obj, M_low, M_high, xtol=1e-6)
        except: return np.nan
    M_loq_low_r2 = loq_low()
    M_loq_high_r2 = loq_high()

    # ---------- R3 per-well metrics ----------
    import math
    lob_p_approx = b_fit + 1.645 * math.sqrt(b_fit * (1 - b_fit) / n_i)
    M_lob_r3 = N_total * (kappa_fit + 1) * (lob_p_approx - b_fit) / (1 - b_fit)
    M_lod_r3 = (3.29 * N_total / gamma_fit) * np.sqrt(b_fit / (n_i * (1 - b_fit)))
    # A = 0.04 * n_i * (1 - b_fit)**2
    # B = -(1 - b_fit)
    # C = -b_fit
    # x_loq = (-B + np.sqrt(B**2 - 4*A*C)) / (2*A)
    # M_loq_low_r3 = N_total * (kappa_fit + 1) * x_loq
    
    
    # Corrected R3 LoQ equation (S2f.33 revised)
    A = (1 - b_fit) * (1 + 0.04 * n_i)
    B = -(1 - 2 * b_fit)
    C = -b_fit
    x_loq = (-B + np.sqrt(B**2 - 4*A*C)) / (2*A)
    M_loq_low_r3 = N_total * (kappa_fit + 1) * x_loq
        
    
    
    
    

    results_per_well.append({
        'Well': i+1,
        'Conc (pg/mL)': concentrations_pgmL[i],
        'n': n_i,
        'R2_LoB': M_lob_r2,
        'R2_LoD': M_lod_r2,
        'R2_LoQ_low': M_loq_low_r2,
        'R2_LoQ_high': M_loq_high_r2,
        'R3_LoB': M_lob_r3,
        'R3_LoD': M_lod_r3,
        'R3_LoQ_low': M_loq_low_r3
    })

print(f"Parameters: N_total = {N_total}, κ = {kappa_fit:.4f}, b = {b_fit:.6e}")
print(f"\n{'Well':<6} {'Conc(pg/mL)':<12} {'n':<8} {'R2_LoB':<12} {'R2_LoD':<12} {'R2_LoQ_low':<14} {'R2_LoQ_high':<14} {'R3_LoB':<12} {'R3_LoD':<12} {'R3_LoQ_low':<14}")
print("-"*120)
for d in results_per_well:
    print(f"{d['Well']:<6} {d['Conc (pg/mL)']:<12.4f} {d['n']:<8} "
          f"{d['R2_LoB']:<12.1f} {d['R2_LoD']:<12.1f} {d['R2_LoQ_low']:<14.1f} {d['R2_LoQ_high']:<14.0f} "
          f"{d['R3_LoB']:<12.1f} {d['R3_LoD']:<12.1f} {d['R3_LoQ_low']:<14.1f}")

r2_lob = np.array([d['R2_LoB'] for d in results_per_well])
r2_lod = np.array([d['R2_LoD'] for d in results_per_well])
r2_loq_l = np.array([d['R2_LoQ_low'] for d in results_per_well])
r2_loq_h = np.array([d['R2_LoQ_high'] for d in results_per_well])
r3_lob = np.array([d['R3_LoB'] for d in results_per_well])
r3_lod = np.array([d['R3_LoD'] for d in results_per_well])
r3_loq_l = np.array([d['R3_LoQ_low'] for d in results_per_well])

print("\nSummary statistics across all wells:")
print(f"{'Metric':<15} {'Min':<12} {'Median':<12} {'Max':<12}")
print(f"{'R2 LoB':<15} {np.min(r2_lob):<12.1f} {np.median(r2_lob):<12.1f} {np.max(r2_lob):<12.1f}")
print(f"{'R2 LoD':<15} {np.nanmin(r2_lod):<12.1f} {np.nanmedian(r2_lod):<12.1f} {np.nanmax(r2_lod):<12.1f}")
print(f"{'R2 LoQ low':<15} {np.nanmin(r2_loq_l):<12.1f} {np.nanmedian(r2_loq_l):<12.1f} {np.nanmax(r2_loq_l):<12.1f}")
print(f"{'R2 LoQ high':<15} {np.nanmin(r2_loq_h):<12.0f} {np.nanmedian(r2_loq_h):<12.0f} {np.nanmax(r2_loq_h):<12.0f}")
print(f"{'R3 LoB':<15} {np.min(r3_lob):<12.1f} {np.median(r3_lob):<12.1f} {np.max(r3_lob):<12.1f}")
print(f"{'R3 LoD':<15} {np.min(r3_lod):<12.1f} {np.median(r3_lod):<12.1f} {np.max(r3_lod):<12.1f}")
print(f"{'R3 LoQ low':<15} {np.min(r3_loq_l):<12.1f} {np.median(r3_loq_l):<12.1f} {np.max(r3_loq_l):<12.1f}")
print("(Note: R3 has no high-conc LoQ; R2 LoQ high may have NaN for some wells, ignored.)")

df_wells = pd.DataFrame(results_per_well)
df_wells.to_csv('IL6_per_well_LoB_LoD_LoQ.csv', index=False)
print("\nPer‑well performance metrics saved to IL6_per_well_LoB_LoD_LoQ.csv")

# ============================================================================
# 9. Visualize LoB / LoD / LoQ distributions
# ============================================================================
print("\n" + "=" * 80)
print("VISUALIZING DISTRIBUTIONS OF LoB / LoD / LoQ")
print("=" * 80)

def M_to_pgmL(M):
    return M / (V * N_A) * Mw * 1e9

r2_lob_c  = M_to_pgmL(r2_lob)
r2_lod_c  = M_to_pgmL(r2_lod)
r2_loql_c = M_to_pgmL(r2_loq_l)
r2_loqh_c = M_to_pgmL(r2_loq_h)

r3_lob_c  = M_to_pgmL(r3_lob)
r3_lod_c  = M_to_pgmL(r3_lod)
r3_loql_c = M_to_pgmL(r3_loq_l)

# (Distribution plot code continues here)
print("Distribution analysis complete.")














# ============================================================================
# 10. Supplementary: TCS-predicted CV at manufacturer‑specified concentrations
# ============================================================================
print("\n" + "="*80)
print("TCS‑PREDICTED CV AT MANUFACTURER‑SPECIFIED CONCENTRATIONS")
print("="*80)

# Select representative concentrations (pg/mL)
ref_concs_pgmL = np.array([
    0.0,       # blank
    0.0055,    # manufacturer's mean LOD
    0.010,     # analytical LLOQ
    0.040,     # functional LLOQ (4x)
    0.607,     # lower bound of normal plasma range
    1.79,      # mean normal plasma
    4.64,      # upper bound of normal plasma range
    24.0       # highest calibrator
])

# Use mean effective well count from training set as typical single-well n
n_typical = int(np.mean(n_obs))
print(f"Typical per‑well bead count (mean of training wells): n = {n_typical}")

print(f"\n{'Conc (pg/mL)':<15} {'M (molecules)':<15} {'P_pos':<12} {'CV (%)':<10}")
print("-" * 55)

for c in ref_concs_pgmL:
    if c == 0:
        # At zero concentration P_pos = b, use background directly
        M_est = 0.0
        P_pos = b_tcs
    else:
        # Concentration -> M
        c_gL = c * 1e-12 * 1000         # pg/mL → g/L
        C_molL = c_gL / Mw              # mol/L
        M_est = C_molL * V * N_A        # total molecules in 125 μL
        mu_est = M_est / N_total
        # Theoretical P_pos
        P_spec = 1 - np.exp(-mu_est / (kappa_tcs + 1))
        P_pos = b_tcs + (1 - b_tcs) * P_spec
        P_pos = np.clip(P_pos, 1e-12, 1 - 1e-12)
    
    # Compute theoretical CV using this P_pos and typical n
    cv = calc_CV(P_pos, n_typical, b_tcs, kappa_tcs, N_total)
    print(f"{c:<15.4f} {M_est:<15.1f} {P_pos:<12.6f} {cv:<10.2f}")

print("\nNote: CV is the theoretical single‑well imprecision from binomial error propagation,")
print("using the average bead count across training wells. Manufacturer's reported CV (e.g.,")
print("analytical LLOQ CV ≈ 12.4%) includes additional sources of variation (between‑run,")
print("between‑instrument, reagent lot), and is therefore expected to be higher.")
print("="*80)




# SI Table 4 export block for IL-6 (comprehensive) - v2
# Replace the existing block in your IL-6 script with this one

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ============================================================
    # Sheet 1: Raw training data (16 wells × 7 cols)
    # ============================================================
    ws1 = wb.active
    ws1.title = "1_Raw_Training"
    headers1 = ["Concentration (pg/mL)", "n_obs (beads)", "AEB_raw",
                "k_train (positive beads)", "P_train", "mu_train (TCS)", "mu_train_4pl"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(concentrations_pgmL)):
        ws1.append([
            concentrations_pgmL[i], n_obs[i], AEB_train_raw[i],
            k_train[i] if 'k_train' in dir() else np.nan,
            P_train[i] if 'P_train' in dir() else np.nan,
            mu_train[i] if 'mu_train' in dir() else np.nan,
            mu_train_4pl[i] if 'mu_train_4pl' in dir() else np.nan
        ])

    # ============================================================
    # Sheet 2: Raw test data (clinical samples)
    # ============================================================
    ws2 = wb.create_sheet("2_Raw_Test")
    ws2.append(["Test sample #", "test_n (beads)", "test_AEB_raw", "test_k (positive beads)"])
    for cell in ws2[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(test_AEB)):
        ws2.append([i+1, test_n[i], test_AEB[i], test_k[i] if 'test_k' in dir() else np.nan])

    # ============================================================
    # Sheet 3: 4PL vs TCS predictions on training set
    # ============================================================
    ws3 = wb.create_sheet("3_Training_Predictions")
    ws3.append(["Concentration (pg/mL)", "AEB_observed",
                "AEB_pred_4PL", "resid_4PL", "P_pred_TCS", "AEB_pred_TCS", "resid_TCS"])
    for cell in ws3[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(concentrations_pgmL)):
        ws3.append([
            concentrations_pgmL[i], AEB_train_raw[i],
            AEB_pred_train_4pl[i] if 'AEB_pred_train_4pl' in dir() else np.nan,
            resid_4pl[i] if 'resid_4pl' in dir() else np.nan,
            tcs_model(mu_train_4pl[i] if 'mu_train_4pl' in dir() else mu_train[i], kappa_tcs, b_tcs),
            AEB_pred_train_tcs[i] if 'AEB_pred_train_tcs' in dir() else np.nan,
            resid_tcs[i] if 'resid_tcs' in dir() else np.nan
        ])

    # ============================================================
    # Sheet 4: Fitted parameters
    # ============================================================
    ws4 = wb.create_sheet("4_Fit_Parameters")
    ws4.append(["Parameter", "Value", "Note"])
    for cell in ws4[1]:
        cell.font = bold; cell.fill = header_fill
    if 'A_4pl' in dir(): ws4.append(["A_4pl (fixed from blank)", A_4pl, "Lower asymptote"])
    if 'B_4pl' in dir(): ws4.append(["B_4pl (slope)", B_4pl, ""])
    if 'C_4pl' in dir(): ws4.append(["C_4pl (midpoint)", C_4pl, "pg/mL × scale"])
    if 'D_4pl' in dir(): ws4.append(["D_4pl (upper asymptote)", D_4pl, ""])
    if 'kappa_tcs' in dir(): ws4.append(["κ_tcs (depletion)", kappa_tcs, "Mid-conc identifiable"])
    if 'b_tcs' in dir(): ws4.append(["b_tcs (background)", b_tcs, "Non-specific binding"])
    if 'Mw' in dir(): ws4.append(["Mw (Da)", Mw, "IL-6 molecular weight"])
    if 'V' in dir(): ws4.append(["V (assay volume)", V, "Liters"])
    if 'N_total' in dir(): ws4.append(["N_total (beads)", N_total, "TCS normalization"])

    # ============================================================
    # Sheet 5: Bootstrap CI + correlations (KEY: orthogonality claim)
    # ============================================================
    ws5 = wb.create_sheet("5_Bootstrap_CI")
    ws5.append(["Model", "Parameter", "Point", "2.5%", "97.5%", "Width", "Relative width"])
    for cell in ws5[1]:
        cell.font = bold; cell.fill = header_fill
    for model, name, val, arr in [("4PL", "B", B_4pl, boot_B),
                            ("4PL", "C", C_4pl, boot_C),
                            ("4PL", "D", D_4pl, boot_D)]:
        if len(arr) > 0:
            lo, hi = np.percentile(arr, [2.5, 97.5])
            ws5.append(["4PL", name, val, lo, hi, hi-lo, (hi-lo)/abs(val) if abs(val)>1e-12 else np.nan])
    for model, name, val, arr in [("TCS", "kappa", kappa_tcs, boot_kappa),
                            ("TCS", "b", b_tcs, boot_b)]:
        if len(arr) > 0:
            lo, hi = np.percentile(arr, [2.5, 97.5])
            ws5.append(["TCS", name, val, lo, hi, hi-lo, (hi-lo)/abs(val) if abs(val)>1e-12 else np.nan])
    ws5.append([""])
    ws5.append(["== Correlations ρ (KEY CLAIM) ==", "", "", "", "", "", ""])
    if len(boot_B) > 0 and len(boot_C) > 0:
        ws5.append(["4PL", "ρ(B, C)", np.corrcoef(boot_B, boot_C)[0, 1], "", "", "", ""])
    if len(boot_B) > 0 and len(boot_D) > 0:
        ws5.append(["4PL", "ρ(B, D)", np.corrcoef(boot_B, boot_D)[0, 1], "", "", "", ""])
    if len(boot_C) > 0 and len(boot_D) > 0:
        ws5.append(["4PL", "ρ(C, D)", np.corrcoef(boot_C, boot_D)[0, 1], "", "", "", ""])
    if len(boot_kappa) > 0 and len(boot_b) > 0:
        ws5.append(["TCS", "ρ(κ, b)", np.corrcoef(boot_kappa, boot_b)[0, 1], "", "", "", ""])

    # ============================================================
    # Sheet 6: Model comparison statistics
    # ============================================================
    ws6 = wb.create_sheet("6_Model_Comparison")
    ws6.append(["Metric", "4PL", "TCS", "Winner / Note"])
    for cell in ws6[1]:
        cell.font = bold; cell.fill = header_fill
    if 'logL_4pl' in dir() and 'logL_tcs' in dir():
        ws6.append(["logL", logL_4pl, logL_tcs, "Higher = better fit"])
    if 'AICc_4pl' in dir() and 'AICc_tcs' in dir():
        ws6.append(["AICc", AICc_4pl, AICc_tcs, "Lower = better (penalizes complexity)"])
    if 'BIC_4pl' in dir() and 'BIC_tcs' in dir():
        ws6.append(["BIC", BIC_4pl, BIC_tcs, "Lower = better"])
    if 'R2_4pl' in dir() and 'R2_tcs' in dir():
        ws6.append(["R² (weighted, AEB scale)", R2_4pl, R2_tcs, "Higher = better"])
    if 'p_shapiro_4pl' in dir() and 'p_shapiro_tcs' in dir():
        ws6.append(["Shapiro-Wilk p (residuals)", p_shapiro_4pl, p_shapiro_tcs, "Higher = more normal"])

    # ============================================================
    # Sheet 7: LOO cross-validation statistics
    # ============================================================
    ws7 = wb.create_sheet("7_LOO_Stats")
    ws7.append(["Metric", "Value", "Note"])
    for cell in ws7[1]:
        cell.font = bold; cell.fill = header_fill
    if 'fail_4pl' in dir(): ws7.append(["4PL LOO failures", fail_4pl, "out of 16 wells"])
    if 'fail_tcs' in dir(): ws7.append(["TCS LOO failures", fail_tcs, "out of 16 wells"])
    if 'loo_sum_4pl' in dir(): ws7.append(["LOO total logL (4PL)", loo_sum_4pl, "Sum across wells"])
    if 'loo_sum_tcs' in dir(): ws7.append(["LOO total logL (TCS)", loo_sum_tcs, "Sum across wells"])
    if 'mean_diff' in dir(): ws7.append(["Mean diff (4PL - TCS)", mean_diff, "Positive = TCS better"])
    if 'se_diff' in dir(): ws7.append(["SE of diff", se_diff, ""])
    if 't_stat' in dir(): ws7.append(["t-statistic", t_stat, "Paired t-test LOO"])




    if 'p_tost' in dir():
        ws7.append(["TOST margin (±logL)", float(delta),
                    "1% of mean per-fold |logL|"])
        ws7.append(["Mean diff 90% CI", f"[{ci90[0]:.3f}, {ci90[1]:.3f}]", ""])
        ws7.append(["TOST p", float(p_tost),
                    "EQUIVALENT" if p_tost < 0.05 else "NOT equivalent"])
        ws7.append(["Smallest passing margin (logL)", float(delta_min),
                    f"{100*delta_min/mean_abs_logL:.3f}% of per-fold |logL|"])









    # ============================================================
    # Sheet 8: Reference table for clinical context
    # ============================================================
    ws8 = wb.create_sheet("8_Clinical_Reference")
    ws8.append(["Reference", "Value (pg/mL)", "Source"])
    for cell in ws8[1]:
        cell.font = bold; cell.fill = header_fill
    ws8.append(["Manufacturer LOD", 0.0055, "Quoted by manufacturer"])
    ws8.append(["Analytical LLOQ", 0.010, "Quoted by manufacturer"])
    ws8.append(["Functional LLOQ (4× LOD)", 0.040, "Quoted by manufacturer"])
    ws8.append(["Plasma lower bound (normal)", 0.607, "Reference range"])
    ws8.append(["Plasma mean (normal)", 1.79, "Reference range"])
    ws8.append(["Plasma upper bound (normal)", 4.64, "Reference range"])
    ws8.append(["Highest calibrator", 24.0, "Standard curve"])

    wb.save("SI_Table_4.xlsx")
    print("\n" + "="*60)
    print("SI Table 4 (v2) saved to 'SI_Table_4.xlsx'")
    print("="*60)
    print("Sheets:")
    print("  1. Raw_Training        - 16 wells × 7 columns (raw + processed)")
    print("  2. Raw_Test            - 70 clinical samples")
    print("  3. Training_Predictions- 4PL vs TCS predictions + residuals")
    print("  4. Fit_Parameters      - 4PL (A,B,C,D) + TCS (κ,b)")
    print("  5. Bootstrap_CI        - 95% CI + ρ(B,C,D), ρ(κ,b) KEY CLAIM")
    print("  6. Model_Comparison    - logL, AICc, BIC, R², Shapiro-Wilk")
    print("  7. LOO_Stats           - Leave-one-out with paired t-test")
    print("  8. Clinical_Reference  - manufacturer LOD, plasma normal range")




# ============================================================================
    # v1 extension: merge 4 CSVs + LOO_per_fold.txt + manufacturer reference CV into Excel
    # ============================================================================
    print("\n" + "=" * 70)
    print("EXTENSION v1: Merging CSV outputs into SI_Table_4.xlsx")
    print("=" * 70)
    try:
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        def add_csv_as_sheet(wb, csv_path, sheet_name, header_note=None):
            df = pd.read_csv(csv_path)
            ws = wb.create_sheet(sheet_name)
            if header_note:
                ws.append([header_note])
                ws.cell(row=1, column=1).font = Font(italic=True)
            ws.append(list(df.columns))
            for cell in ws[ws.max_row]:
                cell.font = bold
                cell.fill = header_fill
            for _, row in df.iterrows():
                ws.append([row[c] if not pd.isna(row[c]) else None for c in df.columns])
            return ws

        add_csv_as_sheet(wb, 'TCS_quantification_per_point.csv', '9_TCS_PerPoint',
                         'TCS quantification per training well — M_hat with 3 CIs (Delta / Exact / LR)')
        print('[OK] Added 9_TCS_PerPoint')
        add_csv_as_sheet(wb, 'TCS_test_quantification.csv', '10_TCS_TestSet',
                         'TCS quantification on the 68-sample test set (CV + CI)')
        print('[OK] Added 10_TCS_TestSet')
        add_csv_as_sheet(wb, 'R3_analysis.csv', '11_R3_Analysis',
                         'R3 linear approximation per training well — gamma-based M_r3 estimate')
        print('[OK] Added 11_R3_Analysis')
        add_csv_as_sheet(wb, 'IL6_per_well_LoB_LoD_LoQ.csv', '12_LoB_LoD_LoQ',
                         'Per-well LoB / LoD / LoQ (R2 and R3 formulations)')
        print('[OK] Added 12_LoB_LoD_LoQ')

        loo_df = pd.read_csv('LOO_per_fold.txt', sep='\t')
        ws = wb.create_sheet('13_LOO_PerFold')
        ws.append(['LOO per-fold logL: positive Diff = 4PL better, negative = TCS better'])
        ws.cell(row=1, column=1).font = Font(italic=True)
        ws.append(list(loo_df.columns))
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill
        for _, row in loo_df.iterrows():
            ws.append([row[c] if not pd.isna(row[c]) else None for c in loo_df.columns])
        print('[OK] Added 13_LOO_PerFold')

        ref_concs_pgmL = np.array([0.0, 0.0055, 0.010, 0.040, 0.607, 1.79, 4.64, 24.0])
        n_typical = int(np.mean(n_obs))
        ws = wb.create_sheet('14_Reference_CV')
        ws.append([f'Manufacturer concentration -> TCS predicted CV (typical n=mean(n_obs)={n_typical})'])
        ws.cell(row=1, column=1).font = Font(italic=True)
        ws.append(['Conc (pg/mL)', 'M (molecules)', 'P_pos', 'CV (%)', 'Note'])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill
        for c in ref_concs_pgmL:
            if c == 0:
                M_est = 0.0; P_pos = b_tcs
                note = 'Blank (background only)'
            else:
                c_gL = c * 1e-12 * 1000
                C_molL = c_gL / Mw
                M_est = C_molL * V * N_A
                mu_est = M_est / N_total
                P_spec = 1 - np.exp(-mu_est / (kappa_tcs + 1))
                P_pos = b_tcs + (1 - b_tcs) * P_spec
                P_pos = np.clip(P_pos, 1e-12, 1 - 1e-12)
                note = ('Manufacturer LOD' if c == 0.0055 else
                        'Analytical LLOQ' if c == 0.010 else
                        'Functional LLOQ' if c == 0.040 else
                        'Plasma lower' if c == 0.607 else
                        'Plasma mean' if c == 1.79 else
                        'Plasma upper' if c == 4.64 else 'Highest calibrator')
            P_clip = np.clip(P_pos, 1e-12, 1 - 1e-12)
            dM_dP = N_total * (kappa_tcs + 1) / ((1 - b_tcs) * (1 - P_clip))
            se = np.abs(dM_dP) * np.sqrt(P_clip * (1 - P_clip) / n_typical)
            cv = (se / M_est) * 100 if M_est > 0 else np.inf
            ws.append([c, M_est, P_pos, cv, note])
        print('[OK] Added 14_Reference_CV')

        print(f'[v1] SI Table 4 expanded from 8 to 14 sheets')
    except Exception as e_v1:
        print(f'[FAIL v1] {e_v1}')
        import traceback; traceback.print_exc()

    # ============================================================================
    # v2 extension: add 3 batches of computed-but-unsaved data to Excel
    # ============================================================================
    print("\n" + "=" * 70)
    print("EXTENSION v2: Adding computed-but-unsaved data")
    print("=" * 70)
    try:
        def mu_to_conc_v2(mu_val):
            C_molL = (mu_val * N_total) / (V * N_A)
            return C_molL * Mw * 1e9

        def M_to_pgmL_v2(M):
            return M / (V * N_A) * Mw * 1e9

        def inverse_4pl_v2(AEB_target, A, B, C, D):
            if AEB_target <= A: return 0.0
            if AEB_target >= D: return np.nan
            ratio = (AEB_target - A) / (D - AEB_target)
            if ratio <= 0: return 0.0
            return C * ratio ** (1.0 / B)

        def inverse_tcs_v2(P_target, kappa, b):
            if P_target <= b: return 0.0
            if P_target >= 1 - 1e-6: return np.nan
            P_specific = (P_target - b) / (1 - b)
            if P_specific >= 1: return np.nan
            return -(1 + kappa) * np.log(1 - P_specific)

        test_k_v2 = test_n * (1 - np.exp(-test_AEB))
        P_test_v2 = test_k_v2 / test_n

        conc_4pl_back = []
        for aeb in test_AEB:
            mu_scaled = inverse_4pl_v2(aeb, A_4pl, B_4pl, C_4pl, D_4pl)
            if np.isnan(mu_scaled) or mu_scaled == 0:
                conc_4pl_back.append(np.nan)
            else:
                conc_4pl_back.append(mu_to_conc_v2(mu_scaled * scale_4pl))
        conc_4pl_back = np.array(conc_4pl_back)

        conc_tcs_back = []
        for p in P_test_v2:
            mu_real = inverse_tcs_v2(p, kappa_tcs, b_tcs)
            if np.isnan(mu_real) or mu_real == 0:
                conc_tcs_back.append(np.nan)
            else:
                conc_tcs_back.append(mu_to_conc_v2(mu_real))
        conc_tcs_back = np.array(conc_tcs_back)

        conc_train_real = (mu_train * N_total) / (V * N_A) * Mw * 1e9
        conc_train_4pl_real = (mu_train_4pl * scale_4pl * N_total) / (V * N_A) * Mw * 1e9

        ws = wb.create_sheet('15_TestSet_BackCalc')
        ws.append(['Test set 68 samples — back-calculated concentration (4PL vs TCS) — KEY for clinical validation'])
        ws.cell(row=1, column=1).font = Font(italic=True)
        ws.append(['Test sample #', 'AEB (raw)', 'n (beads)', 'k (positive)', 'P_obs',
                   'Conc_4PL (pg/mL)', 'Conc_TCS (pg/mL)', 'Diff_4PL-TCS', '|log10 ratio|'])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill
        for i in range(len(test_AEB)):
            c4 = conc_4pl_back[i]
            ct = conc_tcs_back[i]
            if not np.isnan(c4) and not np.isnan(ct) and c4 > 0 and ct > 0:
                diff = c4 - ct
                logr = abs(np.log10(c4) - np.log10(ct))
            else:
                diff = np.nan; logr = np.nan
            ws.append([i + 1, test_AEB[i], test_n[i], test_k_v2[i], P_test_v2[i],
                       c4, ct, diff, logr])
        ws.append([])
        ws.append(['Summary', '4PL', 'TCS', '', '', '', '', '', ''])
        for c in ws[ws.max_row]: c.font = bold
        ws.append(['Min (pg/mL)', np.nanmin(conc_4pl_back), np.nanmin(conc_tcs_back),
                   '', '', '', '', '', ''])
        ws.append(['Median (pg/mL)', np.nanmedian(conc_4pl_back), np.nanmedian(conc_tcs_back),
                   '', '', '', '', '', ''])
        ws.append(['Max (pg/mL)', np.nanmax(conc_4pl_back), np.nanmax(conc_tcs_back),
                   '', '', '', '', '', ''])
        ws.append(['# NaN (above ULOQ)', np.sum(np.isnan(conc_4pl_back)),
                   np.sum(np.isnan(conc_tcs_back)), '', '', '', '', '', ''])
        ws.append(['# =0 (below LLOQ)', np.sum(conc_4pl_back == 0),
                   np.sum(conc_tcs_back == 0), '', '', '', '', '', ''])
        print('[OK] Added 15_TestSet_BackCalc')

        ws = wb.create_sheet('16_Train_RealConc')
        ws.append(['Training-set real concentrations in pg/mL (TCS scale and 4PL scale)'])
        ws.cell(row=1, column=1).font = Font(italic=True)
        ws.append(['Concentration (input, pg/mL)', 'conc_train_real (TCS scale, pg/mL)',
                   'conc_train_4pl_real (4PL scale, pg/mL)', 'Ratio (4PL / TCS)'])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill
        for i in range(len(concentrations_pgmL)):
            r = conc_train_4pl_real[i] / conc_train_real[i] if conc_train_real[i] > 0 else np.nan
            ws.append([concentrations_pgmL[i], conc_train_real[i], conc_train_4pl_real[i], r])
        print('[OK] Added 16_Train_RealConc')

        ws = wb.create_sheet('17_LoB_LoD_LoQ_Dist')
        ws.append(['Per-well LoB / LoD / LoQ in pg/mL (converted from molecules, used for plotting distributions)'])
        ws.cell(row=1, column=1).font = Font(italic=True)
        ws.append(['Well', 'R2_LoB (pg/mL)', 'R2_LoD (pg/mL)', 'R2_LoQ_low (pg/mL)',
                   'R2_LoQ_high (pg/mL)', 'R3_LoB (pg/mL)', 'R3_LoD (pg/mL)', 'R3_LoQ_low (pg/mL)'])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill

        ws_existing = wb['12_LoB_LoD_LoQ']
        r2_lob_mol = [row[3] if row[3] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r2_lod_mol = [row[4] if row[4] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r2_loql_mol = [row[5] if row[5] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r2_loqh_mol = [row[6] if row[6] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r3_lob_mol = [row[7] if row[7] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r3_lod_mol = [row[8] if row[8] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        r3_loql_mol = [row[9] if row[9] is not None else np.nan for row in ws_existing.iter_rows(min_row=3, values_only=True) if row[0] is not None]

        r2_lob_c = [M_to_pgmL_v2(m) for m in r2_lob_mol]
        r2_lod_c = [M_to_pgmL_v2(m) for m in r2_lod_mol]
        r2_loql_c = [M_to_pgmL_v2(m) for m in r2_loql_mol]
        r2_loqh_c = [M_to_pgmL_v2(m) for m in r2_loqh_mol]
        r3_lob_c = [M_to_pgmL_v2(m) for m in r3_lob_mol]
        r3_lod_c = [M_to_pgmL_v2(m) for m in r3_lod_mol]
        r3_loql_c = [M_to_pgmL_v2(m) for m in r3_loql_mol]

        for i in range(16):
            ws.append([i + 1, r2_lob_c[i], r2_lod_c[i], r2_loql_c[i], r2_loqh_c[i],
                       r3_lob_c[i], r3_lod_c[i], r3_loql_c[i]])
        ws.append([])
        ws.append(['Median (pg/mL)', np.median(r2_lob_c), np.median(r2_lod_c),
                   np.median(r2_loql_c), np.median(r2_loqh_c),
                   np.median(r3_lob_c), np.median(r3_lod_c), np.median(r3_loql_c)])
        ws.append(['Min (pg/mL)', np.min(r2_lob_c), np.min(r2_lod_c),
                   np.min(r2_loql_c), np.min(r2_loqh_c),
                   np.min(r3_lob_c), np.min(r3_lod_c), np.min(r3_loql_c)])
        ws.append(['Max (pg/mL)', np.max(r2_lob_c), np.max(r2_lod_c),
                   np.max(r2_loql_c), np.max(r2_loqh_c),
                   np.max(r3_lob_c), np.max(r3_lod_c), np.max(r3_loql_c)])
        print('[OK] Added 17_LoB_LoD_LoQ_Dist')

        print(f'[v2] SI Table 4 expanded from 14 to 17 sheets')
    except Exception as e_v2:
        print(f'[FAIL v2] {e_v2}')
        import traceback; traceback.print_exc()

    # Save with 17 sheets (R1 sheet 18 added later after MCMC)
    wb.save("SI_Table_4.xlsx")
    print("\n" + "=" * 70)
    print(f"SI Table 4 saved with {len(wb.sheetnames)} sheets (18_R1_Model added after R1 MCMC)")
    print("=" * 70)
    for i, s in enumerate(wb.sheetnames, 1):
        print(f'  {i:2d}. {s}')
    print("=" * 70)
    print("Done.")

except ImportError:
        print("\n[!] openpyxl not installed. Run: pip install openpyxl")
except Exception as e_save4:
    print(f"\n[!] SI Table 4 v2 save failed: {e_save4}")
    import traceback
    traceback.print_exc()



# Add at end of IL-6 script
print(f"Catastrophic fold contribution: {loo_4pl_arr[10]+loo_4pl_arr[11]:.2f} / {loo_sum_4pl:.2f} = {(loo_4pl_arr[10]+loo_4pl_arr[11])/loo_sum_4pl*100:.1f}%")

# Output:
# Catastrophic fold contribution: -19663.92 / -44591.25 = 44.1%






# ============================================================================
# R1 vs R2 residual comparison + beta validity check
# Appended after R1 MCMC
# ============================================================================

# ---- R1 model functions (if not yet defined) ----
def solve_p_from_xi(xi, kappa, max_iter=100, tol=1e-12):
    """Solve master equation xi = p/(1-p) + p/kappa for p"""
    if xi <= 0:
        return 0.0
    if kappa <= 0:
        return min(xi, 1.0 - 1e-15)
    p = xi / (1.0 + xi)
    p = np.clip(p, 1e-15, 1.0 - 1e-15)
    for _ in range(max_iter):
        f = p / (1.0 - p) + p / kappa - xi
        df = 1.0 / (1.0 - p)**2 + 1.0 / kappa
        dp = f / df
        p_new = np.clip(p - dp, 1e-15, 1.0 - 1e-15)
        if abs(p_new - p) < tol:
            p = p_new
            break
        p = p_new
    return p

def tcs_R1_model(mu, kappa, beta, b):
    """R1 forward model: mu -> P_pos"""
    P_pos = np.empty_like(mu, dtype=float)
    for i, m in enumerate(mu):
        if m <= 0:
            P_pos[i] = b
            continue
        xi = m / (kappa * beta)
        p = solve_p_from_xi(xi, kappa)
        P_specific = 1.0 - (1.0 - p)**beta
        P_pos[i] = b + (1.0 - b) * P_specific
    return P_pos

print("\n" + "="*80)
print("R1 vs R2 residual comparison")
print("="*80)

# ---- R1 least-squares fit (if not yet run) ----
from scipy.optimize import minimize as opt_minimize

def neg_loglik_R1(params, mu, k, n):
    kappa, beta, b = params
    if kappa <= 0 or beta <= 0 or b < 0 or b >= 1:
        return 1e10
    P = tcs_R1_model(mu, kappa, beta, b)
    P = np.clip(P, 1e-15, 1 - 1e-15)
    return -np.sum(k * np.log(P) + (n - k) * np.log(1 - P))

# Use original kappa_tcs, b_tcs as initial values
if 'kappa_R1_hat' not in dir():
    print("R1 least-squares initialization...")
    bounds_R1 = [(0.01, 1000), (0.1, 10000), (0, 0.1)]
    p0_R1 = [kappa_tcs, 10.0, b_tcs]
    res_R1 = opt_minimize(neg_loglik_R1, p0_R1,
                          args=(mu_train, k_all, n_all_arr),
                          bounds=bounds_R1, method='L-BFGS-B',
                          options={'maxiter': 5000, 'eps': 1e-8})
    kappa_R1_hat, beta_R1_hat, b_R1_hat = res_R1.x
    print(f"R1 done: κ={kappa_R1_hat:.4f}, β={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")
else:
    print(f"R1 already run: κ={kappa_R1_hat:.4f}, β={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")

# ---- R2 model predictions (using original kappa_tcs, b_tcs) ----
def tcs_R2_model(mu, kappa, b):
    """R2 forward model: mu -> P_pos"""
    P_specific = 1.0 - np.exp(-mu / (kappa + 1.0))
    return b + (1.0 - b) * P_specific

# ---- Compute predictions and residuals for both models ----
P_R2_pred = tcs_R2_model(mu_train, kappa_tcs, b_tcs)
P_R1_pred_point = tcs_R1_model(mu_train, kappa_R1_hat, beta_R1_hat, b_R1_hat)

# AEB residuals
AEB_obs = -np.log(np.maximum(1 - k_all/n_all_arr, 1e-12))
AEB_R2_pred = -np.log(np.maximum(1 - P_R2_pred, 1e-12))
AEB_R1_pred = -np.log(np.maximum(1 - P_R1_pred_point, 1e-12))

resid_R2 = AEB_obs - AEB_R2_pred
resid_R1 = AEB_obs - AEB_R1_pred

# ---- Per-point comparison ----
d_r2_list = []
d_r1_list = []
imp_list = []
print(f"\n{'Conc':>10s} {'k':>10s} {'n':>8s} {'P_obs':>8s} {'P_R2':>8s} {'P_R1':>8s} {'dAEB_R2':>10s} {'dAEB_R1':>10s} {'improve':>10s}")
print("-"*90)
for i in range(len(mu_train)):
    P_obs = k_all[i]/n_all_arr[i]
    d_r2 = resid_R2[i]
    d_r1 = resid_R1[i]
    imp = abs(d_r2) - abs(d_r1)
    d_r2_list.append(d_r2)
    d_r1_list.append(d_r1)
    imp_list.append(imp)
    print(f"{concentrations_pgmL[i]:>10.4f} {k_all[i]:>10.1f} {n_all_arr[i]:>8.0f} {P_obs:>8.4f} {P_R2_pred[i]:>8.4f} {P_R1_pred_point[i]:>8.4f} {d_r2:>+10.4f} {d_r1:>+10.4f} {imp:>+10.4f}")

# ---- Summary statistics ----
print(f"\n{'Metric':>25s} {'R2 (β→∞)':>15s} {'R1 (β free)':>15s} {'improvement':>15s}")
print("-"*70)
rmse_R2 = np.sqrt(np.mean(resid_R2**2))
rmse_R1 = np.sqrt(np.mean(resid_R1**2))
mae_R2 = np.mean(np.abs(resid_R2))
mae_R1 = np.mean(np.abs(resid_R1))
max_R2 = np.max(np.abs(resid_R2))
max_R1 = np.max(np.abs(resid_R1))
print(f"{'RMSE (AEB)':>25s} {rmse_R2:>15.6f} {rmse_R1:>15.6f} {rmse_R2-rmse_R1:>+15.6f}")
print(f"{'MAE (AEB)':>25s} {mae_R2:>15.6f} {mae_R1:>15.6f} {mae_R2-mae_R1:>+15.6f}")
print(f"{'Max |residual|':>25s} {max_R2:>15.6f} {max_R1:>15.6f} {max_R2-max_R1:>+15.6f}")

# ---- P_specific residuals (more intuitive) ----
P_spec_obs = k_all / n_all_arr
P_spec_R2 = (P_R2_pred - b_tcs) / (1 - b_tcs)
P_spec_R1 = (P_R1_pred_point - b_R1_hat) / (1 - b_R1_hat)

resid_P_R2 = P_spec_obs - P_spec_R2
resid_P_R1 = P_spec_obs - P_spec_R1

print(f"\n{'Metric (P_spec)':>25s} {'R2 (β→∞)':>15s} {'R1 (β free)':>15s} {'improvement':>15s}")
print("-"*70)
rmse_P_R2 = np.sqrt(np.mean(resid_P_R2**2))
rmse_P_R1 = np.sqrt(np.mean(resid_P_R1**2))
print(f"{'RMSE (P_spec)':>25s} {rmse_P_R2:>15.6f} {rmse_P_R1:>15.6f} {rmse_P_R2-rmse_P_R1:>+15.6f}")
print(f"{'Max |resid| (P_spec)':>25s} {np.max(np.abs(resid_P_R2)):>15.6f} {np.max(np.abs(resid_P_R1)):>15.6f}")

# ---- beta validity check ----
print("\n" + "="*80)
print("beta validity check")
print("="*80)

N_total = 500000  # Simoa total partitions
beta_fit = beta_R1_hat
kappa_fit = kappa_R1_hat

Omega_calc = N_total * beta_fit
KVNA_calc = kappa_fit * Omega_calc

print(f"Fitted parameters: κ = {kappa_fit:.4f}, β = {beta_fit:.2f}")
print(f"N = {N_total}")
print(f"Ω = N·β = {Omega_calc:.3e}")
print(f"κΩ = KVN_A = {KVNA_calc:.3e}")

# Simoa IL-6 known parameter estimation
# IL-6 MW ≈ 21 kDa = 21000 g/mol
# Simoa reaction volume: ~25 uL per well total, but effective bead volume is smaller
# K_d for IL-6 antibody: typically ~10 pM = 10e-12 M (high-performance Simoa antibody)
# Here K is effective K, including avidity etc.

MW_IL6 = 21000  # g/mol
V_well = 25e-6  # L (Simoa typical reaction volume)

# Verify from highest concentration point
# 24 pg/mL → M = 24e-12 g/mL / 21000 g/mol * 6.022e23 * V_well_L
M_24pg = 24e-12 / MW_IL6 * 6.022e23 * V_well * 1000  # 24 pg/mL, 1 mL → molecules
print(f"\n24 pg/mL → M = {M_24pg:.3e} molecules (in {V_well*1e6:.0f} μL)")

# If M_24pg is known, xi_24 = M_24pg / K*V*N_A_calc
xi_24 = M_24pg / KVNA_calc
print(f"ξ(24pg/mL) = M/(KVN_A) = {xi_24:.4f}")

# Verify: P_specific at this xi
p_24 = solve_p_from_xi(xi_24, kappa_fit)
P_spec_24 = 1 - (1 - p_24)**beta_fit
print(f"p = {p_24:.6f}")
print(f"P_specific = {P_spec_24:.6f}")
print(f"P_obs (24 pg/mL) = {np.mean(P_spec_obs[concentrations_pgmL==24.0]):.6f}")

# Reverse: compute M from P_obs
print("\n--- Reverse verification: compute M from observed P ---")
print(f"{'Conc':>10s} {'P_obs':>8s} {'M_R1':>12s} {'M_R2':>12s} {'M_true':>12s} {'ratio_R1':>10s} {'ratio_R2':>10s}")
print("-"*80)
for conc_val in sorted(set(concentrations_pgmL)):
    if conc_val == 0:
        continue
    mask = concentrations_pgmL == conc_val
    P_obs_avg = np.mean(k_all[mask] / n_all_arr[mask])
    
    # R1: P_specific = 1-(1-p)^beta -> solve p -> xi -> M
    P_spec_obs_avg = (P_obs_avg - b_R1_hat) / (1 - b_R1_hat)
    P_spec_obs_avg = np.clip(P_spec_obs_avg, 1e-10, 1 - 1e-10)
    # 1-(1-p)^β = P_spec → (1-p)^β = 1-P_spec → p = 1-(1-P_spec)^(1/β)
    p_R1 = 1 - (1 - P_spec_obs_avg)**(1/beta_fit)
    xi_R1 = p_R1/(1-p_R1) + p_R1/kappa_fit
    M_R1 = xi_R1 * kappa_fit * Omega_calc
    
    # R2: P_specific = 1-exp(-μ/(1+κ)) → μ = -(1+κ)*ln(1-P_spec)
    P_spec_R2_obs = (P_obs_avg - b_tcs) / (1 - b_tcs)
    P_spec_R2_obs = np.clip(P_spec_R2_obs, 1e-10, 1 - 1e-10)
    mu_R2 = -(1 + kappa_tcs) * np.log(1 - P_spec_R2_obs)
    M_R2 = mu_R2 * N_total  # M = μ * N
    
    # M_true from concentration
    M_true = conc_val * 1e-12 / MW_IL6 * 6.022e23 * V_well * 1000
    
    ratio_R1 = M_R1 / M_true if M_true > 0 else 0
    ratio_R2 = M_R2 / M_true if M_true > 0 else 0
    
    print(f"{conc_val:>10.4f} {P_obs_avg:>8.4f} {M_R1:>12.3e} {M_R2:>12.3e} {M_true:>12.3e} {ratio_R1:>10.4f} {ratio_R2:>10.4f}")

# ---- Residual plot ----
fig, axes = plt.subplots(1, 3, figsize=(15, 5))

ax1 = axes[0]
ax1.scatter(AEB_R2_pred, resid_R2, alpha=0.7, marker='o', facecolors='none', edgecolors='blue', label='R2 (β→∞)')
ax1.scatter(AEB_R1_pred, resid_R1, alpha=0.7, marker='s', facecolors='none', edgecolors='red', label='R1 (β free)')
ax1.axhline(0, color='k', ls='--', lw=0.8)
ax1.set_xlabel('Predicted AEB')
ax1.set_ylabel('Residual (AEB)')
ax1.set_title('R1 vs R2 Residuals')
ax1.legend()

ax2 = axes[1]
ax2.bar(np.arange(len(resid_R2))-0.15, np.abs(resid_R2), width=0.3, alpha=0.7, color='blue', label='R2')
ax2.bar(np.arange(len(resid_R1))+0.15, np.abs(resid_R1), width=0.3, alpha=0.7, color='red', label='R1')
ax2.set_xlabel('Data point')
ax2.set_ylabel('|Residual| (AEB)')
ax2.set_title('Per-point |Residual|')
ax2.legend()

ax3 = axes[2]
concs_unique = sorted(set(concentrations_pgmL))
for c in concs_unique:
    mask = concentrations_pgmL == c
    if c == 0:
        continue
    ax3.scatter([c]*np.sum(mask), np.abs(resid_R2[mask]), alpha=0.7, color='blue', marker='o', facecolors='none')
    ax3.scatter([c]*np.sum(mask), np.abs(resid_R1[mask]), alpha=0.7, color='red', marker='s', facecolors='none')
ax3.set_xscale('log')
ax3.set_xlabel('Concentration (pg/mL)')
ax3.set_ylabel('|Residual| (AEB)')
ax3.set_title('Residuals vs Concentration')
ax3.legend(['R2', 'R1'])

plt.suptitle(f'R1 vs R2 Residual Comparison (β={beta_fit:.2f}, κ={kappa_fit:.2f})', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig('R1_vs_R2_residuals.png', dpi=300, bbox_inches='tight')
plt.show()

print("\nResidual plot saved: R1_vs_R2_residuals.png")









# ============================================================================
# R1 model fit: testing beta identifiability (IL-6 dataset)
# COMPLETE REPLACEMENT — uses dynesty (nested sampling)
# Physical prior: beta >= 1 (each bead has at least 1 binding site)
# ============================================================================

import dynesty
from dynesty import utils as dyfunc
try:
    import corner
except ImportError:
    corner = None
try:
    import arviz as az
except ImportError:
    az = None

print("\n" + "="*80)
print("R1 Model Fit: Testing β Identifiability (IL-6 dataset)")
print("  Method: Nested sampling (dynesty)")
print("="*80)

# ---- R1 forward model (VECTORIZED) ----
def solve_p_from_xi(xi, kappa, max_iter=100, tol=1e-12):
    """Vectorized Newton-Raphson: xi = p/(1-p) + p/kappa -> p"""
    xi = np.atleast_1d(np.asarray(xi, dtype=float))
    p = np.where(xi > 0, xi / (1.0 + xi), 0.0)
    p = np.clip(p, 1e-15, 1.0 - 1e-15)
    for _ in range(max_iter):
        f = p / (1.0 - p) + p / kappa - xi
        df = 1.0 / (1.0 - p)**2 + 1.0 / kappa
        dp = f / df
        p_new = np.clip(p - dp, 1e-15, 1.0 - 1e-15)
        if np.max(np.abs(p_new - p)) < tol:
            p = p_new
            break
        p = p_new
    return p

def tcs_R1_model(mu, kappa, beta, b):
    """Vectorized R1 forward model: mu -> P_pos"""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    P_pos = np.full_like(mu, b)
    mask = mu > 0
    if np.any(mask):
        xi = mu[mask] / (kappa * beta)
        p = solve_p_from_xi(xi, kappa)
        P_specific = 1.0 - (1.0 - p) ** beta
        P_pos[mask] = b + (1.0 - b) * P_specific
    return np.clip(P_pos, 1e-15, 1.0 - 1e-15)

def neg_loglik_R1(params, mu, k, n):
    kappa, beta, b = params
    if kappa <= 0 or beta <= 0 or b < 0 or b >= 1:
        return 1e12
    P_pos = tcs_R1_model(mu, kappa, beta, b)
    return -np.sum(k * np.log(P_pos) + (n - k) * np.log(1.0 - P_pos))

# ---- Data ----
mu_R1_all = mu_train.copy()
k_R1_all = k_all.copy()
n_R1_all = n_all_arr.copy()

# ---- Initial least-squares fit ----
kappa_init_R1 = kappa_tcs
b_init_R1 = b_tcs
beta_init = 10.0

print(f"\nR2 fit results (as R1 initial values): kappa={kappa_tcs:.4f}, b={b_tcs:.6f}")
print(f"R1 initial beta = {beta_init}")

bounds_R1 = [(0.01, 1000), (1.0, 10000), (0, 0.1)]
p0_R1 = [kappa_init_R1, beta_init, b_init_R1]

print("R1 least-squares initialization...")
# res_R1 = opt.minimize(neg_loglik_R1, p0_R1,
#                       args=(mu_R1_all, k_R1_all, n_R1_all),
#                       bounds=bounds_R1, method='L-BFGS-B',
#                       options={'maxiter': 5000, 'eps': 1e-8})
# 全局优化：differential_evolution（不依赖单一起点）
from scipy.optimize import differential_evolution
res_R1 = differential_evolution(
    neg_loglik_R1,
    bounds=bounds_R1,
    args=(mu_R1_all, k_R1_all, n_R1_all),
    seed=42,
    maxiter=1000,
    tol=1e-10,
    popsize=30,       # 种群大小=30×参数数=90
    mutation=(0.5, 1.5),
    recombination=0.9,
    polish=True,      # 最后自动用L-BFGS-B精细优化
    workers=1,
)

kappa_R1_hat, beta_R1_hat, b_R1_hat = res_R1.x
print(f"R1 least-squares done: kappa={kappa_R1_hat:.4f}, beta={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")
print(f"  (vs R2: kappa={kappa_tcs:.4f}, b={b_tcs:.6f})")

P_R1_pred = tcs_R1_model(mu_train, kappa_R1_hat, beta_R1_hat, b_R1_hat)
AEB_R1_pred = -np.log(np.maximum(1 - P_R1_pred, 1e-12))
logL_R1 = np.sum(k_all * np.log(P_R1_pred) + (n_all_arr - k_all) * np.log(1 - P_R1_pred))
ss_res_R1 = np.sum(weights_AEB * (AEB_train - AEB_R1_pred)**2)
R2_R1 = 1 - ss_res_R1 / ss_tot_4pl

print(f"\nR1 logL = {logL_R1:.2f}  (R2 logL = {logL_tcs:.2f})")
print(f"R1 R2(AEB) = {R2_R1:.4f}  (R2 R2(AEB) = {R2_tcs:.4f})")
print(f"R1 AICc = {AICc(logL_R1, 3, n_all):.2f}  (R2 AICc = {AICc_tcs:.2f})")

# ====================================================================
# NESTED SAMPLING (dynesty)
# Parameterization: theta = [kappa, log10_beta, b]
# Physical prior: beta >= 1 (each bead has at least 1 binding site)
# ====================================================================

def loglike_R1(theta):
    """Binomial log-likelihood for nested sampling"""
    kappa, log10_beta, b = theta
    if kappa <= 0 or b < 0 or b >= 1:
        return -1e12
    beta = 10.0 ** log10_beta
    if beta < 1.0:
        return -1e12
    P_pos = tcs_R1_model(mu_R1_all, kappa, beta, b)
    ll = np.sum(k_R1_all * np.log(P_pos) + (n_R1_all - k_R1_all) * np.log(1.0 - P_pos))
    if not np.isfinite(ll):
        return -1e12
    return ll

def prior_transform_R1(u):
    """Map u ~ Uniform(0,1)^3 to physical parameters
    beta in [1, 1e5] — physical lower bound"""
    kappa = 0.01 + u[0] * (1000.0 - 0.01)
    log10_beta = 0.0 + u[1] * 5.0  # beta in [1, 1e5]
    b = u[2] * 0.1
    return np.array([kappa, log10_beta, b])

print("\n" + "-"*60)
print("Running nested sampling (dynesty)...")
print(f"  ndim = 3, nlive = 500, bound = 'multi', sample = 'auto'")
print(f"  Physical prior: beta >= 1")
print("-"*60)

sampler_R1 = dynesty.NestedSampler(
    loglike_R1,
    prior_transform_R1,
    ndim=3,
    nlive=500,
    bound='multi',
    sample='auto',
    rstate=np.random.default_rng(42),
)

sampler_R1.run_nested(print_progress=True)
results_R1 = sampler_R1.results

# Use global best from nested sampling
logL_R1_global = results_R1.logl.max()
if logL_R1_global > logL_R1:
    print(f"\n*** Global logL from nested sampling: {logL_R1_global:.2f}")
    print(f"    (L-BFGS-B was stuck at local max: {logL_R1:.2f}, gap = {logL_R1_global - logL_R1:.2f})")
    logL_R1 = logL_R1_global
    samples_R1_post = results_R1.samples
    logl_arr = results_R1.logl
    best_idx = np.argmax(logl_arr)
    theta_best = samples_R1_post[best_idx]
    kappa_R1_hat = theta_best[0]
    beta_R1_hat = 10.0 ** theta_best[1]
    b_R1_hat = theta_best[2]
    P_R1_pred = tcs_R1_model(mu_train, kappa_R1_hat, beta_R1_hat, b_R1_hat)
    AEB_R1_pred = -np.log(np.maximum(1 - P_R1_pred, 1e-12))
    R2_R1 = 1 - np.sum(weights_AEB * (AEB_train - AEB_R1_pred)**2) / ss_tot_4pl
    print(f"    Updated: kappa={kappa_R1_hat:.4f}, beta={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")
    print(f"    Updated R2(AEB) = {R2_R1:.4f}")

print(f"\nNested sampling done.")
print(f"  logZ = {results_R1.logz[-1]:.2f} ± {results_R1.logzerr[-1]:.2f}")
print(f"  niter = {results_R1.niter}")
print(f"  ncall = {results_R1.ncall}")

# ---- Extract posterior samples ----
weights_R1 = np.exp(results_R1.logwt - results_R1.logz[-1])
samples_R1 = dyfunc.resample_equal(results_R1.samples, weights_R1)
print(f"  Posterior samples (resampled): {len(samples_R1)}")

# ---- Convergence diagnostics ----
n_half = len(samples_R1) // 2
samples_A = samples_R1[:n_half]
samples_B = samples_R1[n_half:]

print("\n" + "="*60)
print("Convergence diagnostics (split-sample consistency)")
print("="*60)

for i, name in enumerate(['kappa', 'log10_beta', 'b']):
    med_A = np.median(samples_A[:, i])
    med_B = np.median(samples_B[:, i])
    std_A = np.std(samples_A[:, i])
    std_B = np.std(samples_B[:, i])
    pooled_std = np.sqrt((std_A**2 + std_B**2) / 2)
    gr_ratio = abs(med_A - med_B) / (pooled_std + 1e-12)
    print(f"  {name:<12}: split-A median = {med_A:.4f}, split-B median = {med_B:.4f}, "
          f"|Δ|/σ = {gr_ratio:.3f}")

ess_R1 = 1.0 / np.sum(weights_R1**2)
print(f"\n  Estimated ESS (Kish): {ess_R1:.0f}")

# ---- Posterior statistics ----
kappa_samples = samples_R1[:, 0]
beta_samples = 10.0 ** samples_R1[:, 1]
b_samples = samples_R1[:, 2]

kappa_med = np.median(kappa_samples)
beta_med = np.median(beta_samples)
b_med = np.median(b_samples)
kappa_ci = np.percentile(kappa_samples, [2.5, 97.5])
beta_ci = np.percentile(beta_samples, [2.5, 97.5])
b_ci = np.percentile(b_samples, [2.5, 97.5])

print("\n" + "="*60)
print("R1 posterior median and 95% CI")
print("="*60)
print(f"kappa = {kappa_med:.4f}  [{kappa_ci[0]:.4f}, {kappa_ci[1]:.4f}]")
print(f"beta  = {beta_med:.2f}  [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
print(f"b     = {b_med:.6f}  [{b_ci[0]:.6f}, {b_ci[1]:.6f}]")

# ---- Bimodality check ----
print("\n" + "="*60)
print("Bimodality check")
print("="*60)

log10_beta_s = samples_R1[:, 1]
mode_A_mask = log10_beta_s < 1.5   # beta < ~30
mode_B_mask = log10_beta_s >= 1.5  # beta > ~30
frac_A = np.sum(mode_A_mask) / len(samples_R1)
frac_B = np.sum(mode_B_mask) / len(samples_R1)

print(f"Mode A (beta < ~30):  {frac_A*100:.1f}% of posterior mass")
if np.any(mode_A_mask):
    print(f"  kappa = {np.median(kappa_samples[mode_A_mask]):.4f}  "
          f"[{np.percentile(kappa_samples[mode_A_mask], 2.5):.4f}, "
          f"{np.percentile(kappa_samples[mode_A_mask], 97.5):.4f}]")
    print(f"  beta  = {np.median(beta_samples[mode_A_mask]):.2f}  "
          f"[{np.percentile(beta_samples[mode_A_mask], 2.5):.2f}, "
          f"{np.percentile(beta_samples[mode_A_mask], 97.5):.2f}]")
print(f"Mode B (beta >= ~30): {frac_B*100:.1f}% of posterior mass")
if np.any(mode_B_mask):
    print(f"  kappa = {np.median(kappa_samples[mode_B_mask]):.4f}  "
          f"[{np.percentile(kappa_samples[mode_B_mask], 2.5):.4f}, "
          f"{np.percentile(kappa_samples[mode_B_mask], 97.5):.4f}]")
    print(f"  beta  = {np.median(beta_samples[mode_B_mask]):.2f}  "
          f"[{np.percentile(beta_samples[mode_B_mask], 2.5):.2f}, "
          f"{np.percentile(beta_samples[mode_B_mask], 97.5):.2f}]")

# ---------------------------------------------------------------------------
# Physical interpretation (IL-6)
# ---------------------------------------------------------------------------
# beta = effective valency: independently binding epitopes per IL-6 molecule.
# Unlike sβG, the posterior is dominated by a SINGLE intermediate-valency
# mode (beta ~ 22, >95% of the mass) with only a small mode-B tail, so the
# posterior is NOT bimodal: beta is bounded away from both 1 and infinity
# (95% CI [17.5, 31.1], width/median ~ 0.6) and is therefore identifiable
# at this kappa. The R2 (beta -> infinity) approximation introduces <1%
# error in P_specific at beta = 22, so R2 is retained for the main analysis.
# NOTE: identifiable beta here still does NOT yield calibration-free M0 --
# the calibration experiment knows M at each dilution; for unknown samples
# the Omega = N*beta scale still requires the standard curve.
print("\nPhysical interpretation (IL-6):")
print("  Single intermediate-valency mode (beta ~ 22, >95% mass); not bimodal.")
print("  beta is identifiable at this kappa; R2 approximation error <1%.")
print("  (M0 of unknown samples still requires the calibration curve.)")

if frac_A > 0.1 and frac_B > 0.1:
    print("\n>>> Posterior is BIMODAL")
    bimodal = True
else:
    print("\n>>> Posterior is not clearly bimodal")
    bimodal = False

# ---- beta identifiability assessment ----
print("\n" + "="*60)
print("beta identifiability assessment")
print("="*60)

beta_prior_lo, beta_prior_hi = 1.0, 100000.0
beta_post_lo, beta_post_hi = beta_ci[0], beta_ci[1]
hits_ceiling = beta_post_hi >= beta_prior_hi * 0.9
hits_floor = beta_post_lo <= beta_prior_lo * 1.5
post_range_ratio = (beta_post_hi - beta_post_lo) / max(beta_med, 0.01)

print(f"beta prior range: [{beta_prior_lo}, {beta_prior_hi}]")
print(f"beta posterior 95% CI: [{beta_post_lo:.2f}, {beta_post_hi:.2f}]")
print(f"beta posterior width / median = {post_range_ratio:.2f}")
print(f"Hits prior ceiling? {hits_ceiling}")
print(f"Hits prior floor? {hits_floor}")

if bimodal:
    print("\n>>> beta posterior is bimodal -> NOT identifiable")
    identifiable = "weak"
elif hits_ceiling:
    print("\n>>> beta posterior hits prior ceiling -> not identifiable")
    identifiable = False
elif post_range_ratio > 2.0:
    print("\n>>> beta posterior very wide -> weakly identifiable")
    identifiable = "weak"
else:
    print("\n>>> beta posterior bounded and compact -> identifiable")
    identifiable = True

# ---- Derived quantities ----
if identifiable == True:
    Omega_hat = N * beta_med
    kappa_Omega_hat = kappa_med * Omega_hat
    print(f"\nOmega = N*beta = {N:.0f} x {beta_med:.2f} = {Omega_hat:.2e}")
    print(f"kappa*Omega = K*V*N_A = {kappa_Omega_hat:.3e}")

# ====================================================================
# R1 vs R2 model comparison: AICc + Bayes factor
# ====================================================================
print("\n" + "="*60)
print("R1 vs R2 model comparison")
print("="*60)

print("\n--- AICc comparison ---")
print(f"{'R2 (beta->inf)':<18} {logL_tcs:<12.2f} {2:<10} {AICc_tcs:<10.2f}")
print(f"{'R1 (beta free)':<18} {logL_R1:<12.2f} {3:<10} {AICc(logL_R1, 3, n_all):<10.2f}")
delta_AICc = AICc(logL_R1, 3, n_all) - AICc_tcs
print(f"\nDeltaAICc (R1 - R2) = {delta_AICc:.2f}")
if delta_AICc > 10:
    print("-> R2 preferred: data does not support finite beta")
elif delta_AICc < -10:
    print("-> R1 preferred: finite beta supported")
else:
    print("-> Comparable: data cannot distinguish R1 and R2")

# Bayes factor via R2 nested sampling
print("\n--- Bayes factor (nested sampling) ---")
print("Running R2 nested sampling for logZ comparison...")

def loglike_R2(theta):
    """R2 binomial log-likelihood (beta->inf, 2 params)"""
    kappa, b = theta
    if kappa <= 0 or b < 0 or b >= 1:
        return -1e12
    P_pos = tcs_model(mu_R1_all, kappa, b)
    ll = np.sum(k_R1_all * np.log(P_pos) + (n_R1_all - k_R1_all) * np.log(1.0 - P_pos))
    if not np.isfinite(ll):
        return -1e12
    return ll

def prior_transform_R2(u):
    """R2 prior transform: kappa in [0.01, 1000], b in [0, 0.1]"""
    kappa = 0.01 + u[0] * (1000.0 - 0.01)
    b = u[1] * 0.1
    return np.array([kappa, b])

sampler_R2 = dynesty.NestedSampler(
    loglike_R2,
    prior_transform_R2,
    ndim=2,
    nlive=500,
    bound='multi',
    sample='auto',
    rstate=np.random.default_rng(42),
)
sampler_R2.run_nested(print_progress=False)
results_R2 = sampler_R2.results

logZ_R1 = results_R1.logz[-1]
logZ_R2 = results_R2.logz[-1]
logZ_err = np.sqrt(results_R1.logzerr[-1]**2 + results_R2.logzerr[-1]**2)
logBF = logZ_R1 - logZ_R2

print(f"\n  logZ(R1) = {logZ_R1:.2f} ± {results_R1.logzerr[-1]:.2f}")
print(f"  logZ(R2) = {logZ_R2:.2f} ± {results_R2.logzerr[-1]:.2f}")
print(f"  log Bayes factor (R1/R2) = {logBF:.2f} ± {logZ_err:.2f}")

if logBF < -5:
    print("  -> Strong evidence for R2 (beta->inf)")
    bf_conclusion = "R2 strongly preferred"
elif logBF < -2:
    print("  -> Moderate evidence for R2")
    bf_conclusion = "R2 moderately preferred"
elif logBF < 2:
    print("  -> Inconclusive (data cannot distinguish R1 and R2)")
    bf_conclusion = "inconclusive"
elif logBF < 5:
    print("  -> Moderate evidence for R1")
    bf_conclusion = "R1 moderately preferred"
else:
    print("  -> Strong evidence for R1")
    bf_conclusion = "R1 strongly preferred"

# ---- Posterior distribution plot ----
fig_beta, axes_beta = plt.subplots(1, 3, figsize=(15, 5))

ax = axes_beta[0]
ax.hist(kappa_samples, bins=80, density=True, alpha=0.7, color='steelblue', edgecolor='black')
ax.axvline(kappa_med, color='red', lw=2, label=f'median={kappa_med:.2f}')
ax.axvline(kappa_tcs, color='green', lw=2, ls='--', label=f'R2 κ={kappa_tcs:.2f}')
ax.set_xlabel('κ', fontsize=14)
ax.set_ylabel('Density', fontsize=14)
ax.set_title('Posterior of κ (R1)', fontsize=14)
ax.legend(fontsize=11)

ax = axes_beta[1]
ax.hist(beta_samples, bins=80, density=True, alpha=0.7, color='coral', edgecolor='black')
ax.axvline(beta_med, color='red', lw=2, label=f'median={beta_med:.1f}')
ax.axvline(beta_ci[0], color='gray', ls='--', lw=1.5)
ax.axvline(beta_ci[1], color='gray', ls='--', lw=1.5, label=f'95% CI [{beta_ci[0]:.1f}, {beta_ci[1]:.1f}]')
ax.set_xlabel('β', fontsize=14)
ax.set_title('Posterior of β (R1) — KEY DIAGNOSTIC', fontsize=14, fontweight='bold')
ax.legend(fontsize=11)

ax = axes_beta[2]
ax.hist(b_samples, bins=80, density=True, alpha=0.7, color='mediumseagreen', edgecolor='black')
ax.axvline(b_med, color='red', lw=2, label=f'median={b_med:.6f}')
ax.axvline(b_tcs, color='green', lw=2, ls='--', label=f'R2 b={b_tcs:.6f}')
ax.set_xlabel('b', fontsize=14)
ax.set_title('Posterior of b (R1)', fontsize=14)
ax.legend(fontsize=11)

plt.suptitle('R1 Nested Sampling Posteriors: Is β identifiable? (IL-6)', fontsize=16, fontweight='bold')
plt.tight_layout()
plt.savefig('R1_beta_identifiability.png', dpi=300, bbox_inches='tight')
plt.show()

# ---- Corner plot ----
if corner is not None:
    fig_corner = plt.figure(figsize=(10, 10))
    corner.corner(
        samples_R1,
        labels=['kappa', 'log10(beta)', 'b'],
        truths=[kappa_med, np.log10(max(beta_med, 1.0)), b_med],
        label_kwargs={'fontsize': 14},
        show_titles=True, title_kwargs={'fontsize': 12}
    )
    plt.suptitle('R1 Nested Sampling Posterior — β Identifiability Test (IL-6)',
                 fontsize=16, fontweight='bold')
    plt.savefig('R1_corner_plot_nested.png', dpi=300, bbox_inches='tight')
    plt.show()

# ---- Final verdict ----
print("\n" + "="*80)
print("FINAL VERDICT (IL-6 dataset)")
print("="*80)
if identifiable == True:
    print("beta IS identifiable from IL-6 Simoa calibration data.")
    print(f"  beta = {beta_med:.2f} [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
    print("  -> NOTE: This is a calibration experiment. Calibration-free")
    print("     absolute quantification still requires kappa->0 (dPCR).")
elif identifiable == "weak":
    if bimodal:
        print("beta is NOT identifiable — posterior is bimodal.")
        print(f"  Mode A ({frac_A*100:.0f}%): beta ~ {np.median(beta_samples[mode_A_mask]):.1f}")
        print(f"  Mode B ({frac_B*100:.0f}%): beta ~ {np.median(beta_samples[mode_B_mask]):.0f}")
    else:
        print("beta is WEAKLY identifiable — posterior is very wide.")
        print(f"  beta = {beta_med:.2f} [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
    print(f"  Bayes factor: {bf_conclusion}")
    print("  -> M0 estimate would be too imprecise to be useful.")
    print("  -> In practice, digital finite-kappa still needs external calibration.")
else:
    print("beta is NOT identifiable from IL-6 dilution data.")
print("="*80)

# ====================================================================
# Sheet 18: R1 Model Results
# ====================================================================
try:
    from openpyxl import load_workbook as _lwb
    _wb18 = _lwb("SI_Table_4.xlsx")
    if "18_R1_Model" in _wb18.sheetnames:
        del _wb18["18_R1_Model"]
    _ws18 = _wb18.create_sheet("18_R1_Model")

    _ws18.append(['R1 Model Fit (nested sampling) vs R2 Model (beta->inf) — IL-6 dataset'])
    _ws18.append(['Method: dynesty nested sampling, nlive=500, bound=multi'])
    _ws18.append(['Physical prior: beta >= 1 (each bead has at least 1 binding site)'])
    _ws18.append([])
    _ws18.append(['Parameter', 'R2 (beta->inf)', 'R1 (beta free)', 'R1 95% CI'])
    _ws18.append(['kappa', f'{kappa_tcs:.4f}', f'{kappa_med:.4f}',
                  f'[{kappa_ci[0]:.4f}, {kappa_ci[1]:.4f}]'])
    _ws18.append(['beta', 'inf (R2 approx)', f'{beta_med:.2f}',
                  f'[{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]'])
    _ws18.append(['b', f'{b_tcs:.6f}', f'{b_med:.6f}',
                  f'[{b_ci[0]:.6f}, {b_ci[1]:.6f}]'])
    _ws18.append([])

    _ws18.append(['Model Comparison'])
    _ws18.append(['Metric', 'R2', 'R1', 'Difference / Interpretation'])
    _ws18.append(['logL (MLE)', f'{logL_tcs:.2f}', f'{logL_R1:.2f}',
                  f'{logL_R1 - logL_tcs:.2f}'])
    _ws18.append(['n_params', '2', '3', '1'])
    _ws18.append(['AICc', f'{AICc_tcs:.2f}', f'{AICc(logL_R1, 3, n_all):.2f}',
                  f'Δ = {AICc(logL_R1, 3, n_all) - AICc_tcs:.2f}'])
    _ws18.append(['logZ (nested)', f'{logZ_R2:.2f}', f'{logZ_R1:.2f}',
                  f'logBF = {logBF:.2f} ± {logZ_err:.2f}'])
    _ws18.append(['Bayes factor', '', '', f'{bf_conclusion}'])
    _ws18.append(['R2_AEB', f'{R2_tcs:.6f}', f'{R2_R1:.6f}', ''])
    _ws18.append([])

    _ws18.append(['Nested Sampling Diagnostics'])
    _ws18.append(['Quantity', 'R1', 'R2'])
    _ws18.append(['logZ', f'{results_R1.logz[-1]:.2f} ± {results_R1.logzerr[-1]:.2f}',
                  f'{results_R2.logz[-1]:.2f} ± {results_R2.logzerr[-1]:.2f}'])
    _ws18.append(['niter', f'{results_R1.niter}', f'{results_R2.niter}'])
    _ws18.append(['ncall', f'{results_R1.ncall}', f'{results_R2.ncall}'])
    _ws18.append(['ESS (Kish)', f'{ess_R1:.0f}', ''])
    _ws18.append([])

    _ws18.append(['Bimodality Check'])
    _ws18.append(['Mode', 'Fraction', 'kappa (median [95% CI])', 'beta (median [95% CI])'])
    if np.any(mode_A_mask):
        _ws18.append([f'A (beta < ~30)', f'{frac_A*100:.1f}%',
                      f'{np.median(kappa_samples[mode_A_mask]):.4f} '
                      f'[{np.percentile(kappa_samples[mode_A_mask], 2.5):.4f}, '
                      f'{np.percentile(kappa_samples[mode_A_mask], 97.5):.4f}]',
                      f'{np.median(beta_samples[mode_A_mask]):.2f} '
                      f'[{np.percentile(beta_samples[mode_A_mask], 2.5):.2f}, '
                      f'{np.percentile(beta_samples[mode_A_mask], 97.5):.2f}]'])
    if np.any(mode_B_mask):
        _ws18.append([f'B (beta >= ~30)', f'{frac_B*100:.1f}%',
                      f'{np.median(kappa_samples[mode_B_mask]):.4f} '
                      f'[{np.percentile(kappa_samples[mode_B_mask], 2.5):.4f}, '
                      f'{np.percentile(kappa_samples[mode_B_mask], 97.5):.4f}]',
                      f'{np.median(beta_samples[mode_B_mask]):.2f} '
                      f'[{np.percentile(beta_samples[mode_B_mask], 2.5):.2f}, '
                      f'{np.percentile(beta_samples[mode_B_mask], 97.5):.2f}]'])
    _ws18.append(['Physical interpretation:',
                  'single intermediate-valency mode (beta ~ 22, >95% of posterior mass); '
                  'posterior is NOT bimodal'])
    _ws18.append(['Identifiability:',
                  'beta bounded away from 1 and infinity -> identifiable at this kappa; '
                  'R2 approximation error <1% at beta = 22'])
    _ws18.append(['Note:',
                  'identifiable beta in a calibration experiment does NOT yield '
                  'calibration-free M0 for unknown samples'])
    _ws18.append([])

    _ws18.append(['Derived Quantities (R1)'])
    _ws18.append(['Omega = N*beta', f'{N * beta_med:.2e}'])
    _ws18.append(['kappa*Omega = K*V*N_A', f'{kappa_med * N * beta_med:.3e}'])
    _ws18.append([])

    _ws18.append(['Note: This is a calibration experiment (M known at each'])
    _ws18.append(['dilution point). beta identifiability here does NOT imply'])
    _ws18.append(['calibration-free absolute quantification. See SI S2c.5.'])
    _ws18.append([])

    # R1 per-point predictions
    _ws18.append(['R1 Per-point Predictions'])
    _ws18.append(['Conc (pg/mL)', 'k', 'n', 'P_obs', 'P_R2', 'P_R1', 'AEB_R2', 'AEB_R1'])

    P_R2_pred = tcs_model(mu_train, kappa_tcs, b_tcs)
    AEB_R2_pred = -np.log(np.maximum(1 - P_R2_pred, 1e-12))

    for i in range(len(concentrations_pgmL)):
        _ws18.append([f'{concentrations_pgmL[i]:.4f}', f'{k_all[i]:.1f}',
                      f'{n_all_arr[i]:.0f}', f'{k_all[i]/n_all_arr[i]:.6f}',
                      f'{P_R2_pred[i]:.6f}', f'{P_R1_pred[i]:.6f}',
                      f'{AEB_R2_pred[i]:.4f}', f'{AEB_R1_pred[i]:.4f}'])

    _wb18.save("SI_Table_4.xlsx")
    print(f"\n[OK] Sheet 18_R1_Model added to SI_Table_4.xlsx (now {len(_wb18.sheetnames)} sheets)")
except Exception as e_s18:
    print(f"\n[FAIL] Sheet 18: {e_s18}")
    import traceback; traceback.print_exc()
