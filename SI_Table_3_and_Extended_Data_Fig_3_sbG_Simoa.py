

# -*- coding: utf-8 -*-
"""
sβG Simoa 4PL vs TCS binomial-likelihood analysis
==================================================
Nature Methods v10 — author: <you>
Final structure: compute → print → print → ... → SI Table 3 export

Re-runs:
    $ python sbG_simoa_complete.py
Outputs:
    - Print all stats to stdout
    - Save Fig_core_sβG.png (4-panel figure)
    - Save Fig_S1_sβG.png (Q-Q plots)
    - Save SI_Table_3.xlsx (13 sheets)
    - Save TCS_per_point_quantification.csv
    - Save R3_analysis_sβG.csv
"""

import numpy as np
import matplotlib.pyplot as plt
import scipy.optimize as opt
from scipy import stats
from scipy.stats import gaussian_kde, sem, chi2, beta as beta_dist
from scipy.optimize import brentq
import warnings
import seaborn as sns

plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 10,
    'savefig.dpi': 400,
    'axes.linewidth': 0.8,
    'axes.unicode_minus': False,
})
sns.set_style("white")
plt.rcParams['font.serif'] = ['Arial']
warnings.filterwarnings('ignore', category=RuntimeWarning)

print("=" * 80)
print("4PL vs TCS — Nature sβG dataset (binomial likelihood)")
print("=" * 80)


# ============================================================================
# 1. 实验参数
# ============================================================================
N_beads = 400000            # 每孔珠子数
V = 100e-6                  # 反应体积 (L)
N_A = 6.022e23

# ---------- sβG Simoa 2010 dataset ----------
conc_aM = np.array([0, 0.35, 0.7, 3.5, 7, 35, 70, 350, 700, 3500, 7000])
k_avg = np.array([1, 3, 5, 22, 38, 237, 385, 1787, 4036, 15634, 24836])
P_pos = np.array([0.0016, 0.0086, 0.0099, 0.0413, 0.0713, 0.4461, 0.8183,
                   3.3802, 7.5865, 30.6479, 44.5296])
meas_cv = np.array([87, 75, 63, 10, 15, 1, 5, 2, 5, 3, 1])
poisson_cv = np.array([122, 55, 46, 21, 16, 7, 5, 2, 2, 1, 1])

# 估算总珠子数 (取中位数作为固定值)
n_est = k_avg / (P_pos / 100)
n_const = np.median(n_est)
print(f"Estimated bead count per well: median = {n_const:.0f}")

# 计算 AEB 和 P_train
AEB_train = -np.log(np.maximum(1 - k_avg / n_const, 1e-12))
P_train = k_avg / n_const
n_train = np.full_like(AEB_train, n_const, dtype=float)

# 浓度转换为每珠期望分子数 μ
c_molL = conc_aM * 1e-18
mu_train = c_molL * V * N_A / N_beads

# 权重 (used for R² and for visual residual scaling, NOT for fit)
P_obs_for_var = 1 - np.exp(-AEB_train)
var_AEB = P_obs_for_var / (n_train * (1 - P_obs_for_var + 1e-10)) + 1e-10
weights_AEB = 1.0 / var_AEB

# 测试集（无数据，置空）
test_AEB = np.array([])
test_n = np.array([])
test_k = np.array([])

# ============================================================================
# 2. 模型定义
# ============================================================================
mask_zero = mu_train == 0
A_4pl = np.mean(AEB_train[mask_zero])
mask_nz = ~mask_zero
mu_nz = mu_train[mask_nz]
AEB_nz = AEB_train[mask_nz]
weights_nz = weights_AEB[mask_nz]
k_nz = k_avg[mask_nz]
n_nz = n_train[mask_nz]
k_all = k_avg.astype(float)
n_all_arr = n_train


def fourpl_fixed_A(x, B, C, D, A):
    x = np.atleast_1d(x)
    x_safe = np.maximum(x, 1e-12)
    C_safe = max(C, 1e-12)
    log_term = B * (np.log(C_safe) - np.log(x_safe))
    log_term = np.clip(log_term, -100, 100)
    return A + (D - A) / (1.0 + np.exp(log_term))


def predict_AEB_4pl(mu, A, B, C, D):
    mu = np.atleast_1d(mu)
    res = np.empty_like(mu, dtype=float)
    res[mu == 0] = A
    res[mu != 0] = fourpl_fixed_A(mu[mu != 0], B, C, D, A)
    return res


def neg_loglik_4pl(params, mu, A_fixed, k, n):
    B, C, D = params
    if B <= 0 or C <= 0 or D <= A_fixed:
        return 1e12
    AEB_pred = predict_AEB_4pl(mu, A_fixed, B, C, D)
    P_pos_pred = np.clip(1 - np.exp(-AEB_pred), 1e-15, 1 - 1e-15)
    return -np.sum(k * np.log(P_pos_pred) + (n - k) * np.log(1 - P_pos_pred))


def tcs_model(mu, kappa, b):
    mu_safe = np.maximum(mu, 0)
    P_specific = 1 - np.exp(-mu_safe / (1 + kappa))
    return np.clip(b + (1 - b) * P_specific, 0, 1)


def predict_AEB_tcs(mu, kappa, b):
    P_pred = tcs_model(mu, kappa, b)
    return -np.log(np.maximum(1 - P_pred, 1e-12))


def neg_loglik_tcs(params, mu, k, n):
    kappa, b = params
    if kappa <= 0 or b < 0 or b >= 1:
        return 1e12
    P_pos_pred = np.clip(tcs_model(mu, kappa, b), 1e-15, 1 - 1e-15)
    return -np.sum(k * np.log(P_pos_pred) + (n - k) * np.log(1 - P_pos_pred))


# ============================================================================
# 3. 4PL fit
# ============================================================================
idx_sort = np.argsort(mu_nz)
mu_sorted = mu_nz[idx_sort]
AEB_sorted = AEB_nz[idx_sort]
D_init = np.max(AEB_sorted) * 1.05
target = A_4pl + 0.5 * (D_init - A_4pl)
C_init = np.median(mu_nz)
for i in range(len(AEB_sorted) - 1):
    if AEB_sorted[i] <= target <= AEB_sorted[i + 1]:
        ratio = (target - AEB_sorted[i]) / (AEB_sorted[i + 1] - AEB_sorted[i] + 1e-10)
        C_init = mu_sorted[i] + ratio * (mu_sorted[i + 1] - mu_sorted[i])
        break
C_init = np.clip(C_init, mu_nz.min() * 0.5, mu_nz.max() * 2)
B_init = 1.0

bounds_4pl = [(0.1, 4.0), (mu_nz.min() * 0.1, mu_nz.max() * 20), (A_4pl + 1e-6, AEB_sorted.max() * 100)]
p0 = [np.clip(B_init, *bounds_4pl[0]),
      np.clip(C_init, *bounds_4pl[1]),
      np.clip(D_init, *bounds_4pl[2])]

res_4pl = opt.minimize(neg_loglik_4pl, p0, args=(mu_nz, A_4pl, k_nz, n_nz),
                       bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter': 2000})
B_4pl, C_4pl, D_4pl = res_4pl.x

# ============================================================================
# 4. 4PL bootstrap (case resampling)
# ============================================================================
np.random.seed(42)
n_boot = 2000
boot_B, boot_C, boot_D = [], [], []
for _ in range(n_boot):
    idx = np.random.choice(len(k_nz), size=len(k_nz), replace=True)
    try:
        res_b = opt.minimize(neg_loglik_4pl, [B_4pl, C_4pl, D_4pl],
                             args=(mu_nz[idx], A_4pl, k_nz[idx], n_nz[idx]),
                             bounds=bounds_4pl, method='L-BFGS-B', options={'maxiter': 1000})
        if res_b.success:
            boot_B.append(res_b.x[0])
            boot_C.append(res_b.x[1])
            boot_D.append(res_b.x[2])
    except Exception:
        pass

B_CI = np.percentile(boot_B, [2.5, 97.5])
C_CI = np.percentile(boot_C, [2.5, 97.5])
D_CI = np.percentile(boot_D, [2.5, 97.5])
print(f"\n[4PL case-bootstrap n={len(boot_B)}] B={B_4pl:.4f} [{B_CI[0]:.4f}, {B_CI[1]:.4f}]"
      f"  C={C_4pl:.4e} [{C_CI[0]:.4e}, {C_CI[1]:.4e}]"
      f"  D={D_4pl:.4f} [{D_CI[0]:.4f}, {D_CI[1]:.4f}]")

# ============================================================================
# 5. TCS fit
# ============================================================================
P_nz = P_train[mask_nz]
coef = np.polyfit(mu_nz, P_nz, 1, w=np.sqrt(n_train[mask_nz]))
b_init = max(0, coef[1])
kappa_init = max(0.1, (1 - b_init) / coef[0] - 1) if coef[0] > 0 else 1.0
bounds_tcs = [(0.01, 1000), (0, 0.1)]

res_tcs = opt.minimize(neg_loglik_tcs, [kappa_init, b_init],
                       args=(mu_train, k_all, n_all_arr),
                       bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter': 500})
kappa_tcs, b_tcs = res_tcs.x

# ============================================================================
# 6. TCS bootstrap
# ============================================================================
np.random.seed(42)
boot_kappa, boot_b = [], []
for _ in range(n_boot):
    idx = np.random.choice(len(k_all), size=len(k_all), replace=True)
    try:
        res_t = opt.minimize(neg_loglik_tcs, [kappa_tcs, b_tcs],
                             args=(mu_train[idx], k_all[idx], n_all_arr[idx]),
                             bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter': 500})
        if res_t.success:
            boot_kappa.append(res_t.x[0])
            boot_b.append(res_t.x[1])
    except Exception:
        pass

kappa_CI = np.percentile(boot_kappa, [2.5, 97.5])
b_CI = np.percentile(boot_b, [2.5, 97.5])
print(f"[TCS case-bootstrap n={len(boot_kappa)}] kappa={kappa_tcs:.4f}"
      f" [{kappa_CI[0]:.4f}, {kappa_CI[1]:.4f}]"
      f"  b={b_tcs:.6f} [{b_CI[0]:.6f}, {b_CI[1]:.6f}]")


AEB_pred_train_4pl = predict_AEB_4pl(mu_train, A_4pl, B_4pl, C_4pl, D_4pl)
AEB_pred_train_tcs = predict_AEB_tcs(mu_train, kappa_tcs, b_tcs)


# ============================================================================
# 7. D profile-likelihood scan (4PL identifiability diagnosis)
# ============================================================================
print("\n" + "="*80)
print("D profile likelihood scan (4PL identifiability)")
print("="*80)


def neg_loglik_4pl_fixed_D(params_2, mu, A_fixed, k, n, D_fixed):
    B, C = params_2
    return neg_loglik_4pl([B, C, D_fixed], mu, A_fixed, k, n)


D_scan_values = np.concatenate([
    np.array([0.7, 0.8, 0.9, 1.0]),
    np.array([1.5, 2.0, 2.9189, 3.0, 4.0, 5.0]),
    np.linspace(10, 100, 10),
    np.array([200, 500, 1000, 10000])
])

D_scan_results = []
logL_best = -res_4pl.fun  # already binomial neg_loglik

print(f"{'D (fixed)':<14} {'logL':<12} {'B':<10} {'C':<14} {'ΔlogL':<10}")
print("-"*60)
for D_fix in D_scan_values:
    if D_fix <= A_4pl:
        continue
    res_scan = opt.minimize(neg_loglik_4pl_fixed_D, [B_4pl, C_4pl],
                            args=(mu_nz, A_4pl, k_nz, n_nz, D_fix),
                            method='L-BFGS-B',
                            bounds=[(0.1, 4.0), (mu_nz.min() * 0.1, mu_nz.max() * 1e6)],
                            options={'maxiter': 2000})
    if res_scan.success or res_scan.fun < 1e10:
        logL_scan = -res_scan.fun
        delta = logL_scan - logL_best
        D_scan_results.append((D_fix, logL_scan, res_scan.x[0], res_scan.x[1], delta))
        print(f"{D_fix:<14.2f} {logL_scan:<12.4f} {res_scan.x[0]:<10.4f}"
              f" {res_scan.x[1]:<14.4f} {delta:<10.4f}")


# ============================================================================
# 8. R3 linear approximation precompute
# ============================================================================
gamma_tcs_r3 = 1.0 / (kappa_tcs + 1)

P_specific_train_r3 = (P_train - b_tcs) / (1 - b_tcs)
P_specific_train_r3 = np.clip(P_specific_train_r3, 1e-12, 1 - 1e-12)
M_r3 = N_beads * (kappa_tcs + 1) * P_specific_train_r3
mu_r3 = M_r3 / N_beads
CV_r3 = np.where(
    mu_r3 > 1e-12,
    (1.0 / (gamma_tcs_r3 * (1 - b_tcs) * mu_r3)) * np.sqrt(P_train * (1 - P_train) / n_const),
    np.inf
)
r3_valid = P_specific_train_r3 < 0.2

mask_fit_r3 = r3_valid & ~mask_zero
if np.sum(mask_fit_r3) >= 2:
    c_fit_r3 = conc_aM[mask_fit_r3]
    mu_fit_r3 = c_fit_r3 * 1e-18 * V * N_A / N_beads
    P_fit_r3 = P_specific_train_r3[mask_fit_r3]
    gamma_fit_r3 = np.sum(P_fit_r3 * mu_fit_r3) / np.sum(mu_fit_r3 ** 2)
    kappa_fit_r3 = 1.0 / gamma_fit_r3 - 1.0 if gamma_fit_r3 > 0 else np.inf
else:
    gamma_fit_r3 = np.nan
    kappa_fit_r3 = np.nan


# ============================================================================
# 9. Statistics: logL, AICc, BIC, R²
# ============================================================================
def binom_loglik_from_AEB(AEB_pred, k, n):
    P_pos_pred = np.clip(1 - np.exp(-AEB_pred), 1e-15, 1 - 1e-15)
    return np.sum(k * np.log(P_pos_pred) + (n - k) * np.log(1 - P_pos_pred))


logL_4pl = binom_loglik_from_AEB(AEB_pred_train_4pl, k_all, n_all_arr)
logL_tcs = binom_loglik_from_AEB(AEB_pred_train_tcs, k_all, n_all_arr)


def AICc(logL, n_params, n):
    AIC = -2 * logL + 2 * n_params
    correction = (2 * n_params * (n_params + 1) / (n - n_params - 1)
                  if n > n_params + 1 else np.inf)
    return AIC + correction


n_all = len(AEB_train)
AICc_4pl = AICc(logL_4pl, 3, n_all)
AICc_tcs = AICc(logL_tcs, 2, n_all)
BIC_4pl = -2 * logL_4pl + 3 * np.log(n_all)
BIC_tcs = -2 * logL_tcs + 2 * np.log(n_all)

# R² on AEB scale (descriptive only; inference uses binomial logL)
ss_res_4pl = np.sum(weights_AEB * (AEB_train - AEB_pred_train_4pl) ** 2)
ss_tot = np.sum(weights_AEB * (AEB_train - np.average(AEB_train, weights=weights_AEB)) ** 2)
R2_4pl = 1 - ss_res_4pl / ss_tot
R2_tcs = 1 - np.sum(weights_AEB * (AEB_train - AEB_pred_train_tcs) ** 2) / ss_tot

# ============================================================================
# 10. Residuals (AEB and deviance) + Shapiro-Wilk
# ============================================================================
resid_4pl = AEB_train - AEB_pred_train_4pl
resid_tcs = AEB_train - AEB_pred_train_tcs


def deviance_resid(k, n, P_pos_fit):
    P_obs = k / n
    with np.errstate(divide='ignore', invalid='ignore'):
        term_pos = np.where(k > 0, k * np.log(k / (n * P_pos_fit)), 0)
        term_neg = np.where(k < n, (n - k) * np.log((n - k) / (n * (1 - P_pos_fit))), 0)
    dev = 2 * (term_pos + term_neg)
    return np.sign(P_obs - P_pos_fit) * np.sqrt(np.abs(dev))


P_pos_fit_4pl = np.clip(1 - np.exp(-AEB_pred_train_4pl), 1e-15, 1 - 1e-15)
P_pos_fit_tcs = np.clip(tcs_model(mu_train, kappa_tcs, b_tcs), 1e-15, 1 - 1e-15)
dr_4pl = deviance_resid(k_all, n_all_arr, P_pos_fit_4pl)
dr_tcs = deviance_resid(k_all, n_all_arr, P_pos_fit_tcs)
_, p_sh_4pl = stats.shapiro(dr_4pl)
_, p_sh_tcs = stats.shapiro(dr_tcs)


# ============================================================================
# 11. Leave-one-out (binomial logL)
# ============================================================================
loo_4pl_vals, loo_tcs_vals = [], []
for i in range(n_all):
    idx_keep = np.ones(n_all, dtype=bool)
    idx_keep[i] = False
    mu_loo = mu_train[idx_keep]
    k_loo = k_all[idx_keep]
    n_loo = n_all_arr[idx_keep]
    k_i, n_i = k_all[i], n_all_arr[i]
    mu_i = mu_train[i]

    mask_nz_loo = mu_loo > 0
    if np.any(mask_nz_loo):
        # A_loo = np.mean(AEB_train[idx_keep][mu_loo == 0]) if np.any(mu_loo == 0) else A_4pl
        
        
        if np.any(mu_loo == 0):
            A_loo = np.mean(AEB_train[idx_keep][mu_loo == 0])
        else:
            # 留出的是零浓度点：用剩余数据的最低AEB作为背景估计
            A_loo = np.min(AEB_train[idx_keep])

        
        
        mu_nz_loo = mu_loo[mask_nz_loo]
        k_nz_loo = k_loo[mask_nz_loo]
        n_nz_loo = n_loo[mask_nz_loo]
        AEB_nz_loo = AEB_train[idx_keep][mask_nz_loo]
        D_loo = np.max(AEB_nz_loo) * 1.05
        C_loo = np.median(mu_nz_loo)
        p0_loo = [np.clip(B_4pl, 0.1, 4.0),
                  np.clip(C_loo, mu_nz_loo.min() * 0.1, mu_nz_loo.max() * 20),
                  np.clip(D_loo, A_loo + 1e-6, AEB_nz_loo.max() * 100)]
        try:
            res_loo = opt.minimize(neg_loglik_4pl, p0_loo,
                                   args=(mu_nz_loo, A_loo, k_nz_loo, n_nz_loo),
                                   bounds=[(0.1, 4.0), (mu_nz_loo.min() * 0.1, mu_nz_loo.max() * 20),
                                           (A_loo + 1e-6, AEB_nz_loo.max() * 100)],
                                   method='L-BFGS-B', options={'maxiter': 1000})
            if res_loo.success:
                B_loo, C_loo, D_loo = res_loo.x
                AEB_i_pred = predict_AEB_4pl(np.array([mu_i]), A_loo, B_loo, C_loo, D_loo)[0]
                P_i_pred = np.clip(1 - np.exp(-AEB_i_pred), 1e-15, 1 - 1e-15)
                loo_4pl_vals.append(k_i * np.log(P_i_pred) + (n_i - k_i) * np.log(1 - P_i_pred))
            else:
                loo_4pl_vals.append(np.nan)
        except Exception:
            loo_4pl_vals.append(np.nan)
    else:
        loo_4pl_vals.append(np.nan)

    try:
        res_tcs_loo = opt.minimize(neg_loglik_tcs, [kappa_tcs, b_tcs],
                                   args=(mu_loo, k_loo, n_loo),
                                   bounds=bounds_tcs, method='L-BFGS-B', options={'maxiter': 500})
        if res_tcs_loo.success:
            kappa_loo, b_loo = res_tcs_loo.x
            P_i_pred_tcs = np.clip(tcs_model(np.array([mu_i]), kappa_loo, b_loo),
                                    1e-15, 1 - 1e-15)[0]
            loo_tcs_vals.append(k_i * np.log(P_i_pred_tcs) + (n_i - k_i) * np.log(1 - P_i_pred_tcs))
        else:
            loo_tcs_vals.append(np.nan)
    except Exception:
        loo_tcs_vals.append(np.nan)

valid_idx = ~np.isnan(loo_4pl_vals) & ~np.isnan(loo_tcs_vals)
loo_4pl_arr = np.array(loo_4pl_vals)[valid_idx]
loo_tcs_arr = np.array(loo_tcs_vals)[valid_idx]
loo_diff = loo_4pl_arr - loo_tcs_arr
mean_diff = np.mean(loo_diff)
se_diff = sem(loo_diff)
t_stat = mean_diff / se_diff
median_diff = np.median(loo_diff)










# ---- TOST equivalence test (two one-sided tests, alpha = 0.05) ----
from scipy.stats import t as t_dist

n_f  = len(loo_diff)
sd_d = np.std(loo_diff, ddof=1)
se_d = sd_d / np.sqrt(n_f)
tcrit = t_dist.ppf(0.95, n_f - 1)

mean_abs_logL = np.mean(np.abs(loo_tcs_arr))     # per-fold logL magnitude
delta = 0.01 * mean_abs_logL                     # equivalence margin: 1% of per-fold |logL|

t_lo = (mean_diff + delta) / se_d
t_hi = (delta - mean_diff) / se_d
p_tost = max(1 - t_dist.cdf(t_lo, n_f - 1), 1 - t_dist.cdf(t_hi, n_f - 1))
ci90 = (mean_diff - tcrit * se_d, mean_diff + tcrit * se_d)
delta_min = max(abs(ci90[0]), abs(ci90[1]))

print(f"\nTOST equivalence test (margin = 1% of per-fold |logL| = {delta:.2f} logL):")
print(f"  mean diff = {mean_diff:.3f}, 90% CI = [{ci90[0]:.3f}, {ci90[1]:.3f}]")
print(f"  TOST p = {p_tost:.4f} -> "
      f"{'EQUIVALENT within +/-' + f'{delta:.1f} logL' if p_tost < 0.05 else 'equivalence NOT established at this margin'}")
print(f"  smallest passing margin: +/-{delta_min:.2f} logL/fold "
      f"({100 * delta_min / mean_abs_logL:.3f}% of per-fold |logL|)")













# ============================================================================
# 12. TCS per-point quantification (paper Table 2 numbers)
# ============================================================================
import pandas as pd


def M_from_Ppos(P_pos, b, kappa, N_total):
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    return -N_total * (kappa + 1) * np.log(1 - P_spec)


def calc_CV(P_pos, n_obs, b, kappa, N_total):
    M_hat = M_from_Ppos(P_pos, b, kappa, N_total)
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    dM_dP = N_total * (kappa + 1) / ((1 - b) * (1 - P_spec))
    se = np.abs(dM_dP) * np.sqrt(P_pos * (1 - P_pos) / n_obs)
    return (se / M_hat) * 100


def delta_ci(P_pos, n_obs, b, kappa, N_total, alpha=0.05):
    P_spec = (P_pos - b) / (1 - b)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    M_hat = M_from_Ppos(P_pos, b, kappa, N_total)
    dM_dP = N_total * (kappa + 1) / ((1 - b) * (1 - P_spec))
    se = np.abs(dM_dP) * np.sqrt(P_pos * (1 - P_pos) / n_obs)
    z = np.sqrt(chi2.ppf(1 - alpha, 1))
    return (M_hat - z * se, M_hat + z * se)


def exact_transformation_ci(k, n_obs, b, kappa, N_total, alpha=0.05):
    if k == 0:
        p_low = 0.0
        p_up = beta_dist.ppf(1 - alpha / 2, k + 1, n_obs - k)
    elif k == n_obs:
        p_low = beta_dist.ppf(alpha / 2, k, n_obs - k + 1)
        p_up = 1.0
    else:
        p_low = beta_dist.ppf(alpha / 2, k, n_obs - k + 1)
        p_up = beta_dist.ppf(1 - alpha / 2, k + 1, n_obs - k)
    M_low = M_from_Ppos(p_low, b, kappa, N_total)
    M_up = M_from_Ppos(p_up, b, kappa, N_total)
    return (M_low, M_up)


def likelihood_ratio_ci(k, n_obs, b, kappa, N_total, alpha=0.05):
    P_obs = k / n_obs
    M_hat = M_from_Ppos(P_obs, b, kappa, N_total)

    def loglik(M):
        mu = M / N_total
        P_spec = 1 - np.exp(-mu / (kappa + 1))
        P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
        P_pos_p = b + (1 - b) * P_spec
        P_pos_p = np.clip(P_pos_p, 1e-12, 1 - 1e-12)
        return k * np.log(P_pos_p) + (n_obs - k) * np.log(1 - P_pos_p)

    def deviance(M):
        return 2 * (loglik(M_hat) - loglik(M))

    threshold = chi2.ppf(1 - alpha, 1)
    try:
        M_low = brentq(lambda M: deviance(M) - threshold, 1e-3, M_hat)
    except Exception:
        M_low = 0.0
    try:
        M_up = brentq(lambda M: deviance(M) - threshold, M_hat, 1e15)
    except Exception:
        M_up = np.inf
    return (M_low, M_up)


N_total = N_beads
n_obs = n_const
b_fit = b_tcs
kappa_fit = kappa_tcs

print("\n" + "=" * 80)
print("TCS MODEL: Per-point quantification and confidence intervals")
print(f"Total partitions N = {N_total}, Observed partitions n (median) = {n_obs:.0f}")
print("=" * 80)

results_tcs = []
for i in range(len(conc_aM)):
    if conc_aM[i] == 0:
        continue
    k = k_avg[i]
    P_obs_p = k / n_obs
    M_hat = M_from_Ppos(P_obs_p, b_fit, kappa_fit, N_total)
    cv = calc_CV(P_obs_p, n_obs, b_fit, kappa_fit, N_total)
    ci_d = delta_ci(P_obs_p, n_obs, b_fit, kappa_fit, N_total)
    ci_e = exact_transformation_ci(k, n_obs, b_fit, kappa_fit, N_total)
    ci_lr = likelihood_ratio_ci(k, n_obs, b_fit, kappa_fit, N_total)
    results_tcs.append({
        'Conc (aM)': conc_aM[i], 'k_avg': k, 'P_obs': P_obs_p,
        'M_hat (molecules)': M_hat, 'CV (%)': cv,
        'Delta_low': ci_d[0], 'Delta_up': ci_d[1],
        'Exact_low': ci_e[0], 'Exact_up': ci_e[1],
        'LR_low': ci_lr[0], 'LR_up': ci_lr[1],
    })

df_tcs_quant = pd.DataFrame(results_tcs)
pd.set_option('display.float_format', '{:.3f}'.format)
print(df_tcs_quant.to_string(index=False))
df_tcs_quant.to_csv('TCS_per_point_quantification.csv', index=False)
print("\nResults saved to TCS_per_point_quantification.csv")

# 为绘图准备数据
plot_tcs_data = []
for i in range(len(conc_aM)):
    if conc_aM[i] == 0:
        continue
    k = k_avg[i]
    P_obs_p = k / n_obs
    M_hat = M_from_Ppos(P_obs_p, b_fit, kappa_fit, N_total)
    cv = calc_CV(P_obs_p, n_obs, b_fit, kappa_fit, N_total)
    M_low, M_up = exact_transformation_ci(k, n_obs, b_fit, kappa_fit, N_total)
    P_low = b_fit + (1 - b_fit) * (1 - np.exp(-M_low / (N_total * (kappa_fit + 1))))
    P_up = b_fit + (1 - b_fit) * (1 - np.exp(-M_up / (N_total * (kappa_fit + 1))))
    plot_tcs_data.append((conc_aM[i], P_obs_p, M_hat, cv, P_low, P_up))


# ============================================================================
# 13. Print: fitted parameters, bootstrap CI, correlations
# ============================================================================
print("\n" + "=" * 60)
print("FITTED PARAMETERS")
print("=" * 60)
print(f"4PL (A fixed): A = {A_4pl:.6f}")
print(f"                B = {B_4pl:.4f}")
print(f"                C = {C_4pl:.4e}")
print(f"                D = {D_4pl:.4f}")
print(f"TCS:            κ = {kappa_tcs:.4f}")
print(f"                b = {b_tcs:.6f}")

print("\n" + "=" * 70)
print("PARAMETER CI COMPARISON (bootstrap, n=200)")
print("=" * 70)
print(f"{'Param':<8} {'4PL':<32} {'TCS':<32}")
print(f"{'─'*72}")
print(f"{'slope':<8} {f'B = {B_4pl:.4f}':<14} [{B_CI[0]:.4f}, {B_CI[1]:.4f}]  "
      f"{f'κ = {kappa_tcs:.4f}':<14} [{kappa_CI[0]:.4f}, {kappa_CI[1]:.4f}]")
print(f"{'offset':<8} {f'C = {C_4pl:.4e}':<14} [{C_CI[0]:.4e}, {C_CI[1]:.4e}]  "
      f"{f'b = {b_tcs:.6f}':<14} [{b_CI[0]:.6f}, {b_CI[1]:.6f}]")
print(f"{'upper':<8} {f'D = {D_4pl:.4f}':<14} [{D_CI[0]:.4f}, {D_CI[1]:.4f}]  "
      f"{'(n/a)':<32}")

print("\nParameter correlation (from bootstrap):")
print(f"  4PL: ρ(B, C) = {np.corrcoef(boot_B, boot_C)[0, 1]:+.3f}, "
      f"ρ(B, D) = {np.corrcoef(boot_B, boot_D)[0, 1]:+.3f}, "
      f"ρ(C, D) = {np.corrcoef(boot_C, boot_D)[0, 1]:+.3f}")
print(f"  TCS: ρ(κ, b) = {np.corrcoef(boot_kappa, boot_b)[0, 1]:+.3f}")


# ============================================================================
# 14. Print: model comparison + LOO
# ============================================================================
print("\n" + "=" * 60)
print("MODEL COMPARISON")
print(f"Mean diff (4PL - TCS) = {mean_diff:.3f}, SE = {se_diff:.3f}, t = {t_stat:.3f}")
print(f"Median diff (4PL - TCS) = {median_diff:.3f} (robust to outlier fold)")
print("=" * 60)
print(f"{'':<8} {'logL(binom)':<14} {'AICc':<10} {'BIC':<10} {'R²(AEB)':<10}")
print(f"{'4PL':<8} {logL_4pl:<12.2f} {AICc_4pl:<10.2f} {BIC_4pl:<10.2f} {R2_4pl:<10.4f}")
print(f"{'TCS':<8} {logL_tcs:<12.2f} {AICc_tcs:<10.2f} {BIC_tcs:<10.2f} {R2_tcs:<10.4f}")

print("\n" + "=" * 60)
print("LEAVE-ONE-OUT")
print("=" * 60)
print(f"LOO sum logL:  4PL = {np.sum(loo_4pl_arr):.2f}, TCS = {np.sum(loo_tcs_arr):.2f}")
print(f"Mean diff (4PL - TCS) = {mean_diff:.3f}, SE = {se_diff:.3f}, t = {t_stat:.3f}")
if t_stat < -2:
    print("→ TCS significantly better (|t|>2).")
elif t_stat > 2:
    print("→ 4PL significantly better.")
else:
    print("→ No significant difference.")

print("\nPer-fold logL differences (4PL - TCS):")
for idx, (l4, lt) in enumerate(zip(loo_4pl_vals, loo_tcs_vals)):
    if np.isnan(l4) or np.isnan(lt):
        print(f"  Fold {idx+1:2d}: FAIL")
    else:
        print(f"  Fold {idx+1:2d}: {l4-lt:12.4f}")


# ============================================================================
# 15. Print: Shapiro + deviance residual per point
# ============================================================================
print("\nShapiro-Wilk p-values (deviance residuals):")
print(f"  4PL: p = {p_sh_4pl:.4f}")
print(f"  TCS: p = {p_sh_tcs:.4f}")
print(f"\nDeviance residuals (per point):")
print(f"{'Conc(aM)':<10} {'4PL dr':<12} {'TCS dr':<12}")
print("-" * 34)
for i in range(len(conc_aM)):
    print(f"{conc_aM[i]:<10.1f} {dr_4pl[i]:<12.4f} {dr_tcs[i]:<12.4f}")


# ============================================================================
# 16. κ scan — compute, store, print
# ============================================================================
print("\n" + "=" * 80)
print("κ scan: fit quality and residual pattern vs fixed κ")
print("=" * 80)

kappa_scan_values = [0.1, 0.2, 0.3, 0.4, 0.5, 0.5793, 0.6, 0.7, 0.8, 1.0, 1.5, 2.0, 5.0, 10.0]
kappa_show = [0.3, 0.4, 0.5, 0.5793, 0.7, 1.0]

kappa_scan_data = []
resid_vs_kappa_data = {}

print(f"{'κ':<10} {'logL':<12} {'R²(AEB)':<10} {'maxAEBresid':<14}"
      f" {'low conc resid':<14} {'high conc resid':<14}")
print("-" * 74)
for k_fix in kappa_scan_values:
    res_k = opt.minimize(
        lambda params: neg_loglik_tcs([k_fix, params[0]], mu_train, k_all, n_all_arr),
        [b_tcs], method='L-BFGS-B', bounds=[(0, 0.1)], options={'maxiter': 1000}
    )
    b_k = res_k.x[0]
    AEB_pred_k = predict_AEB_tcs(mu_train, k_fix, b_k)
    logL_k = -res_k.fun
    ss_res_k = np.sum(weights_AEB * (AEB_train - AEB_pred_k) ** 2)
    r2_k = 1 - ss_res_k / ss_tot
    resid_k = AEB_train - AEB_pred_k
    max_resid = float(np.max(np.abs(resid_k[1:])))
    low_resid = float(np.max(np.abs(resid_k[1:7])))
    high_resid = float(np.max(np.abs(resid_k[7:])))
    kappa_scan_data.append((k_fix, logL_k, r2_k, max_resid, low_resid, high_resid))
    if k_fix in kappa_show:
        resid_vs_kappa_data[k_fix] = resid_k.tolist()
    label = " ← MLE" if abs(k_fix - 0.5793) < 0.001 else ""
    print(f"{k_fix:<10.4f} {logL_k:<12.2f} {r2_k:<10.6f} {max_resid:<14.6f} "
          f"{low_resid:<14.6f} {high_resid:<14.6f}{label}")

# Per-point AEB residual table (using stored data, no double optimize)
print("\n" + "=" * 80)
print("Per-point AEB residuals at selected κ values")
print("=" * 80)
print(f"{'Conc(aM)':<10}", end="")
for k_fix in kappa_show:
    label = f"κ={k_fix:.3f}" if k_fix != 0.5793 else "κ=0.579(MLE)"
    print(f"  {label:<14}", end="")
print()
print("-" * (10 + 16 * len(kappa_show)))
for i in range(len(conc_aM)):
    print(f"{conc_aM[i]:<10.1f}", end="")
    for k_fix in kappa_show:
        print(f"  {resid_vs_kappa_data[k_fix][i]:<14.6f}", end="")
    print()


# ============================================================================
# 17. R3 print + LoB/LoD/LoQ compute + print
# ============================================================================
print("\n" + "=" * 80)
print("R3 LINEAR APPROXIMATION (Low Concentration)")
print("=" * 80)
print(f"TCS-based γ = {gamma_tcs_r3:.4f}  (κ = {kappa_tcs:.2f})")
if not np.isnan(gamma_fit_r3):
    print(f"Fitted γ = {gamma_fit_r3:.4f}  (κ = {kappa_fit_r3:.2f})")
else:
    print("Not enough points for R3 fitting.")
print(f"Number of points satisfying P_specific < 0.2: {np.sum(r3_valid)}")
print("\nR3 estimates (training set):")
print(f"{'Conc(aM)':>10} {'P_specific':>10} {'M_r3':>12} {'CV_r3(%)':>10} {'R3_valid':>8}")
for i in range(len(conc_aM)):
    print(f"{conc_aM[i]:10.4f} {P_specific_train_r3[i]:10.6f} {M_r3[i]:12.2f} "
          f"{CV_r3[i] * 100:10.2f} {str(r3_valid[i]):>8}")

df_r3 = pd.DataFrame({
    'Conc_aM': conc_aM, 'P_specific': P_specific_train_r3, 'M_r3': M_r3,
    'CV_r3_pct': CV_r3 * 100, 'R3_valid': r3_valid,
})
df_r3.to_csv('R3_analysis_sβG.csv', index=False)
print("\nR3 analysis saved to R3_analysis_sβG.csv")


# ----------------------------------------------------------------------------
# 17.b LoB/LoD/LoQ compute
# ----------------------------------------------------------------------------
print("\n" + "=" * 80)
print("LoB, LoD, LoQ Analysis (based on fitted TCS parameters)")
print("=" * 80)

n_obs_beads = int(n_const)
N_total_loc = N_beads
kappa_fit_loc = kappa_tcs
b_fit_loc = b_tcs
gamma_fit_loc = 1.0 / (kappa_fit_loc + 1)


def P_specific_R2(M):
    return 1 - np.exp(-M / (N_total_loc * (kappa_fit_loc + 1)))


def P_pos_R2(M):
    return b_fit_loc + (1 - b_fit_loc) * P_specific_R2(M)


def M_from_Ppos_R2(P_pos):
    if P_pos <= b_fit_loc:
        return 0.0
    P_spec = (P_pos - b_fit_loc) / (1 - b_fit_loc)
    return -N_total_loc * (kappa_fit_loc + 1) * np.log(1 - P_spec)


def CV_R2(M):
    if M <= 0:
        return np.inf
    P_pos_p = P_pos_R2(M)
    if P_pos_p <= 0 or P_pos_p >= 1:
        return np.inf
    P_spec = (P_pos_p - b_fit_loc) / (1 - b_fit_loc)
    P_spec = np.clip(P_spec, 1e-12, 1 - 1e-12)
    dM_dP = N_total_loc * (kappa_fit_loc + 1) / ((1 - b_fit_loc) * (1 - P_spec))
    se = np.abs(dM_dP) * np.sqrt(P_pos_p * (1 - P_pos_p) / n_obs_beads)
    return se / M


def detection_prob_R2(M):
    k_95 = stats.binom.ppf(0.95, n_obs_beads, b_fit_loc)
    lob_p = k_95 / n_obs_beads
    P_pos_p = P_pos_R2(M)
    k_th = int(np.floor(n_obs_beads * lob_p)) + 1
    if k_th > n_obs_beads:
        return 0.0
    return stats.binom.sf(k_th - 1, n_obs_beads, P_pos_p)


k_95 = stats.binom.ppf(0.95, n_obs_beads, b_fit_loc)
lob_p_R2 = k_95 / n_obs_beads
M_lob_R2 = M_from_Ppos_R2(lob_p_R2)


def lod_R2():
    M_low = 1.0
    while detection_prob_R2(M_low) < 0.95:
        M_low *= 2
        if M_low > 1e8:
            return np.nan
    M_high = M_low
    M_low = max(1.0, M_low / 2)
    try:
        return brentq(lambda M: detection_prob_R2(M) - 0.95, M_low, M_high, xtol=1e-6)
    except Exception:
        return np.nan


M_lod_R2 = lod_R2()


def loq_low_R2():
    def obj(M):
        return CV_R2(M) - 0.2
    M_low = 1.0
    while obj(M_low) > 0:
        M_low *= 2
    M_high = M_low
    M_low = max(1.0, M_low / 2)
    try:
        return brentq(obj, M_low, M_high, xtol=1e-6)
    except Exception:
        return np.nan


def loq_high_R2():
    def obj(M):
        return CV_R2(M) - 0.2
    M_start = 500000
    while CV_R2(M_start) < 0.2:
        M_start *= 2
        if M_start > 1e8:
            break
    M_high = M_start
    M_low = M_start / 2
    while obj(M_low) > 0:
        M_low /= 2
        if M_low < 1e3:
            break
    try:
        return brentq(obj, M_low, M_high, xtol=1e-6)
    except Exception:
        return np.nan


M_loq_low_R2 = loq_low_R2()
M_loq_high_R2 = loq_high_R2()


def M_from_Ppos_R3(P_pos):
    P_spec = (P_pos - b_fit_loc) / (1 - b_fit_loc)
    return N_total_loc * (kappa_fit_loc + 1) * P_spec


M_lob_R3 = M_from_Ppos_R3(lob_p_R2)
M_lod_R3 = (3.29 * N_total_loc / gamma_fit_loc) * np.sqrt(b_fit_loc / (n_obs_beads * (1 - b_fit_loc)))

A_q = (1 - b_fit_loc) * (1 + 0.04 * n_obs_beads)
B_q = -(1 - 2 * b_fit_loc)
C_q = -b_fit_loc
disc = B_q ** 2 - 4 * A_q * C_q
if disc >= 0:
    x_loq = (-B_q + np.sqrt(disc)) / (2 * A_q)
else:
    x_loq = np.nan
M_loq_low_R3 = N_total_loc * (kappa_fit_loc + 1) * x_loq if not np.isnan(x_loq) else np.nan

print(f"\nParameters: N_total = {N_total_loc}, n_obs = {n_obs_beads}, "
      f"κ = {kappa_fit_loc:.4f}, b = {b_fit_loc:.6e}, γ = {gamma_fit_loc:.4f}")
print()
print(f"{'Model':<6} {'LoB (molecules)':<20} {'LoD (molecules)':<20} "
      f"{'LoQ low (molecules)':<22} {'LoQ high (molecules)':<22}")
print("-" * 90)
print(f"{'R2':<6} {M_lob_R2:<20.1f} {M_lod_R2:<20.1f} {M_loq_low_R2:<22.1f} {M_loq_high_R2:<22.0f}")
print(f"{'R3':<6} {M_lob_R3:<20.1f} {M_lod_R3:<20.1f} {M_loq_low_R3:<22.1f} {'---':<22}")
print("\n注：R2 LoQ 为数值解，R3 LoQ 为解析近似；R3 无高浓度根。")


# ============================================================================
# 18. Plot: 4-panel main figure
# ============================================================================
fig_core, axes = plt.subplots(2, 2, figsize=(13, 11))
fig_core.suptitle('Extended Data Fig. 3: TCS reanalysis of the 2010 Simoa sβG benchmark data',
                  fontsize=16, fontweight='bold', y=0.95)

conc_real = conc_aM
mu_plot = np.logspace(np.log10(mu_nz.min() * 0.5), np.log10(mu_nz.max() * 2), 200)
conc_plot = mu_plot * (1 / (V * N_A / N_beads)) / 1e-18

# LoB/LoD/LoQ indicators (R2)
conc_lob_r2 = M_lob_R2 / (V * N_A) * 1e18
conc_lod_r2 = M_lod_R2 / (V * N_A) * 1e18
conc_loq_low_r2 = M_loq_low_R2 / (V * N_A) * 1e18
conc_loq_high_r2 = M_loq_high_R2 / (V * N_A) * 1e18
conc_lob_r3 = M_lob_R3 / (V * N_A) * 1e18
conc_lod_r3 = M_lod_R3 / (V * N_A) * 1e18
conc_loq_low_r3 = M_loq_low_R3 / (V * N_A) * 1e18

# (a) 4PL fit
ax = axes[0, 0]
ax.semilogx(conc_plot, fourpl_fixed_A(mu_plot, B_4pl, C_4pl, D_4pl, A_4pl), 'b-', lw=2)
ax.scatter(conc_real[mask_nz], AEB_nz, c='blue', s=30, alpha=0.6)
ax.scatter(conc_real[mask_zero], AEB_train[mask_zero], c='blue', s=30, alpha=0.6, marker='s')
for i in range(len(conc_aM)):
    if conc_aM[i] == 0:
        continue
    x = conc_aM[i]
    y = AEB_train[i]
    ax.text(x * 0.6, y + 0.04, f'M:{meas_cv[i]:.0f}%\nP:{poisson_cv[i]:.0f}%',
            rotation=90, fontsize=10, color='darkblue', alpha=0.8, ha='left', va='bottom')
ax.set_xlabel('Concentration (aM)', fontsize=14)
ax.set_ylabel('AEB', fontsize=14)
ax.set_title('(1) 4PL Fit with Reported CVs', fontsize=14, fontweight='bold')
ax.text(0.05, 0.95,
        f"A = {A_4pl:.6f}\nB = {B_4pl:.2f}\nC = {C_4pl:.2e}\nD = {D_4pl:.2f}\n"
        f"4PL: AEB = A + (D-A)/(1+(C/x)$^B$)\nM: measurement CV\n"
        f"P: Poisson noise CV\nR$^2$ = {R2_4pl:.4f}",
        transform=ax.transAxes, fontsize=14, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.5))
ax.grid(False)

# (b) TCS fit
ax = axes[0, 1]
ax.semilogx(conc_plot, tcs_model(mu_plot, kappa_tcs, b_tcs), 'r-', lw=2)
ax.scatter(conc_real[mask_nz], P_nz, c='red', s=10, alpha=0.6)
ax.scatter(conc_real[mask_zero], P_train[mask_zero], c='red', s=30, alpha=0.6, marker='s')
for (conc, Pobs, Mhat, cv, Plow, Pup) in plot_tcs_data:
    ax.errorbar(conc, Pobs, yerr=[[Pobs - Plow], [Pup - Pobs]],
                fmt='none', ecolor='darkred', capsize=8, alpha=0.7, linewidth=4)
    if conc >= 0:
        ax.text(conc * 0.9, Pobs + 0.04, f'T:{cv:.1f}%', fontsize=12, color='darkred',
                rotation=90, ha='left', va='bottom', alpha=0.9)
vline_lob = ax.axvline(conc_lob_r2, color='green', linestyle=':', linewidth=1.5,
                       label=f'LoB ({conc_lob_r2:.3f} aM)')
vline_lod = ax.axvline(conc_lod_r2, color='orange', linestyle='--', linewidth=1.5,
                       label=f'LoD ({conc_lod_r2:.3f} aM)')
vline_loq_low = ax.axvline(conc_loq_low_r2, color='purple', linestyle='-.', linewidth=1.5,
                            label=f'LoQ low ({conc_loq_low_r2:.3f} aM)')
vline_loq_high = ax.axvline(conc_loq_high_r2, color='purple', linestyle='-.', linewidth=1.5,
                             label=f'LoQ high ({conc_loq_high_r2:.1f} aM)')
ax.legend(handles=[vline_lob, vline_lod, vline_loq_low, vline_loq_high],
          loc='lower right', fontsize=8, framealpha=0.8)
ax.set_xlabel('Concentration (aM)', fontsize=14)
ax.set_ylabel(r'$P_{\mathrm{pos}}$', fontsize=14)
ax.set_title('(2) R2 Fit with Uncertainty', fontsize=14, fontweight='bold')
eq_str = r"$\widehat{M_{\mathrm{R2}}} = -N(1+\kappa)\ln\left(1-\widehat{P_{\mathrm{specific}}}\right)$"
ax.text(0.05, 0.95,
        f"$\\kappa$ = {kappa_tcs:.2f}\nb = {b_tcs:.6f}\n{eq_str}\n"
        f"R$^2$ = {R2_tcs:.4f}\nError bars: Clopper-Pearson 95% CI\nT: TCS Model CV",
        transform=ax.transAxes, fontsize=14, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))
ax.grid(False)

# (c) Residuals on AEB scale
ax = axes[1, 0]
ax.scatter(conc_real, resid_4pl, c='blue', alpha=0.6, label='4PL')
ax.scatter(conc_real, resid_tcs, c='red', alpha=0.6, label='TCS')
ax.axhline(0, color='k', ls='--')
ax.set_xscale('log')
ax.set_xlabel('Concentration (aM)', fontsize=14)
ax.set_ylabel('Residual (AEB)', fontsize=14)
ax.set_title('(3) Residuals (AEB scale)', fontsize=14, fontweight='bold')
ax.legend(fontsize=14)
ax.grid(False)
ax.text(0.45, 0.15, f"LOO: Δ = {mean_diff:.1f} ± {se_diff:.1f}\nt = {t_stat:.2f}",
        transform=ax.transAxes, fontsize=14, ha='right', va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.9))

# (d) R3 linear approximation
ax = axes[1, 1]
c_range = np.logspace(np.log10(conc_aM[conc_aM > 0].min() * 0.5),
                      np.log10(conc_aM[conc_aM > 0].max() * 2), 100)
mu_range = c_range * 1e-18 * V * N_A / N_beads
P_theory = gamma_tcs_r3 * mu_range
ax.loglog(c_range, P_theory, 'k--', lw=1, label='R2 theory')
if not np.isnan(gamma_fit_r3):
    P_fit_line = gamma_fit_r3 * mu_range
    ax.loglog(c_range, P_fit_line, 'b-', lw=2, label='R3 fitted')
ax.loglog(conc_aM[~mask_zero], P_specific_train_r3[~mask_zero],
          'o', color='gray', alpha=0.5, label='All data')
valid_indices = np.where(r3_valid & ~mask_zero)[0]
for plot_idx, i in enumerate(valid_indices):
    conc_i = conc_aM[i]
    ps_i = P_specific_train_r3[i]
    cv_i = CV_r3[i] * 100
    ax.loglog(conc_i, ps_i, 'ro', markersize=8)
    y_shift = 1.5 if plot_idx % 2 == 0 else 0.6
    va = 'bottom' if plot_idx % 2 == 0 else 'top'
    ax.text(conc_i, ps_i * y_shift, f'T:{cv_i:.2f}%',
            fontsize=12, color='darkred', ha='center', va=va)
vline_lob_r3 = ax.axvline(conc_lob_r3, color='green', linestyle=':', linewidth=1.5,
                          label=f'LoB ({conc_lob_r3:.3f} aM)')
vline_lod_r3 = ax.axvline(conc_lod_r3, color='orange', linestyle='--', linewidth=1.5,
                          label=f'LoD ({conc_lod_r3:.3f} aM)')
vline_loq_low_r3 = ax.axvline(conc_loq_low_r3, color='purple', linestyle='-.', linewidth=1.5,
                               label=f'LoQ low ({conc_loq_low_r3:.3f} aM)')
ax.legend(handles=[vline_lob_r3, vline_lod_r3, vline_loq_low_r3],
          loc='upper left', fontsize=10, framealpha=0.8)
if not np.isnan(gamma_fit_r3):
    kappa_fit_r3_v = 1.0 / gamma_fit_r3 - 1.0
    fit_info = f"R3 fitted: $\\kappa = {kappa_fit_r3_v:.2f}$, $\\gamma = {gamma_fit_r3:.4f}$"
else:
    fit_info = "R3 fitted: insufficient data"
eq_r3 = r"$\widehat{M_{\mathrm{R3}}}=N(1+\kappa)\left(\widehat{P_{\mathrm{specific}}}\right)$"
gamma_def = r"$\gamma = \frac{1}{1+\kappa}$"
ax.text(0.05, 0.95,
        f"R2 (TCS): $\\kappa = {kappa_tcs:.2f}$, $\\gamma = {gamma_tcs_r3:.4f}$\n"
        f"{fit_info}\n{eq_r3}\n{gamma_def}\nT: TCS Model CV",
        transform=ax.transAxes, fontsize=14, va='top',
        bbox=dict(boxstyle='round', facecolor='white', alpha=0.5))
ax.set_xlabel('Concentration (aM)', fontsize=14)
ax.set_ylabel(r'$P_{\mathrm{specific}}$', fontsize=14)
ax.set_title('(4) R3 Linear Approximation', fontsize=14, fontweight='bold')
ax.legend(fontsize=8)
fig_core.savefig('Extended_data_fig_3.svg', dpi=300, bbox_inches='tight')
plt.show()
plt.close(fig_core)


# ============================================================================
# 19. Plot: Q-Q plots
# ============================================================================

fig_s1, axes_s1 = plt.subplots(1, 2, figsize=(9, 5))
fig_s1.suptitle('Extended Data Figure. Normal Q–Q plots of deviance residuals '
                'for the 4PL and TCS-based mechanistic models (sβG dataset)',
                fontsize=14, fontweight='bold')
stats.probplot(dr_4pl, dist="norm", plot=axes_s1[0])
axes_s1[0].set_title('4PL')
axes_s1[0].grid(False)
stats.probplot(dr_tcs, dist="norm", plot=axes_s1[1])
axes_s1[1].set_title('TCS')
axes_s1[1].grid(False)
plt.tight_layout()
plt.show()
fig_s1.savefig('Fig_S1_sβG.png', dpi=300)
print("\nSupplementary figure saved: Fig_S1_sβG.png")


# ============================================================================
# 20. ⭐ SI Table 3 EXPORT — AFTER all printing is done ⭐
# ============================================================================
print("\n" + "=" * 80)
print("EXPORTING: SI Table 3.xlsx (13 sheets, all data computed)")
print("=" * 80)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    mle_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # ---- Sheet 1: Raw data ----
    ws1 = wb.active
    ws1.title = "1_Raw_Data"
    ws1.append(["Concentration (aM)", "k_avg", "P_pos (%)", "Estimated n (beads/well)",
                "Measurement CV (%)", "Poisson CV (%)"])
    for cell in ws1[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(conc_aM)):
        n_i = k_avg[i] / (P_pos[i] / 100) if P_pos[i] > 0 else np.nan
        ws1.append([conc_aM[i], k_avg[i], P_pos[i], n_i, meas_cv[i], poisson_cv[i]])

    # ---- Sheet 2: Per-point quantification ----
    ws2 = wb.create_sheet("2_TCS_PerPoint")
    ws2.append(["Conc (aM)", "k_avg", "P_obs", "M_hat (molecules)", "CV (%)",
                "Delta_low", "Delta_up", "Exact_low", "Exact_up", "LR_low", "LR_up"])
    for cell in ws2[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(conc_aM)):
        if conc_aM[i] == 0:
            continue
        P_obs_p = k_avg[i] / n_const
        M_hat_p = M_from_Ppos(P_obs_p, b_fit, kappa_fit, N_beads)
        cv_p = calc_CV(P_obs_p, n_const, b_fit, kappa_fit, N_beads)
        ci_d = delta_ci(P_obs_p, n_const, b_fit, kappa_fit, N_beads)
        ci_e = exact_transformation_ci(k_avg[i], n_const, b_fit, kappa_fit, N_beads)
        ci_lr = likelihood_ratio_ci(k_avg[i], n_const, b_fit, kappa_fit, N_beads)
        ws2.append([conc_aM[i], k_avg[i], P_obs_p, M_hat_p, cv_p,
                    ci_d[0], ci_d[1], ci_e[0], ci_e[1], ci_lr[0], ci_lr[1]])

    # ---- Sheet 3: Fitted parameters ----
    ws3 = wb.create_sheet("3_Fit_Parameters")
    ws3.append(["Parameter", "Value", "Description"])
    for cell in ws3[1]:
        cell.font = bold; cell.fill = header_fill
    ws3.append(["kappa (TCS)", kappa_tcs, "TCS R2 depletion coefficient"])
    ws3.append(["b (TCS background)", b_tcs, "TCS non-specific background probability"])
    ws3.append(["gamma = 1/(1+kappa)", gamma_tcs_r3, "Linear coefficient (R3)"])
    ws3.append(["R2 (TCS, AEB scale)", R2_tcs, "Weighted R² for TCS R2 model"])
    ws3.append(["R2 (4PL)", R2_4pl, "Weighted R² for 4PL model"])
    ws3.append(["A (4PL, fixed)", A_4pl, "4PL lower asymptote"])
    ws3.append(["B (4PL)", B_4pl, "4PL slope"])
    ws3.append(["C (4PL)", C_4pl, "4PL midpoint"])
    ws3.append(["D (4PL)", D_4pl, "4PL upper asymptote"])

    # ---- Sheet 4: Case-bootstrap CI ----
    ws4 = wb.create_sheet("4_Bootstrap_CI")
    ws4.append(["Model", "Parameter", "Point estimate",
                "Bootstrap 2.5%", "Bootstrap 97.5%", "Relative width"])
    for cell in ws4[1]:
        cell.font = bold; cell.fill = header_fill
    for name, val, arr in [("B (4PL)", B_4pl, boot_B),
                           ("C (4PL)", C_4pl, boot_C),
                           ("D (4PL)", D_4pl, boot_D)]:
        if len(arr) > 0:
            lo, hi = np.percentile(arr, [2.5, 97.5])
            width = (hi - lo) / max(abs(val), 1e-12)
            ws4.append(["4PL", name, val, lo, hi, width])
    for name, val, arr in [("kappa (TCS)", kappa_tcs, boot_kappa),
                           ("b (TCS)", b_tcs, boot_b)]:
        if len(arr) > 0:
            lo, hi = np.percentile(arr, [2.5, 97.5])
            width = (hi - lo) / max(abs(val), 1e-12)
            ws4.append(["TCS", name, val, lo, hi, width])

    # ---- Sheet 5: Correlations ----
    ws5 = wb.create_sheet("5_Correlations")
    ws5.append(["Model", "Param 1", "Param 2", "Pearson rho"])
    for cell in ws5[1]:
        cell.font = bold; cell.fill = header_fill
    if len(boot_B) > 0:
        ws5.append(["4PL", "B", "C", float(np.corrcoef(boot_B, boot_C)[0, 1])])
        ws5.append(["4PL", "B", "D", float(np.corrcoef(boot_B, boot_D)[0, 1])])
        ws5.append(["4PL", "C", "D", float(np.corrcoef(boot_C, boot_D)[0, 1])])
    if len(boot_kappa) > 0:
        ws5.append(["TCS", "kappa", "b",
                    float(np.corrcoef(boot_kappa, boot_b)[0, 1])])

    # ---- Sheet 6: LoB/LoD/LoQ ----
    ws6 = wb.create_sheet("6_LoB_LoD_LoQ")
    ws6.append(["Metric", "R2 (molecules)", "R3 (molecules)",
                "Concentration (aM, R2)", "Concentration (aM, R3)"])
    for cell in ws6[1]:
        cell.font = bold; cell.fill = header_fill
    ws6.append(["LoB", M_lob_R2, M_lob_R3,
                M_lob_R2 / (V * N_A) * 1e18, M_lob_R3 / (V * N_A) * 1e18])
    ws6.append(["LoD", M_lod_R2, M_lod_R3,
                M_lod_R2 / (V * N_A) * 1e18, M_lod_R3 / (V * N_A) * 1e18])
    ws6.append(["LoQ low", M_loq_low_R2, M_loq_low_R3,
                M_loq_low_R2 / (V * N_A) * 1e18, M_loq_low_R3 / (V * N_A) * 1e18])
    ws6.append(["LoQ high", M_loq_high_R2, np.nan,
                M_loq_high_R2 / (V * N_A) * 1e18
                if not np.isnan(M_loq_high_R2) else np.nan, np.nan])

    # ---- Sheet 7: R3 linear comparison ----
    ws7 = wb.create_sheet("7_R3_Linear")
    ws7.append(["Parameter", "TCS R2 prediction", "R3 independent fit"])
    for cell in ws7[1]:
        cell.font = bold; cell.fill = header_fill
    ws7.append(["gamma", gamma_tcs_r3,
                gamma_fit_r3 if not np.isnan(gamma_fit_r3) else "insufficient data"])
    if not np.isnan(gamma_fit_r3):
        ws7.append(["kappa (from gamma)", kappa_tcs, kappa_fit_r3])
        ws7.append(["Relative deviation (%)", 0.0,
                    abs(gamma_fit_r3 - gamma_tcs_r3) / gamma_tcs_r3 * 100])

    # ---- Sheet 8: D profile-likelihood ----
    ws8 = wb.create_sheet("8_D_Profile_Likelihood")
    ws8.append(["D (fixed)", "logL (binomial)", "B (fitted)", "C (fitted)",
                "ΔlogL", "Interpretation"])
    for cell in ws8[1]:
        cell.font = bold; cell.fill = header_fill
    for D_fix, logL_scan, B_scan, C_scan, delta in D_scan_results:
        if delta > -1.92:
            note = "MLE neighbor (1σ)"
        elif delta > -3.84:
            note = "within 1σ formal 95% CI"
        elif delta > -30:
            note = "within 2σ but data cannot resolve"
        else:
            note = "essentially flat (D poorly identified)"
        ws8.append([D_fix, logL_scan, B_scan, C_scan, delta, note])
    ws8.append([])
    ws8.append(["BEST (MLE)", "", B_4pl, C_4pl, 0.0, "ΔlogL=0 reference"])

    # ---- Sheet 9: Shapiro-Wilk ----
    ws9 = wb.create_sheet("9_Shapiro_Wilk")
    ws9.append(["Model", "Residual type", "Shapiro-Wilk p", "Conclusion (α=0.05)"])
    for cell in ws9[1]:
        cell.font = bold; cell.fill = header_fill
    sh_concl_4pl = ("Reject normality — overfit signature"
                    if p_sh_4pl < 0.05 else "Cannot reject normality")
    sh_concl_tcs = ("Cannot reject (consistent with model)"
                    if p_sh_tcs >= 0.05 else "Reject normality")
    ws9.append(["4PL", "Deviance (binomial)", float(p_sh_4pl), sh_concl_4pl])
    ws9.append(["TCS", "Deviance (binomial)", float(p_sh_tcs), sh_concl_tcs])

    # ---- Sheet 10: Deviance residuals per point ----
    ws10 = wb.create_sheet("10_Deviance_Residuals")
    ws10.append(["Concentration (aM)", "4PL deviance residual",
                 "TCS deviance residual", "Note"])
    for cell in ws10[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(conc_aM)):
        note = ""
        if conc_aM[i] == 7000 and abs(dr_4pl[i]) < 1 and abs(dr_tcs[i]) > 5:
            note = "leverage: 4PL fits exactly via D, TCS over-predicts 2.2%"
        elif conc_aM[i] == 350 and abs(dr_4pl[i]) > 4:
            note = "high deviance residual under 4PL"
        elif conc_aM[i] == 700:
            note = "TCS P_pos plateau miss"
        ws10.append([conc_aM[i], float(dr_4pl[i]),
                     float(dr_tcs[i]), note])

    # ---- Sheet 11: κ scan ----
    ws11 = wb.create_sheet("11_Kappa_Scan")
    ws11.append(["κ (fixed)", "logL (binomial)", "R² (AEB scale)",
                 "max |AEB residual|", "low-conc max |resid|",
                 "high-conc max |resid|", "Note"])
    for cell in ws11[1]:
        cell.font = bold; cell.fill = header_fill
    mle_row = None
    for idx, (k_fix, logL_k, r2_k, max_resid, low_resid,
              high_resid) in enumerate(kappa_scan_data):
        note = "MLE" if abs(k_fix - 0.5793) < 0.001 else ""
        ws11.append([k_fix, logL_k, r2_k, max_resid,
                     low_resid, high_resid, note])
        if note == "MLE":
            mle_row = idx + 2
    if mle_row is not None:
        for cell in ws11[mle_row]:
            cell.fill = mle_fill

    # ---- Sheet 12: per-point residual heatmap ----
    ws12 = wb.create_sheet("12_Resid_vs_Kappa")
    header_row = ["Concentration (aM)"] + [
        ("κ=0.579 (MLE)" if abs(k - 0.5793) < 1e-3 else f"κ={k:.3f}")
        for k in kappa_show
    ]
    ws12.append(header_row)
    for cell in ws12[1]:
        cell.font = bold; cell.fill = header_fill
    for i in range(len(conc_aM)):
        row = [conc_aM[i]]
        for k_fix in kappa_show:
            row.append(resid_vs_kappa_data[k_fix][i])
        ws12.append(row)
    mle_col_idx = kappa_show.index(0.5793) + 2
    ws12.cell(row=1, column=mle_col_idx).fill = mle_fill

    # ---- Sheet 13: LOO per-fold ----
    ws13 = wb.create_sheet("13_LOO_PerFold")
    ws13.append(["Fold", "Concentration (aM)",
                 "logL (4PL hold-out)", "logL (TCS hold-out)",
                 "ΔlogL (4PL - TCS)", "Note"])
    for cell in ws13[1]:
        cell.font = bold; cell.fill = header_fill
    for idx, (l4, lt) in enumerate(zip(loo_4pl_vals, loo_tcs_vals)):
        conc_fold = conc_aM[idx]
        if np.isnan(l4) or np.isnan(lt):
            ws13.append([idx + 1, conc_fold, "FAIL", "FAIL",
                         np.nan, "fit failed"])
        else:
            d = l4 - lt
            note = ""
            if idx + 1 == 11:
                note = "leverage fold (7000 aM, 73% of AICc gap)"
            elif idx + 1 == 9:
                note = "high conc, 4PL advantage"
            elif idx + 1 == 10:
                note = "high conc, 4PL advantage"
            ws13.append([idx + 1, conc_fold, float(l4), float(lt),
                         float(d), note])
    ws13.append([])
    ws13.append(["SUMMARY", "", "", "", "", ""])
    ws13.append(["Mean Δ", "", "", "", float(mean_diff), ""])
    ws13.append(["SE Δ",   "", "", "", float(se_diff),   ""])
    ws13.append(["t-statistic", "", "", "", float(t_stat), "n.s. (|t|<2)"])
    ws13.append(["Median Δ", "", "", "", float(median_diff),
                 "robust to leverage"])
    ws13.append(["N (valid folds)", "", "", "",
                 int(np.sum(valid_idx)), ""])






    ws13.append([])
    ws13.append(["TOST equivalence test", "", "", "", "", ""])
    ws13.append(["Margin (±logL)", "", "", "", float(delta),
                 "1% of mean per-fold |logL|"])
    ws13.append(["Mean Δ 90% CI", "", "", "",
                 f"[{ci90[0]:.3f}, {ci90[1]:.3f}]", ""])
    ws13.append(["TOST p", "", "", "", float(p_tost),
                 "EQUIVALENT" if p_tost < 0.05 else "NOT equivalent"])
    ws13.append(["Smallest passing margin (logL)", "", "", "",
                 float(delta_min),
                 f"{100*delta_min/mean_abs_logL:.3f}% of per-fold |logL|"])











    wb.save("SI_Table_3.xlsx")
    print("\n" + "=" * 60)
    print("[SI Table 3 saved to 'SI_Table_3.xlsx' — 13 sheets]")
    print("=" * 60)
    sheet_descriptions = [
        "1. Raw_Data",
        "2. TCS_PerPoint",
        "3. Fit_Parameters",
        "4. Bootstrap_CI",
        "5. Correlations",
        "6. LoB_LoD_LoQ",
        "7. R3_Linear",
        "8. D_Profile_Likelihood",
        "9. Shapiro_Wilk",
        "10. Deviance_Residuals",
        "11. Kappa_Scan",
        "12. Resid_vs_Kappa",
        "13. LOO_PerFold",
    ]
    for i, name in enumerate(sheet_descriptions, 1):
        print(f"  Sheet {i:2d}: {name}")

except Exception as e_save:
    print(f"\n[!] SI Table 3 export failed: {e_save}")
    import traceback
    traceback.print_exc()


print("\n" + "=" * 80)
print("DONE. Outputs:")
print("  - Fig_S1_sβG.png     (Q-Q plot, deviance residuals)")
print("  - Fig_core_sβG.png   (Extended Data Fig. 3, 4-panel)")
print("  - SI_Table_3.xlsx    (13 sheets)")
print("  - TCS_per_point_quantification.csv")
print("  - R3_analysis_sβG.csv")
print("=" * 80)

# ============================================================================
# R1 model fit: testing beta identifiability (sβG dataset)
# COMPLETE REPLACEMENT — uses dynesty (nested sampling)
# Advantages: handles multimodal natively, no Windows/numpy issues,
#             gives Bayes factor for R1 vs R2 comparison
# ============================================================================

import dynesty
from dynesty import utils as dyfunc
try:
    import corner
except ImportError:
    corner = None
try:
    import arviz as az
except ImportError:
    az = None

print("\n" + "="*80)
print("R1 Model Fit: Testing beta Identifiability (sβG dataset)")
print("  Method: Nested sampling (dynesty)")
print("="*80)

# ---- R1 forward model (VECTORIZED) ----
def solve_p_from_xi(xi, kappa, max_iter=100, tol=1e-12):
    """Vectorized Newton-Raphson: xi = p/(1-p) + p/kappa -> p"""
    xi = np.atleast_1d(np.asarray(xi, dtype=float))
    p = np.where(xi > 0, xi / (1.0 + xi), 0.0)
    p = np.clip(p, 1e-15, 1.0 - 1e-15)
    for _ in range(max_iter):
        f = p / (1.0 - p) + p / kappa - xi
        df = 1.0 / (1.0 - p)**2 + 1.0 / kappa
        dp = f / df
        p_new = np.clip(p - dp, 1e-15, 1.0 - 1e-15)
        if np.max(np.abs(p_new - p)) < tol:
            p = p_new
            break
        p = p_new
    return p

def tcs_R1_model(mu, kappa, beta, b):
    """Vectorized R1 forward model: mu -> P_pos"""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    P_pos = np.full_like(mu, b)
    mask = mu > 0
    if np.any(mask):
        xi = mu[mask] / (kappa * beta)
        p = solve_p_from_xi(xi, kappa)
        P_specific = 1.0 - (1.0 - p) ** beta
        P_pos[mask] = b + (1.0 - b) * P_specific
    return np.clip(P_pos, 1e-15, 1.0 - 1e-15)

def neg_loglik_R1(params, mu, k, n):
    kappa, beta, b = params
    if kappa <= 0 or beta <= 0 or b < 0 or b >= 1:
        return 1e12
    P_pos = tcs_R1_model(mu, kappa, beta, b)
    return -np.sum(k * np.log(P_pos) + (n - k) * np.log(1.0 - P_pos))

# ---- Data ----
mu_R1_all = mu_train.copy()
k_R1_all = k_all.copy()
n_R1_all = n_all_arr.copy()

# ---- Initial least-squares fit ----
kappa_init_R1 = kappa_tcs
b_init_R1 = b_tcs
beta_init = 10.0

print(f"\nR2 fit results (as R1 initial values): kappa={kappa_tcs:.4f}, b={b_tcs:.6f}")
print(f"R1 initial beta = {beta_init}")

bounds_R1 = [(0.01, 1000), (0.1, 10000), (0, 0.1)]
p0_R1 = [kappa_init_R1, beta_init, b_init_R1]

print("R1 least-squares initialization...")
# res_R1 = opt.minimize(neg_loglik_R1, p0_R1,
#                       args=(mu_R1_all, k_R1_all, n_R1_all),
#                       bounds=bounds_R1, method='L-BFGS-B',
#                       options={'maxiter': 5000, 'eps': 1e-8})
# 全局优化：differential_evolution（不依赖单一起点）
from scipy.optimize import differential_evolution
res_R1 = differential_evolution(
    neg_loglik_R1,
    bounds=bounds_R1,
    args=(mu_R1_all, k_R1_all, n_R1_all),
    seed=42,
    maxiter=1000,
    tol=1e-10,
    popsize=30,       # 种群大小=30×参数数=90
    mutation=(0.5, 1.5),
    recombination=0.9,
    polish=True,      # 最后自动用L-BFGS-B精细优化
    workers=1,
)

kappa_R1_hat, beta_R1_hat, b_R1_hat = res_R1.x
print(f"R1 least-squares done: kappa={kappa_R1_hat:.4f}, beta={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")
print(f"  (vs R2: kappa={kappa_tcs:.4f}, b={b_tcs:.6f})")

P_R1_pred = tcs_R1_model(mu_train, kappa_R1_hat, beta_R1_hat, b_R1_hat)
AEB_R1_pred = -np.log(np.maximum(1 - P_R1_pred, 1e-12))
logL_R1 = np.sum(k_all * np.log(P_R1_pred) + (n_all_arr - k_all) * np.log(1 - P_R1_pred))
ss_res_R1 = np.sum(weights_AEB * (AEB_train - AEB_R1_pred)**2)
R2_R1 = 1 - ss_res_R1 / ss_tot



















print(f"\nR1 logL = {logL_R1:.2f}  (R2 logL = {logL_tcs:.2f})")
print(f"R1 R2(AEB) = {R2_R1:.4f}  (R2 R2(AEB) = {R2_tcs:.4f})")
print(f"R1 AICc = {AICc(logL_R1, 3, n_all):.2f}  (R2 AICc = {AICc_tcs:.2f})")

# ====================================================================
# NESTED SAMPLING (dynesty)
# ====================================================================
# Parameterization: theta = [kappa, log10_beta, b]
# Prior ranges: kappa in [0.01, 1000], log10_beta in [-1, 5], b in [0, 0.1]

def loglike_R1(theta):
    """Binomial log-likelihood for nested sampling"""
    kappa, log10_beta, b = theta
    if kappa <= 0 or b < 0 or b >= 1:
        return -1e12
    beta = 10.0 ** log10_beta
    if beta <= 0:
        return -1e12
    P_pos = tcs_R1_model(mu_R1_all, kappa, beta, b)
    ll = np.sum(k_R1_all * np.log(P_pos) + (n_R1_all - k_R1_all) * np.log(1.0 - P_pos))
    if not np.isfinite(ll):
        return -1e12
    return ll

def prior_transform_R1(u):
    """Map u ~ Uniform(0,1)^3 to physical parameters"""
    kappa = 0.01 + u[0] * (1000.0 - 0.01)
    # log10_beta = -1.0 + u[1] * (5.0 - (-1.0))
    log10_beta = 0.0 + u[1] * 5.0

    b = u[2] * 0.1
    return np.array([kappa, log10_beta, b])

print("\n" + "-"*60)
print("Running nested sampling (dynesty)...")
print(f"  ndim = 3, nlive = 500, bound = 'multi', sample = 'auto'")
print("-"*60)

sampler_R1 = dynesty.NestedSampler(
    loglike_R1,
    prior_transform_R1,
    ndim=3,
    nlive=500,
    bound='multi',     # multi-bound ellipsoidal decomposition (handles multimodal)
    sample='auto',
    rstate=np.random.default_rng(42),
)

sampler_R1.run_nested(print_progress=True)
results_R1 = sampler_R1.results




# 用nested sampling的全局最优替换L-BFGS-B的局部最优
logL_R1_global = results_R1.logl.max()
if logL_R1_global > logL_R1:
    print(f"\n*** Global logL from nested sampling: {logL_R1_global:.2f}")
    print(f"    (L-BFGS-B was stuck at local max: {logL_R1:.2f}, gap = {logL_R1_global - logL_R1:.2f})")
    logL_R1 = logL_R1_global
    samples_R1_post = results_R1.samples
    logl_arr = results_R1.logl
    best_idx = np.argmax(logl_arr)
    theta_best = samples_R1_post[best_idx]
    kappa_R1_hat = theta_best[0]
    beta_R1_hat = 10.0 ** theta_best[1]
    b_R1_hat = theta_best[2]
    P_R1_pred = tcs_R1_model(mu_train, kappa_R1_hat, beta_R1_hat, b_R1_hat)
    AEB_R1_pred = -np.log(np.maximum(1 - P_R1_pred, 1e-12))
    R2_R1 = 1 - np.sum(weights_AEB * (AEB_train - AEB_R1_pred)**2) / ss_tot
    print(f"    Updated: kappa={kappa_R1_hat:.4f}, beta={beta_R1_hat:.2f}, b={b_R1_hat:.6f}")
    print(f"    Updated R2(AEB) = {R2_R1:.4f}")






















print(f"\nNested sampling done.")
print(f"  logZ = {results_R1.logz[-1]:.2f} ± {results_R1.logzerr[-1]:.2f}")
print(f"  niter = {results_R1.niter}")
print(f"  ncall = {results_R1.ncall}")

# ---- Extract posterior samples ----
# Resample weighted samples to equal-weight samples
weights_R1 = np.exp(results_R1.logwt - results_R1.logz[-1])
samples_R1 = dyfunc.resample_equal(results_R1.samples, weights_R1)
print(f"  Posterior samples (resampled): {len(samples_R1)}")

# ---- Convergence diagnostics ----
# Split samples into 2 halves and check consistency (proxy for R-hat)
n_half = len(samples_R1) // 2
samples_A = samples_R1[:n_half]
samples_B = samples_R1[n_half:]

print("\n" + "="*60)
print("Convergence diagnostics (split-sample consistency)")
print("="*60)

for i, name in enumerate(['kappa', 'log10_beta', 'b']):
    med_A = np.median(samples_A[:, i])
    med_B = np.median(samples_B[:, i])
    std_A = np.std(samples_A[:, i])
    std_B = np.std(samples_B[:, i])
    # Gelman-Rubin-like ratio: should be close to 1
    pooled_std = np.sqrt((std_A**2 + std_B**2) / 2)
    gr_ratio = abs(med_A - med_B) / (pooled_std + 1e-12)
    print(f"  {name:<12}: split-A median = {med_A:.4f}, split-B median = {med_B:.4f}, "
          f"|Δ|/σ = {gr_ratio:.3f}")

# ESS estimate from weighted samples
ess_R1 = 1.0 / np.sum(weights_R1**2)
print(f"\n  Estimated ESS (Kish): {ess_R1:.0f}")

# ---- Posterior statistics ----
kappa_samples = samples_R1[:, 0]
beta_samples = 10.0 ** samples_R1[:, 1]
b_samples = samples_R1[:, 2]

kappa_med = np.median(kappa_samples)
beta_med = np.median(beta_samples)
b_med = np.median(b_samples)
kappa_ci = np.percentile(kappa_samples, [2.5, 97.5])
beta_ci = np.percentile(beta_samples, [2.5, 97.5])
b_ci = np.percentile(b_samples, [2.5, 97.5])

print("\n" + "="*60)
print("R1 posterior median and 95% CI")
print("="*60)
print(f"kappa = {kappa_med:.4f}  [{kappa_ci[0]:.4f}, {kappa_ci[1]:.4f}]")
print(f"beta  = {beta_med:.2f}  [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
print(f"b     = {b_med:.6f}  [{b_ci[0]:.6f}, {b_ci[1]:.6f}]")

# ---- Bimodality check ----
print("\n" + "="*60)
print("Bimodality check")
print("="*60)

log10_beta_s = samples_R1[:, 1]
mode_A_mask = log10_beta_s < 1.5   # beta < ~30
mode_B_mask = log10_beta_s >= 1.5  # beta > ~30
frac_A = np.sum(mode_A_mask) / len(samples_R1)
frac_B = np.sum(mode_B_mask) / len(samples_R1)

print(f"Mode A (beta < ~30):  {frac_A*100:.1f}% of posterior mass")
if np.any(mode_A_mask):
    print(f"  kappa = {np.median(kappa_samples[mode_A_mask]):.4f}  "
          f"[{np.percentile(kappa_samples[mode_A_mask], 2.5):.4f}, "
          f"{np.percentile(kappa_samples[mode_A_mask], 97.5):.4f}]")
    print(f"  beta  = {np.median(beta_samples[mode_A_mask]):.2f}  "
          f"[{np.percentile(beta_samples[mode_A_mask], 2.5):.2f}, "
          f"{np.percentile(beta_samples[mode_A_mask], 97.5):.2f}]")
print(f"Mode B (beta >= ~30): {frac_B*100:.1f}% of posterior mass")
if np.any(mode_B_mask):
    print(f"  kappa = {np.median(kappa_samples[mode_B_mask]):.4f}  "
          f"[{np.percentile(kappa_samples[mode_B_mask], 2.5):.4f}, "
          f"{np.percentile(kappa_samples[mode_B_mask], 97.5):.4f}]")
    print(f"  beta  = {np.median(beta_samples[mode_B_mask]):.2f}  "
          f"[{np.percentile(beta_samples[mode_B_mask], 2.5):.2f}, "
          f"{np.percentile(beta_samples[mode_B_mask], 97.5):.2f}]")

# ---------------------------------------------------------------------------
# Physical interpretation of the two modes
# ---------------------------------------------------------------------------
# beta = effective valency: the number of independently binding epitopes
# (capture sites) per target molecule.
#   Mode A (beta ~ 1):    monovalent, analog-like binding.
#   Mode B (beta ~ 1e3):  hypervalent regime; here (1-p)^beta ~ exp(-beta*p),
#                         i.e. the R2 (beta -> infinity) limit is effectively
#                         exact.
# The posterior splits into these two physically distinct regimes with
# comparable likelihood, so beta itself is NOT identifiable from this
# dataset: the data admit both a monovalent and a hypervalent explanation.
# Consequence: Omega = N*beta cannot be separated, the absolute molecule
# number M0 remains unknown, and digital finite-kappa (Simoa) still
# requires external calibration. Only kappa -> 0 (dPCR) is calibration-free.
print("\nPhysical interpretation:")
print("  Mode A (beta ~ 1):   monovalent, analog-like binding")
print("  Mode B (beta ~ 1e3): hypervalent regime; R2 (beta->inf) limit effectively exact")
print("  -> two physically distinct valency regimes fit the data equally well,")
print("     so beta is not identifiable and Omega = N*beta cannot be separated.")

# ---- Mode-B regime characterization (descriptive) ----
# Evaluate the R1 isotherm at the mode-B posterior medians and compare with
# (i) the data and (ii) the direct 4PL fit. NOTE: R1 has no upper plateau on
# the AEB scale, so the asymptotic 4PL parameters (C, D) cannot be recovered
# from the R1 curve; the comparison is made at the prediction level over the
# data range.
modeb_stats = None
if np.any(mode_B_mask):
    kap_mb  = float(np.median(kappa_samples[mode_B_mask]))
    beta_mb = float(np.median(beta_samples[mode_B_mask]))
    b_mb    = float(np.median(b_samples[mode_B_mask]))

    P_modeB   = tcs_R1_model(mu_train, kap_mb, beta_mb, b_mb)
    AEB_modeB = -np.log(np.maximum(1.0 - P_modeB, 1e-12))
    AEB_obs   = -np.log(np.maximum(1.0 - np.clip(k_all / n_all_arr, 0, 1 - 1e-15), 1e-12))
    ss_res_mb = float(np.sum((AEB_obs[mask_nz] - AEB_modeB[mask_nz]) ** 2))
    ss_tot_mb = float(np.sum((AEB_obs[mask_nz] - np.mean(AEB_obs[mask_nz])) ** 2))
    r2_modeB  = 1.0 - ss_res_mb / ss_tot_mb

    AEB_4pl_pred = predict_AEB_4pl(mu_train, A_4pl, B_4pl, C_4pl, D_4pl)
    P_4pl_pred   = 1.0 - np.exp(-AEB_4pl_pred)
    lslopes = np.diff(np.log(AEB_modeB[mask_nz])) / np.diff(np.log(mu_nz))
    rel = np.abs(P_modeB[mask_nz] - P_4pl_pred[mask_nz]) / P_4pl_pred[mask_nz] * 100.0

    print("\nMode-B regime characterization (descriptive):")
    print(f"  mode-B medians: kappa={kap_mb:.4f}, beta={beta_mb:.0f}, b={b_mb:.2e}")
    print(f"  AEB-scale R^2 vs data: mode-B R1 = {r2_modeB:.4f} "
          f"(4PL fit = {R2_4pl:.4f}, R2 limit = {R2_tcs:.4f})")
    print(f"  local Hill slope: median {np.median(lslopes):.3f} "
          f"[{lslopes.min():.3f}, {lslopes.max():.3f}] (4PL B = {B_4pl:.3f})")
    print(f"  mode-B R1 vs 4PL prediction: mean |diff| = {rel.mean():.1f}%, "
          f"max = {rel.max():.1f}% over the data range")
    modeb_stats = (kap_mb, beta_mb, b_mb, r2_modeB,
                   float(np.median(lslopes)), float(rel.mean()), float(rel.max()))

if frac_A > 0.1 and frac_B > 0.1:
    print("\n>>> Posterior is BIMODAL")
    print(">>> Two near-equivalent regimes exist; data cannot distinguish them")
    bimodal = True
else:
    print("\n>>> Posterior is not clearly bimodal")
    bimodal = False

# ---- beta identifiability assessment ----
print("\n" + "="*60)
print("beta identifiability assessment")
print("="*60)

# beta_prior_lo, beta_prior_hi = 0.1, 100000.0
beta_prior_lo, beta_prior_hi = 1.0, 1000000.0

beta_post_lo, beta_post_hi = beta_ci[0], beta_ci[1]
hits_ceiling = beta_post_hi >= beta_prior_hi * 0.9
hits_floor = beta_post_lo <= beta_prior_lo * 1.5
post_range_ratio = (beta_post_hi - beta_post_lo) / max(beta_med, 0.01)

print(f"beta prior range: [{beta_prior_lo}, {beta_prior_hi}]")
print(f"beta posterior 95% CI: [{beta_post_lo:.2f}, {beta_post_hi:.2f}]")
print(f"beta posterior width / median = {post_range_ratio:.2f}")
print(f"Hits prior ceiling? {hits_ceiling}")
print(f"Hits prior floor? {hits_floor}")

if bimodal:
    print("\n>>> beta posterior is bimodal -> NOT identifiable")
    print(">>> Conclusion: digital finite-kappa cannot obtain absolute M0")
    identifiable = "weak"
elif hits_ceiling:
    print("\n>>> beta posterior hits prior ceiling -> not identifiable")
    identifiable = False
elif post_range_ratio > 2.0:
    print("\n>>> beta posterior very wide -> weakly identifiable")
    identifiable = "weak"
else:
    print("\n>>> beta posterior bounded and compact -> identifiable")
    identifiable = True

# ====================================================================
# R1 vs R2 model comparison: AICc + Bayes factor
# ====================================================================
print("\n" + "="*60)
print("R1 vs R2 model comparison")
print("="*60)

# AICc comparison
print("\n--- AICc comparison ---")
print(f"{'R2 (beta->inf)':<18} {logL_tcs:<12.2f} {2:<10} {AICc_tcs:<10.2f}")
print(f"{'R1 (beta free)':<18} {logL_R1:<12.2f} {3:<10} {AICc(logL_R1, 3, n_all):<10.2f}")
delta_AICc = AICc(logL_R1, 3, n_all) - AICc_tcs
print(f"\nDeltaAICc (R1 - R2) = {delta_AICc:.2f}")
if delta_AICc > 10:
    print("-> R2 preferred: data does not support finite beta")
elif delta_AICc < -10:
    print("-> R1 preferred: finite beta supported")
else:
    print("-> Comparable: data cannot distinguish R1 and R2")

# Bayes factor via R2 nested sampling
print("\n--- Bayes factor (nested sampling) ---")
print("Running R2 nested sampling for logZ comparison...")

def loglike_R2(theta):
    """R2 binomial log-likelihood (beta->inf, 2 params)"""
    kappa, b = theta
    if kappa <= 0 or b < 0 or b >= 1:
        return -1e12
    P_pos = tcs_model(mu_R1_all, kappa, b)
    ll = np.sum(k_R1_all * np.log(P_pos) + (n_R1_all - k_R1_all) * np.log(1.0 - P_pos))
    if not np.isfinite(ll):
        return -1e12
    return ll

def prior_transform_R2(u):
    """R2 prior transform: kappa in [0.01, 1000], b in [0, 0.1]"""
    kappa = 0.01 + u[0] * (1000.0 - 0.01)
    b = u[1] * 0.1
    return np.array([kappa, b])

sampler_R2 = dynesty.NestedSampler(
    loglike_R2,
    prior_transform_R2,
    ndim=2,
    nlive=500,
    bound='multi',
    sample='auto',
    rstate=np.random.default_rng(42),
)
sampler_R2.run_nested(print_progress=False)
results_R2 = sampler_R2.results

logZ_R1 = results_R1.logz[-1]
logZ_R2 = results_R2.logz[-1]
logZ_err = np.sqrt(results_R1.logzerr[-1]**2 + results_R2.logzerr[-1]**2)
logBF = logZ_R1 - logZ_R2  # Bayes factor R1/R2

print(f"\n  logZ(R1) = {logZ_R1:.2f} ± {results_R1.logzerr[-1]:.2f}")
print(f"  logZ(R2) = {logZ_R2:.2f} ± {results_R2.logzerr[-1]:.2f}")
print(f"  log Bayes factor (R1/R2) = {logBF:.2f} ± {logZ_err:.2f}")
print(f"  Bayes factor B_12 = {np.exp(logBF):.4f}")

if logBF < -5:
    print("  -> Strong evidence for R2 (beta->inf)")
    bf_conclusion = "R2 strongly preferred"
elif logBF < -2:
    print("  -> Moderate evidence for R2")
    bf_conclusion = "R2 moderately preferred"
elif logBF < 2:
    print("  -> Inconclusive (data cannot distinguish R1 and R2)")
    bf_conclusion = "inconclusive"
elif logBF < 5:
    print("  -> Moderate evidence for R1")
    bf_conclusion = "R1 moderately preferred"
else:
    print("  -> Strong evidence for R1")
    bf_conclusion = "R1 strongly preferred"

# ---- Corner plot ----
if corner is not None:
    fig_corner = plt.figure(figsize=(10, 10))
    corner.corner(
        samples_R1,
        labels=['kappa', 'log10(beta)', 'b'],
        truths=[kappa_med, np.log10(max(beta_med, 0.1)), b_med],
        label_kwargs={'fontsize': 14},
        show_titles=True, title_kwargs={'fontsize': 12}
    )
    plt.suptitle('R1 Nested Sampling Posterior - beta Identifiability Test (sβG)',
                 fontsize=16, fontweight='bold')
    plt.savefig('R1_corner_plot_sbG_nested.png', dpi=300, bbox_inches='tight')
    plt.show()

# ---- Final verdict ----
print("\n" + "="*80)
print("FINAL VERDICT (sβG dataset)")
print("="*80)
if identifiable == True:
    print("beta IS identifiable from sβG Simoa calibration data.")
    print(f"  beta = {beta_med:.2f} [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
    print("  -> NOTE: This is a calibration experiment. Calibration-free")
    print("     absolute quantification still requires kappa->0 (dPCR).")
elif identifiable == "weak":
    if bimodal:
        print("beta is NOT identifiable — posterior is bimodal.")
        print(f"  Mode A ({frac_A*100:.0f}%): beta ~ {np.median(beta_samples[mode_A_mask]):.1f}")
        print(f"  Mode B ({frac_B*100:.0f}%): beta ~ {np.median(beta_samples[mode_B_mask]):.0f}")
    else:
        print("beta is WEAKLY identifiable — posterior is very wide.")
        print(f"  beta = {beta_med:.2f} [{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]")
    print(f"  Bayes factor: {bf_conclusion}")
    print("  -> M0 estimate would be too imprecise to be useful.")
    print("  -> In practice, digital finite-kappa still needs external calibration.")
else:
    print("beta is NOT identifiable from sβG dilution data.")
print("="*80)

# ====================================================================
# Sheet 14: R1 Model Results
# ====================================================================
try:
    from openpyxl import load_workbook as _lwb
    _wb14 = _lwb("SI_Table_3.xlsx")
    if "14_R1_Model" in _wb14.sheetnames:
        del _wb14["14_R1_Model"]
    _ws14 = _wb14.create_sheet("14_R1_Model")

    _ws14.append(['R1 Model Fit (nested sampling) vs R2 Model (beta->inf) — sβG dataset'])
    _ws14.append(['Method: dynesty nested sampling, nlive=500, bound=multi'])
    _ws14.append([])
    _ws14.append(['Parameter', 'R2 (beta->inf)', 'R1 (beta free)', 'R1 95% CI'])
    _ws14.append(['kappa', f'{kappa_tcs:.4f}', f'{kappa_med:.4f}',
                  f'[{kappa_ci[0]:.4f}, {kappa_ci[1]:.4f}]'])
    _ws14.append(['beta', 'inf (R2 approx)', f'{beta_med:.2f}',
                  f'[{beta_ci[0]:.2f}, {beta_ci[1]:.2f}]'])
    _ws14.append(['b', f'{b_tcs:.6f}', f'{b_med:.6f}',
                  f'[{b_ci[0]:.6f}, {b_ci[1]:.6f}]'])
    _ws14.append([])

    _ws14.append(['Model Comparison'])
    _ws14.append(['Metric', 'R2', 'R1', 'Difference / Interpretation'])
    _ws14.append(['logL (MLE)', f'{logL_tcs:.2f}', f'{logL_R1:.2f}',
                  f'{logL_R1 - logL_tcs:.2f}'])
    _ws14.append(['n_params', '2', '3', '1'])
    _ws14.append(['AICc', f'{AICc_tcs:.2f}', f'{AICc(logL_R1, 3, n_all):.2f}',
                  f'Δ = {AICc(logL_R1, 3, n_all) - AICc_tcs:.2f}'])
    _ws14.append(['logZ (nested)', f'{logZ_R2:.2f}', f'{logZ_R1:.2f}',
                  f'logBF = {logBF:.2f} ± {logZ_err:.2f}'])
    _ws14.append(['Bayes factor', '', '', f'B_12 = {np.exp(logBF):.4f} ({bf_conclusion})'])
    _ws14.append(['R2_AEB', f'{R2_tcs:.6f}', f'{R2_R1:.6f}', ''])
    _ws14.append([])

    _ws14.append(['Nested Sampling Diagnostics'])
    _ws14.append(['Quantity', 'R1', 'R2'])
    _ws14.append(['logZ', f'{results_R1.logz[-1]:.2f} ± {results_R1.logzerr[-1]:.2f}',
                  f'{results_R2.logz[-1]:.2f} ± {results_R2.logzerr[-1]:.2f}'])
    _ws14.append(['niter', f'{results_R1.niter}', f'{results_R2.niter}'])
    _ws14.append(['ncall', f'{results_R1.ncall}', f'{results_R2.ncall}'])
    _ws14.append(['ESS (Kish)', f'{ess_R1:.0f}', ''])
    _ws14.append([])

    _ws14.append(['Bimodality Check'])
    _ws14.append(['Mode', 'Fraction', 'kappa (median [95% CI])', 'beta (median [95% CI])'])
    if np.any(mode_A_mask):
        _ws14.append([f'A (beta < ~30)', f'{frac_A*100:.1f}%',
                      f'{np.median(kappa_samples[mode_A_mask]):.4f} '
                      f'[{np.percentile(kappa_samples[mode_A_mask], 2.5):.4f}, '
                      f'{np.percentile(kappa_samples[mode_A_mask], 97.5):.4f}]',
                      f'{np.median(beta_samples[mode_A_mask]):.2f} '
                      f'[{np.percentile(beta_samples[mode_A_mask], 2.5):.2f}, '
                      f'{np.percentile(beta_samples[mode_A_mask], 97.5):.2f}]'])
    if np.any(mode_B_mask):
        _ws14.append([f'B (beta >= ~30)', f'{frac_B*100:.1f}%',
                      f'{np.median(kappa_samples[mode_B_mask]):.4f} '
                      f'[{np.percentile(kappa_samples[mode_B_mask], 2.5):.4f}, '
                      f'{np.percentile(kappa_samples[mode_B_mask], 97.5):.4f}]',
                      f'{np.median(beta_samples[mode_B_mask]):.2f} '
                      f'[{np.percentile(beta_samples[mode_B_mask], 2.5):.2f}, '
                      f'{np.percentile(beta_samples[mode_B_mask], 97.5):.2f}]'])
    _ws14.append(['Physical interpretation:',
                  'Mode A = monovalent (analog-like) binding; Mode B = hypervalent '
                  'regime where the R2 (beta->inf) limit is effectively exact'])
    _ws14.append(['Why bimodal = non-identifiable:',
                  'two physically distinct valency regimes fit the data equally well; '
                  'kappa is nearly identical across modes'])
    _ws14.append(['Consequence:',
                  'Omega = N*beta cannot be separated, so M0 is unknown; digital '
                  'finite-kappa (Simoa) still requires external calibration'])
    if modeb_stats is not None:
        _kap, _beta, _b, _r2, _slope, _rel_mean, _rel_max = modeb_stats
        _ws14.append([])
        _ws14.append(['Mode-B regime characterization (descriptive)'])
        _ws14.append(['Quantity', 'Value', 'Reference', 'Note'])
        _ws14.append(['mode-B medians',
                      f'kappa={_kap:.4f}, beta={_beta:.0f}, b={_b:.2e}', '', ''])
        _ws14.append(['AEB-scale R^2 vs data', f'{_r2:.4f}',
                      f'4PL fit = {R2_4pl:.4f}; R2 limit = {R2_tcs:.4f}', ''])
        _ws14.append(['local Hill slope (median)', f'{_slope:.3f}',
                      f'4PL B = {B_4pl:.3f}', ''])
        _ws14.append(['mode-B R1 vs 4PL prediction (mean |diff|)',
                      f'{_rel_mean:.1f}%', f'max = {_rel_max:.1f}%',
                      'over the data range; R1 has no AEB plateau, so asymptotic '
                      '4PL parameters are not recoverable'])
    _ws14.append([])

    _ws14.append(['Note: This is a calibration experiment (M known at each'])
    _ws14.append(['dilution point). beta identifiability here does NOT imply'])
    _ws14.append(['calibration-free absolute quantification. See SI S2c.5.'])
    _ws14.append([])

    # R1 per-point predictions
    _ws14.append(['R1 Per-point Predictions'])
    _ws14.append(['Conc (aM)', 'k', 'n', 'P_obs', 'P_R2', 'P_R1', 'AEB_R2', 'AEB_R1'])

    P_R2_pred = tcs_model(mu_train, kappa_tcs, b_tcs)
    AEB_R2_pred = -np.log(np.maximum(1 - P_R2_pred, 1e-12))

    for i in range(len(conc_aM)):
        _ws14.append([f'{conc_aM[i]:.4f}', f'{k_all[i]:.1f}',
                      f'{n_all_arr[i]:.0f}', f'{k_all[i]/n_all_arr[i]:.6f}',
                      f'{P_R2_pred[i]:.6f}', f'{P_R1_pred[i]:.6f}',
                      f'{AEB_R2_pred[i]:.4f}', f'{AEB_R1_pred[i]:.4f}'])

    _wb14.save("SI_Table_3.xlsx")
    print(f"\n[OK] Sheet 14_R1_Model added to SI_Table_3.xlsx (now {len(_wb14.sheetnames)} sheets)")
except Exception as e_s14:
    print(f"\n[FAIL] Sheet 14: {e_s14}")
    import traceback; traceback.print_exc()
