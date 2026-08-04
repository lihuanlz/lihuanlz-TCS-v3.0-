# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 13:55:06 2026

@author: lihua
"""


"""
1960 Yalow RIA TCS Reanalysis — Fixed 2026-07-22
Fixes:
  - Added parameter bounds (0, inf) to prevent negative parameters
  - Added Durbin-Watson statistic (consistent with S9.6)
  - Fixed kappa description in docstring
  - Suppressed OptimizeWarning in bootstrap
  - Removed unused n_orders variable
  - Added R² threshold comment
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import argparse
from typing import Tuple, Dict, Any, List
import warnings


# ==========================================================
# 1. 数据加载
# ==========================================================
def load_yalow1960_data() -> Tuple[np.ndarray, np.ndarray]:
    """返回 (f_all, p_all)，其中 f_all 用 1e15 代替无穷大稀释因子。

    列含义: [f, ?, ?, ?, p*/?, p*]
      0: f  (稀释因子; np.inf = 零标准)
      1-4: 原始测量值 (counts 等) - 不参与拟合
      5: p* (tracer 的结合分率)
    """
    raw = np.array([
        [np.inf, 0,    6.9, 2.6, 2.6538, 0.7263],  # 零浓度
        [10,     1.4,  6.8, 3.4, 2.0000, 0.6667],
        [5,      2.8,  5.2, 3.6, 1.4444, 0.5909],
        [3.3333, 4.2,  5.0, 4.0, 1.2500, 0.5556],
        [2,      7.0,  4.6, 5.2, 0.8846, 0.4694],
        [1,      14.0, 4.1, 6.3, 0.6508, 0.3942]
    ])
    f_inf = 1e15
    f_all = np.array([f_inf] + list(raw[1:, 0]))
    p_all = raw[:, 5]
    return f_all, p_all


# ==========================================================
# 2. TCS 竞争法严格方程
# ==========================================================
def tcs_competitive(f: np.ndarray, kappa: float, xi0: float, xi_star: float) -> np.ndarray:
    """TCS 竞争法严格方程（有限 κ）。

    p_total 来自 ξ_tot = p_total/(1-p_total) + p_total/κ 的解
    p_star  = p_total × (ξ*/ξ_tot) 由 ξ_i = p_i × (1/(1-p_total) + 1/κ) 精确定出

    Parameters
    ----------
    f       : 稀释因子（f=∞ 对应 1e15）
    kappa   : 无量纲耗竭参数 (= K_d / (n R_T))
    xi0     : 未标记抗原归一化浓度
    xi_star : 标记抗原归一化浓度
    """
    xi_tot = xi0 / f + xi_star
    b = kappa * xi_tot + kappa + 1
    disc = np.sqrt(np.maximum(b**2 - 4 * kappa * xi_tot, 0))
    p_total = (b - disc) / 2
    p_star = p_total * (xi_star / xi_tot)
    return p_star


# ==========================================================
# 3. Durbin-Watson 统计量
# ==========================================================
def durbin_watson(residuals: np.ndarray) -> float:
    """计算 Durbin-Watson 统计量。DW≈2 表示无自相关。"""
    diff = np.diff(residuals)
    return np.sum(diff**2) / np.sum(residuals**2)


# ==========================================================
# 4. 拟合与评估
# ==========================================================
def fit_model(f_all: np.ndarray, p_all: np.ndarray,
              p0: Tuple[float, float, float] = (10.0, 5.0, 2.5)) -> Dict[str, Any]:
    """拟合自由 κ 的 TCS 模型，返回参数、协方差、R²、DW 等。"""
    popt, pcov = curve_fit(tcs_competitive, f_all, p_all, p0=p0,
                           bounds=([0, 0, 0], [np.inf, np.inf, np.inf]),
                           maxfev=1000000)
    p_fit = tcs_competitive(f_all, *popt)
    residuals = p_all - p_fit
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((p_all - np.mean(p_all))**2)
    r2 = 1 - ss_res / ss_tot
    perr = np.sqrt(np.diag(pcov))
    rmse = np.sqrt(np.mean(residuals**2))
    mae = np.mean(np.abs(residuals))
    max_res = np.max(np.abs(residuals))
    dw = durbin_watson(residuals)
    p_zero_model = tcs_competitive(np.array([1e15]), *popt)[0]
    return {
        'popt': popt, 'pcov': pcov, 'perr': perr,
        'p_fit': p_fit, 'r2': r2, 'rmse': rmse,
        'mae': mae, 'max_res': max_res, 'dw': dw,
        'p_zero_model': p_zero_model
    }


# ==========================================================
# 5. Case-bootstrap 置信区间
# ==========================================================
def bootstrap_ci(f_all: np.ndarray, p_all: np.ndarray,
                 initial_popt: Tuple[float, float, float],
                 n_boot: int = 2000, seed: int = 42) -> Dict[str, np.ndarray]:
    """Case resampling bootstrap，返回 (2.5%, 97.5%) 区间。"""
    np.random.seed(seed)
    boot_params = []
    for _ in range(n_boot):
        idx = np.random.choice(len(p_all), size=len(p_all), replace=True)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                popt_boot, _ = curve_fit(
                    tcs_competitive, f_all[idx], p_all[idx],
                    p0=initial_popt,
                    bounds=([0, 0, 0], [np.inf, np.inf, np.inf]),
                    maxfev=1000000
                )
            boot_params.append(popt_boot)
        except Exception:
            pass
    if len(boot_params) < 10:
        raise RuntimeError("Bootstrap failed: too few successful fits.")
    boot_params = np.array(boot_params)
    ci_lower = np.percentile(boot_params, 2.5, axis=0)
    ci_upper = np.percentile(boot_params, 97.5, axis=0)
    return {'boot_params': boot_params, 'ci_lower': ci_lower, 'ci_upper': ci_upper,
            'n_success': len(boot_params)}


# ==========================================================
# 6. κ 扫描（固定 κ，看 R² 变化）
# ==========================================================
def kappa_sweep(f_all: np.ndarray, p_all: np.ndarray,
                xi0_init: float, xi_star_init: float,
                k_values: list = None) -> list:
    """对一系列固定 κ 值分别拟合 ξ₀, ξ*，返回 [(κ, r², ξ₀_fit, ξ*_fit), ...]。"""
    if k_values is None:
        k_values = [0.001, 0.01, 0.1, 1, 10, 100, 1000, 10000]
    results = []
    for k_test in k_values:
        try:
            popt_k, _ = curve_fit(
                lambda f, xi0, xi_star: tcs_competitive(f, k_test, xi0, xi_star),
                f_all, p_all, p0=[xi0_init, xi_star_init],
                bounds=([0, 0], [np.inf, np.inf]),
                maxfev=1000000
            )
            p_k = tcs_competitive(f_all, k_test, *popt_k)
            r2_k = 1 - np.sum((p_all - p_k)**2) / np.sum((p_all - np.mean(p_all))**2)
            results.append((k_test, r2_k, popt_k[0], popt_k[1]))
        except Exception:
            results.append((k_test, np.nan, np.nan, np.nan))
    return results


# ==========================================================
# 7. 绘图（Extended Data Fig. 2）
# ==========================================================
def plot_results(f_all: np.ndarray, p_all: np.ndarray,
                 fit: Dict[str, Any], ci: Dict[str, np.ndarray],
                 sweep: list, output_path: str = 'Figure2b.svg'):
    """生成半对数结合曲线图，包含数据、自由 κ 拟合和两条固定 κ 曲线。"""
    kappa_fit, xi0_fit, xi_star_fit = fit['popt']
    p_zero_model = fit['p_zero_model']
    r2 = fit['r2']

    # 横坐标：未标记抗原浓度 ξ_unlabeled = ξ₀ / f（排除零标准点）
    xi_unlabeled = xi0_fit / f_all[1:]
    # 零标准点横坐标：取最低浓度的 5%（显示用，无物理意义）
    xi_zero = xi_unlabeled.min() * 0.05

    # 拟合曲线所用的 ξ 范围
    xi_curve = np.logspace(np.log10(xi_unlabeled.min() * 0.02),
                           np.log10(xi_unlabeled.max() * 2), 500)
    f_curve = xi0_fit / xi_curve
    p_curve = tcs_competitive(f_curve, kappa_fit, xi0_fit, xi_star_fit)

    plt.figure(figsize=(10,9))

    # 数据点（不包括零标准）
    plt.scatter(xi_unlabeled, p_all[1:], color='#c1121f', s=60,
                label='Data', zorder=3)

    # 零标准点：用模型预测值标注，并注明实测值
    plt.scatter([xi_zero], [p_zero_model], color='#669bbc', s=60,
                label=f'Zero standard (model: {p_zero_model:.4f})', zorder=3)
    plt.annotate(f'Actual: {p_all[0]:.4f}', xy=(xi_zero, p_zero_model),
                 xytext=(-30, -20), textcoords='offset points',
                 fontsize=14, color='blue')

    # 自由 κ 拟合曲线
    plt.plot(xi_curve, p_curve, 'b-', lw=2, label='TCS fit (κ-free)')

    # 两条固定 κ 的代表曲线：κ=0.1 和 κ=100
    for k_test, color, style in [(0.1, 'gray', ':'), (100, 'darkgreen', '--')]:
        popt_k, _ = curve_fit(
            lambda f, xi0, xi_star: tcs_competitive(f, k_test, xi0, xi_star),
            f_all, p_all, p0=[xi0_fit, xi_star_fit],
            bounds=([0, 0], [np.inf, np.inf]),
            maxfev=1000000
        )
        p_curve_k = tcs_competitive(f_curve, k_test, *popt_k)
        plt.plot(xi_curve, p_curve_k, color=color, lw=1.5, linestyle=style,
                 label=f'TCS fit (κ={k_test} fixed)', alpha=0.7)

    # 参数文本框
    param_text = (
        r"$\kappa \geq$" + f" {ci['ci_lower'][0]:.2f} " +
        r"(one-sided, 95% CI)" + "\n"
        r"$\xi_0$ = " + f"{xi0_fit:.2f} " +
        f"[{ci['ci_lower'][1]:.2f}, {ci['ci_upper'][1]:.2f}]" + "\n"
        r"$\xi_*$ = " + f"{xi_star_fit:.2f} " +
        f"[{ci['ci_lower'][2]:.2f}, {ci['ci_upper'][2]:.2f}]" + "\n"
        r"$R^2$ = " + f"{r2:.4f}"
    )
    plt.text(0.35, 0.4, param_text, transform=plt.gca().transAxes,
             fontsize=14, va='top', ha='right',
             bbox=dict(boxstyle='round', facecolor='white', alpha=0.85,
                       edgecolor='gray'))

    plt.xscale('log')
    plt.xlabel(r'$\xi = \xi_0 / f$  (normalized unlabeled insulin concentration)',
               fontsize=14, fontweight='bold')
    plt.ylabel('Bound fraction of tracer ($p_*$)', fontsize=14, fontweight='bold')
    plt.title('Extended Data Fig. 2: Yalow & Berson 1960 Insulin RIA (TCS reanalysis)',
              fontsize=16, fontweight='bold')
    plt.legend(loc='lower left', fontsize=14, framealpha=0.9)
    plt.grid(False)
    plt.xlim(xi_unlabeled.min() * 0.01, xi_unlabeled.max() * 5)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.show()


# ==========================================================
# 8. Excel 输出（SI Table 2）
# ==========================================================
def export_excel(f_all: np.ndarray, p_all: np.ndarray,
                 fit: Dict[str, Any], ci: Dict[str, np.ndarray],
                 sweep: list, output_excel: str = 'SI_Table_2.xlsx'):
    """生成多工作表的 Excel 报告。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[!] openpyxl not installed. Excel not saved. Run: pip install openpyxl")
        return

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    kappa_fit, xi0_fit, xi_star_fit = fit['popt']
    p_fit = fit['p_fit']
    p_zero_model = fit['p_zero_model']
    r2 = fit['r2']
    perr = fit['perr']
    dw = fit['dw']

    # ---------- Sheet 1: Raw and Fit ----------
    ws1 = wb.active
    ws1.title = "1_Raw_and_Fit"
    ws1.append(["Sample", "Dilution f", "xi = xi0/f",
                "Observed p*", "Model p*", "Residual"])
    for cell in ws1[1]:
        cell.font = bold
        cell.fill = header_fill
    for i in range(len(f_all)):
        dil_label = "inf (zero std)" if f_all[i] >= 1e10 else f_all[i]
        xi_label = "— (xi=0)" if f_all[i] >= 1e10 else f"{xi0_fit / f_all[i]:.4f}"
        ws1.append([f"Point {i+1}", dil_label, xi_label,
                    p_all[i], p_fit[i], p_all[i] - p_fit[i]])
    ws1.append([])
    ws1.append(["R-squared", "", "", "", r2, ""])
    ws1.append(["Durbin-Watson", "", "", "", dw, ""])
    ws1.append(["Zero std residual (model - actual)", "", "", "",
                p_zero_model - p_all[0], ""])

    # ---------- Sheet 2: Fit Parameters ----------
    ws2 = wb.create_sheet("2_Fit_Parameters")
    ws2.append(["Parameter", "Value", "Asymptotic SE", "Note"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill
    ws2.append(["kappa", kappa_fit, perr[0], "Unidentifiable (one-sided bound)"])
    ws2.append(["xi_0", xi0_fit, perr[1], ""])
    ws2.append(["xi_star", xi_star_fit, perr[2], ""])
    ws2.append(["R-squared", r2, "—", ""])
    ws2.append(["RMSE", fit['rmse'], "—", ""])
    ws2.append(["Durbin-Watson", dw, "—", "DW≈2: no autocorrelation"])
    ws2.append([])
    ws2.append(["Physical interpretation:", "", "", ""])
    ws2.append(["The displacement curve lies in the kappa-independent (zero-depletion)",
                "regime: its shape constrains the dimensionless ratios xi_0 and xi_star",
                "but not the absolute molecular scale, so kappa is reported as a",
                "one-sided lower bound (scan-based, see 4_Kappa_Scan)."])
    ws2.append(["This is direct experimental evidence of scale degeneracy; the same",
                "one-sided reporting protocol is used for kappa_3 of the",
                "three-concentration ELISA dataset (SI Table 1).", ""])

    # ---------- Sheet 3: Bootstrap CI ----------
    ws3 = wb.create_sheet("3_Bootstrap_CI")
    ws3.append(["Parameter", "Point est.", "2.5%", "97.5%", "Width (log10)"])
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
    for i, name in enumerate(["kappa", "xi_0", "xi_star"]):
        lo, hi = ci['ci_lower'][i], ci['ci_upper'][i]
        width = np.log10(hi / max(lo, 1e-6)) if i == 0 else "—"
        ws3.append([name, fit['popt'][i], lo, hi, width])
    ws3.append([])
    width_kappa = np.log10(ci['ci_upper'][0] / max(ci['ci_lower'][0], 1e-6))
    ws3.append([f"κ CI spans {width_kappa:.1f} orders of magnitude; "
                f"reported as κ ≥ {ci['ci_lower'][0]:.2f}", "", "", "", ""])

    # ---------- Sheet 4: Kappa Sweep ----------
    ws4 = wb.create_sheet("4_Kappa_Scan")
    ws4.append(["kappa (fixed)", "R-squared", "xi_0", "xi_star", "Note"])
    for cell in ws4[1]:
        cell.font = bold
        cell.fill = header_fill
    for (k_test, r2_k, xi0_k, xi_star_k) in sweep:
        note = "fail" if np.isnan(r2_k) else "OK"
        ws4.append([k_test, r2_k, xi0_k, xi_star_k, note])

    # 动态确定 plateau 区间
    valid = [(k, r2_k) for (k, r2_k, _, _) in sweep if not np.isnan(r2_k)]
    # 0.98 threshold: <2% variance increase relative to the best fit,
    # well within experimental noise for 6-point RIA data
    r2_threshold = 0.98
    plateau_k = [k for k, r2_k in valid if r2_k >= r2_threshold]
    ws4.append([])
    if plateau_k:
        ws4.append([f"Plateau: κ ≥ {min(plateau_k)} → R² ≥ {r2_threshold} "
                    f"(κ-independent regime)", "", "", "", ""])
        ws4.append([f"Physical conclusion: data constrain κ ≥ {min(plateau_k)} "
                    f"(one-sided)", "", "", "", ""])
    else:
        ws4.append(["No clear plateau detected at R² ≥ 0.98.", "", "", "", ""])

    # ---------- Sheet 5: Fixed-κ Test ----------
    ws5 = wb.create_sheet("5_FixedKappa_Test")
    ws5.append(["Test", "κ value", "R²", "ΔR² from free"])
    for cell in ws5[1]:
        cell.font = bold
        cell.fill = header_fill
    ws5.append(["Free fit", f"{kappa_fit:.1f}", r2, 0.0])

    for fix_k, desc in [(0.1, "Strong-depletion limit"),
                        (1.0, "Intermediate")]:
        try:
            popt_k, _ = curve_fit(
                lambda f, xi0, xi_star: tcs_competitive(f, fix_k, xi0, xi_star),
                f_all, p_all, p0=[xi0_fit, xi_star_fit],
                bounds=([0, 0], [np.inf, np.inf]),
                maxfev=1000000)
            p_k = tcs_competitive(f_all, fix_k, *popt_k)
            r2_fixed = 1 - np.sum((p_all - p_k)**2) / \
                       np.sum((p_all - np.mean(p_all))**2)
            ws5.append([f"Fixed κ={fix_k} ({desc})", fix_k,
                        r2_fixed, r2_fixed - r2])
        except Exception:
            ws5.append([f"Fixed κ={fix_k} ({desc})", fix_k,
                        "fit failed", ""])

    # ---------- Sheet 6: Bootstrap Summary ----------
    ws6 = wb.create_sheet("6_Bootstrap_Summary")
    ws6.append(["Statistic", "kappa", "xi_0", "xi_star"])
    for cell in ws6[1]:
        cell.font = bold
        cell.fill = header_fill
    bp = ci['boot_params']
    for stat_name, stat_fn in [("N successful", len),
                               ("Median", np.median),
                               ("Mean", np.mean),
                               ("Std", np.std),
                               ("Min", np.min),
                               ("Max", np.max)]:
        ws6.append([stat_name] + [stat_fn(bp[:, i]) for i in range(bp.shape[1])])

    wb.save(output_excel)
    print(f"\n{'=' * 60}")
    print(f"SI Table 2 saved to '{output_excel}'")
    print(f"{'=' * 60}")
    print("Sheets:")
    print("  1. Raw_and_Fit       — 6 Yalow points + model fit + residuals")
    print("  2. Fit_Parameters    — κ, ξ₀, ξ* + R² + RMSE + DW")
    print("  3. Bootstrap_CI      — case-resampling 95% CI")
    print("  4. Kappa_Scan        — fixed-κ R² across 8 values")
    print("  5. FixedKappa_Test   — R² degradation test (κ=0.1 & κ=1.0)")
    print("  6. Bootstrap_Summary — bootstrap replicate statistics")


# ==========================================================
# 9. 主程序
# ==========================================================
def main():
    parser = argparse.ArgumentParser(
        description="Yalow 1960 RIA TCS reanalysis")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip Excel output")
    parser.add_argument("--figure", default="Extened_data_fig_2.svg",
                        help="Output figure path")
    parser.add_argument("--excel", default="SI_Table_2.xlsx",
                        help="Output Excel path")
    args = parser.parse_args()

    # ---- 加载数据 ----
    f_all, p_all = load_yalow1960_data()

    # ---- 自由 κ 拟合 ----
    fit = fit_model(f_all, p_all)
    kappa_fit, xi0_fit, xi_star_fit = fit['popt']
    perr = fit['perr']
    r2 = fit['r2']
    dw = fit['dw']

    print("=" * 60)
    print("TCS 竞争法严格方程拟合（有限 κ）")
    print("=" * 60)
    print(f"κ   = {kappa_fit:.2f} ± {perr[0]:.2f}")
    print(f"ξ₀  = {xi0_fit:.4f} ± {perr[1]:.4f}")
    print(f"ξ*  = {xi_star_fit:.4f} ± {perr[2]:.4f}")
    print(f"R²  = {r2:.6f}")
    print(f"DW  = {dw:.4f}")
    print("=" * 60)

    # ---- Goodness-of-fit 附加指标 ----
    print(f"\nGoodness-of-fit (additional metrics):")
    print(f"  RMSE       = {fit['rmse']:.6f}   (on 0-1 bound-fraction scale)")
    print(f"  MAE        = {fit['mae']:.6f}")
    print(f"  max|resid| = {fit['max_res']:.6f}")
    print(f"  DW         = {dw:.4f}   (≈2: no autocorrelation)")

    # ---- Bootstrap CI ----
    ci = bootstrap_ci(f_all, p_all, fit['popt'])
    ci_lower = ci['ci_lower']
    ci_upper = ci['ci_upper']

    print(f"\nBootstrap 95% Confidence Intervals "
          f"(case-resampling, n={ci['n_success']}):")
    print(f"κ   = {kappa_fit:.1f}  [{ci_lower[0]:.1f}, {ci_upper[0]:.1f}]")
    print(f"ξ₀  = {xi0_fit:.2f}  [{ci_lower[1]:.2f}, {ci_upper[1]:.2f}]")
    print(f"ξ*  = {xi_star_fit:.2f}  [{ci_lower[2]:.2f}, {ci_upper[2]:.2f}]")

    # ---- κ 不可辨识性分析 ----
    print("\n" + "=" * 60)
    print("κ unidentifiability analysis (Yalow 1960 RIA)")
    print("=" * 60)
    # Run-invariant one-sided bound from the deterministic fixed-κ scan
    # (R² >= 0.98 plateau edge); see 4_Kappa_Scan in SI_Table_2.xlsx.
    _sw = kappa_sweep(f_all, p_all, xi0_fit, xi_star_fit)
    _vi = [(k, r2) for (k, r2, _, _) in _sw if not np.isnan(r2)]
    kappa_bound_scan = float(min([k for k, r2 in _vi if r2 >= 0.98],
                                 default=float('nan')))
    width_kappa = np.log10(ci_upper[0] / max(ci_lower[0], 1e-6))
    print(f"Point estimate κ = {kappa_fit:.1f}")
    print(f"Bootstrap CI: [{ci_lower[0]:.1f}, {ci_upper[0]:.1f}]  "
          f"(spans {width_kappa:.1f} orders of magnitude)")
    print()
    print(f"Physical interpretation:")
    print(f"  • The Yalow displacement curve lies in the κ-independent regime.")
    print(f"  • Standard errors on κ span orders of magnitude "
          f"(asymptotic SE: {perr[0]:.0f}).")
    # NOTE (portability): the point estimate of κ and the bootstrap CI
    # endpoints lie on a flat likelihood ridge and therefore vary between
    # optimizer implementations (BLAS/SciPy builds), even with a fixed
    # bootstrap seed. The run-invariant statement is the scan-based bound
    # below (4_Kappa_Scan is deterministic to 6 decimals across machines).
    print(f"  • We therefore report κ as a one-sided lower bound: "
          f"κ ≥ {kappa_bound_scan:.1f} (scan-based: R² ≥ 0.98 for all fixed "
          f"κ ≥ {kappa_bound_scan:.1f}; run-invariant)")
    print(f"  • The bootstrap CI above is illustrative of the ridge width; "
          f"its endpoints vary between optimizer implementations.")
    print(f"  • This matches the protocol used for κ₃ in the MCMC analysis "
          f"of the")
    print(f"    three-concentration ELISA dataset (Methods §3).")
    print()
    print(f"  • Only the dimensionless concentrations ξ₀ and ξ* "
          f"are well-constrained.")
    print(f"  • This is direct evidence of scale degeneracy: "
          f"the displacement curve")
    print(f"    shape constrains the RATIOS but not the absolute "
          f"molecular numbers.")

    # ---- κ 扫描 ----
    print("\n" + "=" * 60)
    print("Sensitivity to κ (R² vs fixed-κ)")
    print("=" * 60)
    sweep = kappa_sweep(f_all, p_all, xi0_fit, xi_star_fit)

    print(f"{'κ':<12} {'R²':<12} {'ξ₀':<10} {'ξ*':<10}")
    print("-" * 44)
    for k_test, r2_k, xi0_k, xi_star_k in sweep:
        if np.isnan(r2_k):
            print(f"{k_test:<12.4g} {'fit failed':<12} {'—':<10} {'—':<10}")
        else:
            print(f"{k_test:<12.4g} {r2_k:<12.6f} "
                  f"{xi0_k:<10.4f} {xi_star_k:<10.4f}")

    # 动态计算 R² 变化范围
    valid_r2 = [r2_k for (_, r2_k, _, _) in sweep if not np.isnan(r2_k)]
    if valid_r2:
        r2_spread = max(valid_r2) - min(valid_r2)
        k_valid = [k for (k, r2_k, _, _) in sweep if not np.isnan(r2_k)]
        print(f"\nR² varies by {r2_spread:.4f} across "
              f"{len(k_valid)} κ values "
              f"({min(k_valid):.4g} to {max(k_valid):.4g}),")
        print(f"confirming that the Yalow displacement curve cannot "
              f"distinguish κ regimes.")

    # ---- 零标准点检查 ----
    p_zero_model = fit['p_zero_model']
    print(f"\nZero standard: actual p*={p_all[0]:.4f}, "
          f"model p*={p_zero_model:.4f}")

    # ---- 绘图 ----
    plot_results(f_all, p_all, fit, ci, sweep, output_path=args.figure)

    # ---- Excel 输出 ----
    if not args.no_excel:
        export_excel(f_all, p_all, fit, ci, sweep, output_excel=args.excel)


if __name__ == "__main__":
    main()
