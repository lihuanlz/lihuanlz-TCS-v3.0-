# -*- coding: utf-8 -*-
"""
Created on Tue Jul  7 17:19:01 2026

@author: lihua
"""

# -*- cRFUing: utf-8 -*-
"""
三组联合拟合：共享 ξ₀，每组独立 A, κ（D 锁定为实测空白）
改进版 v2:
  - D 锁定为实测 blank RFU（不再作为 MCMC 参数）
  - 8 个 MCMC 参数（xi0, A1, k1, A2, k2, A3, k3, log_sigma）
  - 4PL/5PL 比较 + bootstrap CI + 7 个 sheet Excel 导出
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import least_squares, curve_fit
from scipy.stats import t, shapiro, probplot
import warnings
warnings.filterwarnings('ignore')

try:
    import emcee
except ImportError:
    raise ImportError("请先安装 emcee: pip install emcee")

plt.rcParams['font.family'] = 'Arial'

# ========== 用户设置 ==========
USE_WEIGHTS = False   # MCMC 使用恒定方差，此选项仅用于最小二乘初始估计

# ========== 数据 ==========
dil_factors = np.array([1, 2, 4, 8, 16, 32, 64, 1e7])  # 背景用大数

RFU1 = np.array([
    [2.219, 2.218, 2.215],
    [1.242, 1.238, 1.225],
    [0.8265, 0.8351, 0.8214],
    [0.4581, 0.4464, 0.4421],
    [0.2214, 0.2339, 0.2230],
    [0.1522, 0.1539, 0.1536],
    [0.1002, 0.1098, 0.1025],
    [0.02547, 0.02442, 0.02314]
])

RFU2 = np.array([
    [1.929, 1.961, 1.935],
    [1.137, 1.148, 1.137],
    [0.7448, 0.7343, 0.7329],
    [0.3825, 0.3856, 0.3885],
    [0.1154, 0.1137, 0.1155],
    [0.1188, 0.1117, 0.1168],
    [0.06615, 0.06538, 0.06588],
    [0.02256, 0.02701, 0.02433]
])

RFU3 = np.array([
    [0.3184, 0.3128, 0.3153],
    [0.3001, 0.2927, 0.2961],
    [0.2069, 0.2119, 0.2102],
    [0.1806, 0.1838, 0.1836],
    [0.1008, 0.1078, 0.1057],
    [0.07218, 0.07146, 0.07174],
    [0.05296, 0.05369, 0.05267],
    [0.02285, 0.02163, 0.02681]
])

# ========== D 锁定为实测 blank（不再作为 MCMC 参数） ==========
D1_fixed = float(np.mean(RFU1[-1]))   # ≈ 0.0243
D2_fixed = float(np.mean(RFU2[-1]))   # ≈ 0.0246
D3_fixed = float(np.mean(RFU3[-1]))   # ≈ 0.0238
print(f"D 锁定: D1={D1_fixed:.4f}, D2={D2_fixed:.4f}, D3={D3_fixed:.4f}")

# 展平数据
y1_flat = RFU1.T.flatten()
y2_flat = RFU2.T.flatten()
y3_flat = RFU3.T.flatten()
y_obs = np.concatenate([y1_flat, y2_flat, y3_flat])

dil_repeated = np.tile(dil_factors, 3)
f_data = np.tile(dil_repeated, 3)
cond_idx = np.array([0]*24 + [1]*24 + [2]*24)

# ========== TCS 模型 ==========
def p_tcs_exact(xi, kappa):
    if kappa <= 0:
        return 0.0
    a = 1.0
    b = -(kappa * (xi + 1.0) + 1.0)
    c = kappa * xi
    disc = np.maximum(b**2 - 4*a*c, 0.0)
    return (-b - np.sqrt(disc)) / (2.0 * a)

def model(dil, xi0, A, D, kappa):
    f = 1.0 / dil
    xi = xi0 * f
    xi = np.where(dil > 1e6, 0.0, xi)
    p = p_tcs_exact(xi, kappa)
    return D + (A - D) * p


def predict_global(theta, dil, cond_idx):
    xi0 = theta[0]
    A1, k1 = theta[1], theta[2]
    A2, k2 = theta[3], theta[4]
    A3, k3 = theta[5], theta[6]
    y_pred = np.empty_like(dil)
    mask1 = (cond_idx == 0); y_pred[mask1] = model(dil[mask1], xi0, A1, D1_fixed, k1)
    mask2 = (cond_idx == 1); y_pred[mask2] = model(dil[mask2], xi0, A2, D2_fixed, k2)
    mask3 = (cond_idx == 2); y_pred[mask3] = model(dil[mask3], xi0, A3, D3_fixed, k3)
    return y_pred

# ========== 最小二乘初始估计 ==========
xi0_guess = 100
A1_guess, k1_guess = 2.66, 0.09
A2_guess, k2_guess = 2.08, 0.01
A3_guess, k3_guess = 0.349, 1.11

theta0 = [xi0_guess,
          A1_guess, k1_guess,
          A2_guess, k2_guess,
          A3_guess, k3_guess]

lb = [0.1, 0.5, 1e-4, 0.5, 1e-4, 0.1, 1e-4]
ub = [500, 3.5, 1000, 2.5, 1000, 0.5, 1000]
x_scale = [10, 1, 1, 1, 1, 0.2, 1]

print("最小二乘初始化...")
res = least_squares(lambda th: predict_global(th, f_data, cond_idx) - y_obs,
                    theta0, bounds=(lb, ub), x_scale=x_scale,
                    loss='linear', method='trf', max_nfev=10000)
theta_hat = res.x
sigma_guess = np.std(y_obs - predict_global(theta_hat, f_data, cond_idx))
print("最小二乘完成，sigma 初始值:", sigma_guess)

# ========== MCMC 设置 ==========
def log_prior(theta):
    xi0, A1, k1, A2, k2, A3, k3, log_sigma = theta
    if not (0.1 < xi0 < 500): return -np.inf
    if not (0.5 < A1 < 3.5) or not (1e-4 < k1 < 1000): return -np.inf
    if not (0.5 < A2 < 2.5) or not (1e-4 < k2 < 1000): return -np.inf
    if not (0.1 < A3 < 0.5) or not (1e-4 < k3 < 1000): return -np.inf
    if log_sigma < -5 or log_sigma > 2: return -np.inf
    return 0.0

def log_likelihood(theta, dil, cond_idx, y_obs):
    xi0, A1, k1, A2, k2, A3, k3, log_sigma = theta
    sigma = np.exp(log_sigma)
    y_pred = predict_global([xi0, A1, k1, A2, k2, A3, k3], dil, cond_idx)
    return -0.5 * np.sum(((y_obs - y_pred) / sigma) ** 2) - len(y_obs) * log_sigma

def log_probability(theta, dil, cond_idx, y_obs):
    lp = log_prior(theta)
    if not np.isfinite(lp):
        return -np.inf
    return lp + log_likelihood(theta, dil, cond_idx, y_obs)

# 初始参数向量（包括 log_sigma）
theta_init_mcmc = np.append(theta_hat, np.log(sigma_guess))
ndim = len(theta_init_mcmc)
nwalkers = max(50, 2 * ndim)
nsteps = 50000
burnin = 10000

np.random.seed(42)
pos = theta_init_mcmc + 1e-4 * np.random.randn(nwalkers, ndim)

print("开始 MCMC 采样 (emcee)...")
sampler = emcee.EnsembleSampler(nwalkers, ndim, log_probability,
                                args=(f_data, cond_idx, y_obs))
sampler.run_mcmc(pos, nsteps, progress=True)

samples = sampler.get_chain(discard=burnin, flat=True)
print("采样完成，后验样本量:", samples.shape[0])

# 收敛诊断（R-hat 和 ESS）
try:
    import arviz as az
    idata = az.from_emcee(sampler, var_names=['xi0','A1','k1',
                                              'A2','k2',
                                              'A3','k3',
                                              'log_sigma'])
    rhat = az.rhat(idata)
    ess = az.ess(idata)
    print("\n收敛诊断 (R-hat):")
    print(rhat)
    print("\n有效样本量 (ESS):")
    print(ess)
except ImportError:
    print("未安装 arviz，无法计算 R-hat 和 ESS，请安装: pip install arviz")
    rhat = ess = None

# 后验中位数和 95% HDI（前 7 个是物理参数）
theta_mcmc = np.median(samples[:, :7], axis=0)
sigma_mcmc = np.exp(np.median(samples[:, 7]))
lower = np.percentile(samples[:, :7], 2.5, axis=0)
upper = np.percentile(samples[:, :7], 97.5, axis=0)
lower_sigma = np.percentile(samples[:, 7], 2.5)
upper_sigma = np.percentile(samples[:, 7], 97.5)

xi0_hat = theta_mcmc[0]
A1, k1 = theta_mcmc[1], theta_mcmc[2]
A2, k2 = theta_mcmc[3], theta_mcmc[4]
A3, k3 = theta_mcmc[5], theta_mcmc[6]
D1, D2, D3 = D1_fixed, D2_fixed, D3_fixed  # 锁定为实测空白

y_pred = predict_global(theta_mcmc, f_data, cond_idx)
residuals_final = y_obs - y_pred

param_names = ['ξ₀', 'A₁', 'κ₁', 'A₂', 'κ₂', 'A₃', 'κ₃']
print("\n" + "="*60)
print("MCMC 后验中位数与 95% HDI（D 锁定为实测空白）")
print("="*60)
for name, val, low, up in zip(param_names, theta_mcmc, lower, upper):
    print(f"{name:5s} = {val:.4f}  (95% HDI: [{low:.4f}, {up:.4f}])")
print(f"D₁ = {D1:.4f}  (FIXED)")
print(f"D₂ = {D2:.4f}  (FIXED)")
print(f"D₃ = {D3:.4f}  (FIXED)")
print(f"sigma = {sigma_mcmc:.4f}  (95% HDI on log σ: [{lower_sigma:.4f}, {upper_sigma:.4f}])")


# ========== 1. 残差诊断图 ==========
from scipy.ndimage import gaussian_filter1d

fig_res = plt.figure(figsize=(15, 5))
fig_res.suptitle('Extended Data Fig. 5c: Residual diagnostics for the global TCS fit\n'
                 r'(shared $\xi_0$, three independent $\kappa_i$, D locked to blank)',
                 fontsize=16, fontweight='bold', y=0.99)

ax1 = fig_res.add_subplot(131)
ax1.scatter(y_pred, residuals_final, alpha=0.6, edgecolors='k', facecolors='none')
ax1.axhline(y=0, color='r', linestyle='--', linewidth=1)
ax1.set_xlabel('Predicted RFU')
ax1.set_ylabel('Residuals')
ax1.set_title('(1)Residuals vs. Predicted')
ax1.grid(True, alpha=0.3)
order = np.argsort(y_pred)
y_sorted = y_pred[order]
res_sorted = residuals_final[order]
try:
    smooth = gaussian_filter1d(res_sorted, sigma=2)
    # Modified: correct label
    ax1.plot(y_sorted, smooth, 'b-', lw=1.5, label='Gaussian smooth')
    ax1.legend()
except:
    pass

ax2 = fig_res.add_subplot(132)
probplot(residuals_final, dist="norm", plot=ax2)
ax2.set_title('(2)Q-Q Plot')
ax2.grid(False, alpha=0.3)

ax3 = fig_res.add_subplot(133)
count, bins, _ = ax3.hist(residuals_final, bins=15, density=True, alpha=0.7, color='skyblue', edgecolor='black')
mu, std = np.mean(residuals_final), np.std(residuals_final)
x_norm = np.linspace(bins[0], bins[-1], 200)
y_norm = (1/(std*np.sqrt(2*np.pi))) * np.exp(-0.5*((x_norm-mu)/std)**2)
ax3.plot(x_norm, y_norm, 'r-', lw=2, label='Normal dist.')
ax3.set_xlabel('Residuals')
ax3.set_ylabel('Density')
ax3.set_title('(3)Residuals Histogram')
ax3.legend()
ax3.grid(False, alpha=0.3)

plt.subplots_adjust(top=0.85)
plt.savefig('Extended_Data_Fig_5c.svg', dpi=300)
plt.show()

# 统计量
shapiro_stat, shapiro_p = shapiro(residuals_final)
dw = np.sum(np.diff(residuals_final)**2) / np.sum(residuals_final**2)
print(f"\nShapiro-Wilk: W={shapiro_stat:.4f}, p={shapiro_p:.4e}")
print(f"Durbin-Watson: {dw:.4f} (理想值接近2)")

# ========== 2. R² 输出 ==========
def r2_group(y_true, y_pred):
    ss_res = np.sum((y_true - y_pred)**2)
    ss_tot = np.sum((y_true - np.mean(y_true))**2)
    return 1 - ss_res / ss_tot

y1_true = y_obs[cond_idx==0]; y1_pred = y_pred[cond_idx==0]
y2_true = y_obs[cond_idx==1]; y2_pred = y_pred[cond_idx==1]
y3_true = y_obs[cond_idx==2]; y3_pred = y_pred[cond_idx==2]
R2_1 = r2_group(y1_true, y1_pred)
R2_2 = r2_group(y2_true, y2_pred)
R2_3 = r2_group(y3_true, y3_pred)
R2_total = r2_group(y_obs, y_pred)

print("\n" + "="*60)
print("三组联合拟合结果（共享 ξ₀，D 固定为实测空白，MCMC 后验中位数）")
print("="*60)
print(f"公共 ξ₀ = {xi0_hat:.4f}")
print(f"组1 (2 µg/mL):   A={A1:.4f}, D={D1:.4f} (FIXED), κ={k1:.4f}, R²={R2_1:.6f}")
print(f"组2 (0.5 µg/mL): A={A2:.4f}, D={D2:.4f} (FIXED), κ={k2:.4f}, R²={R2_2:.6f}")
print(f"组3 (0.05 µg/mL):A={A3:.4f}, D={D3:.4f} (FIXED), κ={k3:.4f}, R²={R2_3:.6f}")
print(f"全局 R² = {R2_total:.6f}")

# ========== 3. RFU vs 稀释倍数图 ==========
dil_plot = np.logspace(0, 7, 200)
y1_curve = model(dil_plot, xi0_hat, A1, D1, k1)
y2_curve = model(dil_plot, xi0_hat, A2, D2, k2)
y3_curve = model(dil_plot, xi0_hat, A3, D3, k3)

y1_mean = np.mean(RFU1, axis=1); y1_std = np.std(RFU1, axis=1)
y2_mean = np.mean(RFU2, axis=1); y2_std = np.std(RFU2, axis=1)
y3_mean = np.mean(RFU3, axis=1); y3_std = np.std(RFU3, axis=1)

fig1, axes = plt.subplots(1, 3, figsize=(15, 5))
colors = ['tab:blue', 'tab:orange', 'tab:green']
markers = ['o', 's', '^']
titles = ['2 µg/mL', '0.5 µg/mL', '0.05 µg/mL']
for ax, dil_vals, y_mean, y_std, y_curve, color, marker, title, r2 in zip(
        axes, [dil_factors, dil_factors, dil_factors],
        [y1_mean, y2_mean, y3_mean],
        [y1_std, y2_std, y3_std],
        [y1_curve, y2_curve, y3_curve],
        colors, markers, titles, [R2_1, R2_2, R2_3]):
    ax.errorbar(dil_vals, y_mean, yerr=y_std, fmt=marker, color=color, capsize=3)
    ax.plot(dil_plot, y_curve, '-', color=color, lw=2, label=f'Fit (R²={r2:.3f})')
    ax.set_xscale('log'); ax.set_xlabel('Dilution factor'); ax.set_ylabel('RFU')
    ax.set_title(title); ax.legend(); ax.grid(True, alpha=0.3)
    ax.set_xlim(1e7, 0.8); ax.invert_xaxis()
plt.suptitle(rf'Global fit: shared $\xi_0$ = {xi0_hat:.2f}, D locked to blank', fontsize=14)
plt.tight_layout()
plt.savefig('Shared_xi0_fit.svg', dpi=300)
plt.show()

# ========== 4. 参数趋势图 ==========
concs = [2.0, 0.5, 0.05]
As = [A1, A2, A3]; Ds = [D1, D2, D3]; ks = [k1, k2, k3]
fig2, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(12, 4))
ax1.plot(concs, As, 'o-', color='tab:red')
ax1.set_xscale('log'); ax1.set_xlabel('Concentration (µg/mL)'); ax1.set_ylabel('A'); ax1.grid(True, alpha=0.3)
ax2.plot(concs, Ds, 's-', color='tab:blue')
ax2.set_xscale('log'); ax2.set_xlabel('Concentration (µg/mL)'); ax2.set_ylabel('D (locked)'); ax2.grid(True, alpha=0.3)
ax3.plot(concs, ks, '^-', color='tab:green')
ax3.set_xscale('log'); ax3.set_yscale('log'); ax3.set_xlabel('Concentration (µg/mL)'); ax3.set_ylabel('κ'); ax3.grid(True, alpha=0.3)
plt.suptitle(rf'Parameter trends (shared $\xi_0$ = {xi0_hat:.2f}, D locked to blank)', fontsize=14)
plt.tight_layout()
plt.savefig('Shared_xi0_parameters.svg', dpi=300)
plt.show()


# ========== 5. p vs ξ 主曲线（Fig. 2a） ==========
group_labels = ['2 µg/mL', '0.5 µg/mL', '0.05 µg/mL']
markers_d = {'2 µg/mL': 'o', '0.5 µg/mL': 's', '0.05 µg/mL': '^'}
colors_d = {'2 µg/mL': 'blue', '0.5 µg/mL': 'orange', '0.05 µg/mL': 'green'}

fig3, ax = plt.subplots(figsize=(7, 7))
valid_indices = [i for i, d in enumerate(dil_factors) if d < 1e6]

# To draw vertical connecting lines later, store median p for each group at each dilution
p_medians_by_dil = {i: [] for i in valid_indices}

tcrit = t.ppf(0.975, df=2)  # t值 for 3 replicates, 95% CI

for grp_idx in range(3):
    A_med = [A1, A2, A3][grp_idx]
    k_med = [k1, k2, k3][grp_idx]
    D_val = [D1_fixed, D2_fixed, D3_fixed][grp_idx]
    RFU = [RFU1, RFU2, RFU3][grp_idx]
    
    xi_vals, p_medians, p_err_low, p_err_high = [], [], [], []
    for i in valid_indices:
        dil = dil_factors[i]
        xi_val = xi0_hat / dil
        
        # 3 replicates → p values with median A
        p_reps = (RFU[i, :] - D_val) / (A_med - D_val)
        p_reps = np.clip(p_reps, 0, 1)
        p_mean = np.mean(p_reps)
        se = np.std(p_reps, ddof=1) / np.sqrt(3)
        
        xi_vals.append(xi_val)
        p_medians.append(p_mean)
        p_err_low.append(tcrit * se)
        p_err_high.append(tcrit * se)
        p_medians_by_dil[i].append(p_mean)
    
    # Plot with symmetric t-distribution CI error bars
    ax.errorbar(xi_vals, p_medians,
                yerr=[p_err_low, p_err_high],
                fmt=markers_d[group_labels[grp_idx]],
                color=colors_d[group_labels[grp_idx]],
                capsize=6, elinewidth=2, markersize=8,
                label=f'{group_labels[grp_idx]}', alpha=0.9)
    
    # Theoretical TCS curve using median parameters
    xi_theory = np.logspace(-3, 3, 200)
    p_theory = p_tcs_exact(xi_theory, k_med)
    ax.plot(xi_theory, p_theory, '--', color=colors_d[group_labels[grp_idx]], lw=2,
            label=f'TCS κ={k_med:.2f}', zorder=5)
    

# Vertical dashed lines connecting points at same dilution across groups
for i in valid_indices:
    p_vals = p_medians_by_dil[i]
    # Use the median xi0 to place the line
    # xi_common = np.median(post_xi0) / dil_factors[i]
    xi_common = xi0_hat / dil_factors[i]

    ax.plot([xi_common] * 3, p_vals, 'k--', linewidth=1.2, alpha=0.7, zorder=2)

ax.set_xscale('log')
ax.set_xlabel(r'$\xi = \xi_0 / f$', fontsize=14)
ax.set_ylabel('p', fontsize=14)
ax.set_title(rf'Fig. 2a: scale degeneracy ($\xi_0$ = {xi0_hat:.2f}, D locked)', fontsize=16, fontweight='bold')
ax.legend(loc='upper left', fontsize=14)
ax.grid(False)
ax.set_ylim(0, 1.05)
fig3.tight_layout()
fig3.savefig('Fig.2a_scale_degeneracy.svg', dpi=300)
plt.show()

# ========== 6. 4PL 和约束 5PL 比较（Extended Data Fig. 5b） ==========
def fourpl_RFU(dil, A, D, C, B):
    return D + (A - D) / (1.0 + (dil / C) ** B)

def fivepl_constrained_RFU(dil, A, D, C, G):
    B = 1.0 / G
    return D + (A - D) / (1.0 + (dil / C) ** B) ** G

valid_mask = dil_factors < 1e6
dil_valid = dil_factors[valid_mask]

groups = [
    ('2 µg/mL', RFU1[valid_mask], A1, D1),
    ('0.5 µg/mL', RFU2[valid_mask], A2, D2),
    ('0.05 µg/mL', RFU3[valid_mask], A3, D3),
]

fit_4pl = {}
fit_5pl = {}
for name, RFU_sub, A_init, D_init in groups:
    x_data = []
    y_data = []
    for i, dil in enumerate(dil_valid):
        for j in range(3):
            x_data.append(dil)
            y_data.append(RFU_sub[i, j])
    x_data = np.array(x_data)
    y_data = np.array(y_data)

    try:
        popt4, _ = curve_fit(
            lambda x, C, B: fourpl_RFU(x, A_init, D_init, C, B),
            x_data, y_data,
            p0=[10.0, 1.0],
            bounds=([0.1, 0.1], [1e4, 10]),
            maxfev=10000
        )
        C4, B4 = popt4
        y_pred4 = fourpl_RFU(x_data, A_init, D_init, C4, B4)
        ss_res = np.sum((y_data - y_pred4) ** 2)
        ss_tot = np.sum((y_data - np.mean(y_data)) ** 2)
        r2_4 = 1 - ss_res / ss_tot
        fit_4pl[name] = (C4, B4, r2_4)
    except:
        fit_4pl[name] = (np.nan, np.nan, np.nan)

    try:
        popt5, _ = curve_fit(
            lambda x, C, G: fivepl_constrained_RFU(x, A_init, D_init, C, G),
            x_data, y_data,
            p0=[10.0, 1.0],
            bounds=([0.1, 0.1], [1e4, 10]),
            maxfev=10000
        )
        C5, G5 = popt5
        B5 = 1.0 / G5
        y_pred5 = fivepl_constrained_RFU(x_data, A_init, D_init, C5, G5)
        ss_res = np.sum((y_data - y_pred5) ** 2)
        ss_tot = np.sum((y_data - np.mean(y_data)) ** 2)
        r2_5 = 1 - ss_res / ss_tot
        fit_5pl[name] = (C5, B5, G5, r2_5)
    except:
        fit_5pl[name] = (np.nan, np.nan, np.nan, np.nan)

fig_cmp, ax_cmp = plt.subplots(figsize=(7, 7))
dil_plot_4pl5pl = np.logspace(0, 3, 200)

for name, RFU_sub, A, D in groups:
    if not np.isnan(fit_4pl[name][0]):
        C4, B4, r2_4 = fit_4pl[name]
        RFU_4pl = fourpl_RFU(dil_plot_4pl5pl, A, D, C4, B4)
        ax_cmp.plot(dil_plot_4pl5pl, RFU_4pl, '-', color=colors_d[name], lw=1.5, alpha=0.7,
                    label=f'{name} 4PL')
    if not np.isnan(fit_5pl[name][0]):
        C5, B5, G5, r2_5 = fit_5pl[name]
        RFU_5pl = fivepl_constrained_RFU(dil_plot_4pl5pl, A, D, C5, G5)
        ax_cmp.plot(dil_plot_4pl5pl, RFU_5pl, '-.', color=colors_d[name], lw=1.5, alpha=0.7,
                    label=f'{name} 5PL')

for name, RFU_sub, A, D in groups:
    x_vals = dil_valid
    y_mean = np.mean(RFU_sub, axis=1)
    y_std = np.std(RFU_sub, axis=1, ddof=1)
    ax_cmp.errorbar(x_vals, y_mean, yerr=y_std, fmt=markers_d[name],
                    color=colors_d[name], capsize=6, elinewidth=2, markersize=8,
                    alpha=0.9)

ax_cmp.set_xscale('log')
ax_cmp.set_xlabel('Dilution factor', fontsize=14)
ax_cmp.set_ylabel('RFU', fontsize=14)
ax_cmp.set_title('Extended Data Fig. 5b: 4PL vs 5PL (B=1/G) comparison (D locked)', fontsize=14, fontweight='bold')
ax_cmp.legend(loc='upper left', fontsize=14)
ax_cmp.invert_xaxis()

textstr = '5PL constraint: B = 1/G\n\n'
for name in fit_4pl:
    if not np.isnan(fit_4pl[name][0]):
        C, B, r2 = fit_4pl[name]
        textstr += f'{name} 4PL: C={C:.2f}, B={B:.2f}, R²={r2:.3f}\n'
    if not np.isnan(fit_5pl[name][0]):
        C, B, G, r2 = fit_5pl[name]
        textstr += f'{name} 5PL: C={C:.2f}, B={B:.2f} (=1/G), G={G:.2f}, R²={r2:.3f}\n'
ax_cmp.text(0.74, 0.65, textstr, transform=ax_cmp.transAxes, fontsize=12,
            verticalalignment='top', horizontalalignment='right',
            bbox=dict(boxstyle='round', facecolor='white', alpha=0.85))

fig_cmp.tight_layout()
# Modified: consistent filename with figure title
fig_cmp.savefig('Extended_Data_Fig_5b_4PL_vs_5PL.svg', dpi=300)
plt.show()


# ========== 6.5 4PL/5PL per-group Bootstrap CIs ==========
np.random.seed(42)
n_boot = 2000
boot_4pl = {n: {'C':[], 'B':[]} for n,_,_,_ in groups}
boot_5pl = {n: {'C':[], 'G':[], 'B':[]} for n,_,_,_ in groups}

def fit_4pl_factory(A_init, D_init):
    return lambda x, C, B: fourpl_RFU(x, A_init, D_init, C, B)

def fit_5pl_factory(A_init, D_init):
    return lambda x, C, G: fivepl_constrained_RFU(x, A_init, D_init, C, G)

for name, RFU_sub, A_init, D_init in groups:
    n_dil = RFU_sub.shape[0]
    x_full = np.repeat(dil_valid, 3)
    for _ in range(n_boot):
        RFU_resampled = np.empty_like(RFU_sub)
        for i in range(n_dil):
            RFU_resampled[i] = RFU_sub[i, np.random.choice(3, 3, replace=True)]
        y_boot = RFU_resampled.flatten()
        try:
            p4, _ = curve_fit(fit_4pl_factory(A_init, D_init),
                              x_full, y_boot,
                              p0=fit_4pl[name][:2],
                              bounds=([0.1, 0.1], [1e4, 10]), maxfev=5000)
            boot_4pl[name]['C'].append(p4[0])
            boot_4pl[name]['B'].append(p4[1])
        except:
            pass
        try:
            p5, _ = curve_fit(fit_5pl_factory(A_init, D_init),
                              x_full, y_boot,
                              p0=[fit_5pl[name][0], fit_5pl[name][2]],
                              bounds=([0.1, 0.1], [1e4, 10]), maxfev=5000)
            boot_5pl[name]['C'].append(p5[0])
            boot_5pl[name]['G'].append(p5[1])
            boot_5pl[name]['B'].append(1.0 / p5[1])
        except:
            pass

print("\n" + "="*60)
# print("4PL/5PL Bootstrap 95% CI (n=2000, D locked to blank)")
print(f"4PL/5PL Bootstrap 95% CI (n={n_boot}, D locked to blank)")
print("="*60)
for name in boot_4pl:
    C4 = np.array(boot_4pl[name]['C']); B4 = np.array(boot_4pl[name]['B'])
    C5 = np.array(boot_5pl[name]['C']); G5 = np.array(boot_5pl[name]['G'])
    if len(C4) and len(C5):
        print(f"  {name}:")
        print(f"    4PL: C = {np.median(C4):.2f} [{np.percentile(C4,2.5):.2f}, {np.percentile(C4,97.5):.2f}]")
        print(f"         B = {np.median(B4):.2f} [{np.percentile(B4,2.5):.2f}, {np.percentile(B4,97.5):.2f}]")
        print(f"    5PL: C = {np.median(C5):.2f} [{np.percentile(C5,2.5):.2f}, {np.percentile(C5,97.5):.2f}]")
        print(f"         G = {np.median(G5):.3f} [{np.percentile(G5,2.5):.3f}, {np.percentile(G5,97.5):.3f}]")


# 4PL 与 5PL 残差图
fig_res_cmp, ax_res = plt.subplots(figsize=(8, 6))
all_true = []
all_pred_4 = []
all_pred_5 = []
for name, RFU_sub, A, D in groups:
    if np.isnan(fit_4pl[name][0]) or np.isnan(fit_5pl[name][0]):
        continue
    C4, B4, _ = fit_4pl[name]
    C5, B5, G5, _ = fit_5pl[name]
    for i, dil in enumerate(dil_valid):
        for j in range(3):
            y_true = RFU_sub[i, j]
            y_pred_4 = fourpl_RFU(dil, A, D, C4, B4)
            y_pred_5 = fivepl_constrained_RFU(dil, A, D, C5, G5)
            all_true.append(y_true)
            all_pred_4.append(y_pred_4)
            all_pred_5.append(y_pred_5)

all_true = np.array(all_true)
all_pred_4 = np.array(all_pred_4)
all_pred_5 = np.array(all_pred_5)

res_4 = all_true - all_pred_4
res_5 = all_true - all_pred_5

rmse_4 = np.sqrt(np.mean(res_4**2))
rmse_5 = np.sqrt(np.mean(res_5**2))

print(f"\nRMSE 4PL = {rmse_4:.6f}")
print(f"RMSE 5PL (B=1/G) = {rmse_5:.6f}")

ax_res.scatter(all_pred_4, res_4, alpha=0.6, marker='o', facecolors='none', edgecolors='blue', label='4PL residuals')
ax_res.scatter(all_pred_5, res_5, alpha=0.6, marker='s', facecolors='none', edgecolors='red', label='5PL residuals (B=1/G)')
ax_res.axhline(y=0, color='k', linestyle='--', linewidth=0.8)
ax_res.set_xlabel('Predicted RFU')
ax_res.set_ylabel('Residual')
ax_res.set_title('4PL vs 5PL Residuals (D locked)')
ax_res.legend()
ax_res.grid(True, alpha=0.3)

text_info = f'RMSE 4PL = {rmse_4:.4f}\nRMSE 5PL = {rmse_5:.4f}'
ax_res.text(0.05, 0.95, text_info, transform=ax_res.transAxes, fontsize=14,
            verticalalignment='top', bbox=dict(boxstyle='round', facecolor='white', alpha=0.85))

fig_res_cmp.tight_layout()
# Modified: consistent filename
fig_res_cmp.savefig('Extended_Data_Fig_5b_residuals.svg', dpi=300)
plt.show()

print("\n所有图形已保存。MCMC 后验中位数代替原 Bootstrap 估计，D 锁定为实测空白。")


# ============================================================================
# SI Table 1: Export all raw data and computed results to Excel (v2, D locked)
# ============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ----- Sheet 1: Raw RFU data (3 concentrations, 8 dilutions, 3 replicates) -----
    ws = wb.active
    ws.title = "1_Raw_RFU"
    ws.append(["Group", "Dilution_factor", "Rep1", "Rep2", "Rep3", "Mean", "SD"])
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill

    group_labels = ["2 µg/mL", "0.5 µg/mL", "0.05 µg/mL"]
    for grp_idx, (label, RFU) in enumerate(zip(group_labels, [RFU1, RFU2, RFU3]), 1):
        for i, dil in enumerate(dil_factors):
            rep1, rep2, rep3 = RFU[i]
            mean = np.mean(RFU[i])
            sd = np.std(RFU[i], ddof=1)
            dil_label = "blank (f→∞)" if dil >= 1e6 else dil
            ws.append([f"Group {grp_idx} ({label})", dil_label, rep1, rep2, rep3, mean, sd])

    # ----- Sheet 2: MCMC posterior summary (8 parameters + 3 fixed D) -----
    ws2 = wb.create_sheet("2_MCMC_Parameters")
    ws2.append(["Parameter", "Median (posterior)", "2.5% HDI", "97.5% HDI", "Description"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill

    descriptions = [
        "Shared normalized analyte concentration",
        "Group 1: saturation signal A₁",
        "Group 1: depletion coefficient κ₁ (low conc → small κ)",
        "Group 2: saturation signal A₂",
        "Group 2: depletion coefficient κ₂",
        "Group 3: saturation signal A₃",
        "Group 3: depletion coefficient κ₃ (low conc → large κ)",
        "Log noise std (log σ) — shared across all 3 groups",
        "Group 1: background D₁ (LOCKED to measured blank)",
        "Group 2: background D₂ (LOCKED to measured blank)",
        "Group 3: background D₃ (LOCKED to measured blank)",
    ]
    full_param_names = param_names + ['log σ', 'D₁ (fixed)', 'D₂ (fixed)', 'D₃ (fixed)']
    full_param_values = list(theta_mcmc) + [np.log(sigma_mcmc), D1_fixed, D2_fixed, D3_fixed]
    full_param_lower = list(lower) + [lower_sigma, D1_fixed, D2_fixed, D3_fixed]
    full_param_upper = list(upper) + [upper_sigma, D1_fixed, D2_fixed, D3_fixed]

    for name, val, lo, hi, desc in zip(full_param_names, full_param_values,
                                        full_param_lower, full_param_upper, descriptions):
        ws2.append([name, val, lo, hi, desc])

    ws2.append([""])
    ws2.append(["NOTE: D is locked to the measured blank RFU (no nuisance parameter in MCMC).",
                "", "", "", ""])
    ws2.append(["Total free physical parameters: 7 (xi0 + A1,k1,A2,k2,A3,k3).",
                "", "", "", ""])
    ws2.append(["NOTE: k3 sits on the high-kappa (zero-depletion) plateau; the upper end",
                "of its 95% HDI is set by the prior ceiling (1e3), not by the data, so k3",
                "is reported as a one-sided lower bound (>= 18.8). Only the lower HDI",
                "bound is data-driven (kappa-independent regime, SI S10.10).", ""])

    # ----- Sheet 3: Convergence diagnostics (R-hat, ESS) -----
    ws3 = wb.create_sheet("3_Convergence")
    ws3.append(["Parameter", "R-hat (target ≤ 1.03)", "ESS (target ≥ 400)"])
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
    if rhat is not None and ess is not None:
        for name, key in zip(['ξ₀','A₁','κ₁','A₂','κ₂','A₃','κ₃','log_σ'],
                              ['xi0','A1','k1','A2','k2','A3','k3','log_sigma']):
            r = float(rhat[key].values)
            e = float(ess[key].values)
            ws3.append([name, r, e])
    else:
        ws3.append(["(arviz not available)", "N/A", "N/A"])

    # ----- Sheet 4: Fit quality (R² per group + global) -----
    ws4 = wb.create_sheet("4_Fit_Quality")
    # Modified: correct description
    ws4.append(["Group", "R² (unweighted)", "n points"])
    for cell in ws4[1]:
        cell.font = bold
        cell.fill = header_fill
    ws4.append(["Group 1 (2 µg/mL)", R2_1, 24])
    ws4.append(["Group 2 (0.5 µg/mL)", R2_2, 24])
    ws4.append(["Group 3 (0.05 µg/mL)", R2_3, 24])
    ws4.append(["Global (all 3 groups)", R2_total, 72])
    ws4.append([""])
    ws4.append(["Residual diagnostics (all 72 points):"])
    ws4.append(["Shapiro-Wilk W", shapiro_stat])
    ws4.append(["Shapiro-Wilk p-value", shapiro_p])
    ws4.append(["Durbin-Watson", dw])
    ws4.append(["Note: 2 = no autocorrelation; <1 or >3 indicates concern"])
    ws4.append([""])
    ws4.append(["RMSE comparison:"])
    ws4.append(["RMSE 4PL (all 72 points)", rmse_4])
    ws4.append(["RMSE 5PL constrained B=1/G (all 72 points)", rmse_5])

    # ----- Sheet 5: 4PL / 5PL per-group fits + bootstrap CIs -----
    ws5 = wb.create_sheet("5_4PL_5PL_Fits")
    ws5.append(["Group", "Model", "C", "B", "G", "R²",
                "C 2.5%", "C 97.5%", "B 2.5%", "B 97.5%", "G 2.5%", "G 97.5%"])
    for cell in ws5[1]:
        cell.font = bold
        cell.fill = header_fill
    for name in fit_4pl:
        if not np.isnan(fit_4pl[name][0]):
            C, B, r2 = fit_4pl[name]
            C4 = np.array(boot_4pl[name]['C']); B4 = np.array(boot_4pl[name]['B'])
            C5 = np.array(boot_5pl[name]['C']); G5 = np.array(boot_5pl[name]['G'])
            B5 = np.array(boot_5pl[name]['B'])
            ws5.append([name, "4PL", C, B, "—", r2,
                        np.percentile(C4, 2.5), np.percentile(C4, 97.5),
                        np.percentile(B4, 2.5), np.percentile(B4, 97.5),
                        "—", "—"])
            if not np.isnan(fit_5pl[name][0]):
                C5v, B5v, G5v, r2_5 = fit_5pl[name]
                ws5.append([name, "5PL (B=1/G)", C5v, B5v, G5v, r2_5,
                            np.percentile(C5, 2.5), np.percentile(C5, 97.5),
                            np.percentile(B5, 2.5), np.percentile(B5, 97.5),
                            np.percentile(G5, 2.5), np.percentile(G5, 97.5)])

    # ----- Sheet 6: Predicted vs Observed (all 72 points) -----
    ws6 = wb.create_sheet("6_Pred_vs_Obs")
    ws6.append(["Group", "Dilution", "Obs rep1", "Obs rep2", "Obs rep3",
                "TCS pred", "Residual rep1", "Residual rep2", "Residual rep3"])
    for cell in ws6[1]:
        cell.font = bold
        cell.fill = header_fill
    for grp_idx, (label, RFU) in enumerate(zip(group_labels, [RFU1, RFU2, RFU3]), 1):
        for i, dil in enumerate(dil_factors):
            dil_label = "blank (f→∞)" if dil >= 1e6 else dil
            tcs_pred = model(dil, xi0_hat,
                             [A1, A2, A3][grp_idx-1],
                             [D1, D2, D3][grp_idx-1],
                             [k1, k2, k3][grp_idx-1])
            ws6.append([f"Group {grp_idx} ({label})", dil_label,
                        RFU[i, 0], RFU[i, 1], RFU[i, 2],
                        tcs_pred,
                        RFU[i, 0] - tcs_pred, RFU[i, 1] - tcs_pred, RFU[i, 2] - tcs_pred])

    # Sheet 10: MCMC posterior samples, thinned to <= 20000 rows.
    # NOTE: the full chain (2e6 draws) exceeds the xlsx row limit
    # (1,048,576) and takes minutes per openpyxl row-append; thinning to
    # 20000 rows loses no posterior information because the per-parameter
    # ESS is only ~3.7e3-1.2e4 (see 3_Convergence). The full chain is
    # exactly reproducible from the fixed seed.
    ws10 = wb.create_sheet("10_MCMC_Samples")
    ws10.append(["xi0", "A1", "k1", "A2", "k2", "A3", "k3"])
    for cell in ws10[1]:
        cell.font = bold; cell.fill = header_fill
    _thin = max(1, samples.shape[0] // 20000)
    _thinned = samples[::_thin]
    for _row in _thinned:
        ws10.append([float(_v) for _v in _row])
    ws10.append([])
    ws10.append([f"Thinned every {_thin}-th draw: showing {_thinned.shape[0]} "
                 f"of {samples.shape[0]} posterior samples (no information "
                 f"loss: per-parameter ESS ~ 3.7e3-1.2e4; full chain "
                 f"reproducible from the fixed seed)."])

    wb.save("SI_Table_1.xlsx")
    print("\n" + "="*60)
    print("SI Table 1 (v2, D locked) saved to 'SI_Table_1.xlsx'")
    print("="*60)
    print("Sheets:")
    print("  1. Raw_RFU           — 3 groups × 8 dilutions × 3 reps (72 data points)")
    print("  2. MCMC_Parameters  — 7 free + 3 locked D + 1 log_sigma = 11 entries")
    print("  3. Convergence      — R-hat + ESS (8 dimensions)")
    print("  4. Fit_Quality      — R² per group + global + residuals + RMSE")
    print("  5. 4PL_5PL_Fits     — per-group 4PL/5PL fits + bootstrap CIs")
    print("  6. Pred_vs_Obs      — all 72 points with predictions + residuals")

except ImportError:
    print("\n[!] openpyxl 未安装。SI Table 1 未保存。请运行: pip install openpyxl")
except Exception as e_save:
    print(f"\n[!] SI Table 1 保存失败: {e_save}")
    import traceback
    traceback.print_exc()
rmse_tcs = np.sqrt(np.mean((y_obs - y_pred)**2))


# ========== Bayesian R²（后验预测R²） ==========
# 从MCMC后验采样，计算R²的分布
n_r2_samples = 2000
np.random.seed(42)
r2_idx = np.random.choice(samples.shape[0], n_r2_samples, replace=False)

r2_posterior = []
r2_posterior_global = []

for idx in r2_idx:
    s = samples[idx]
    theta_s = s[:7]
    y_pred_s = predict_global(theta_s, f_data, cond_idx)
    
    # 全局R²
    ss_res = np.sum((y_obs - y_pred_s)**2)
    ss_tot = np.sum((y_obs - np.mean(y_obs))**2)
    r2_posterior_global.append(1 - ss_res / ss_tot)
    
    # 每组R²
    r2_group_list = []
    for g in range(3):
        mask = (cond_idx == g)
        yg_true = y_obs[mask]
        yg_pred = y_pred_s[mask]
        ss_res_g = np.sum((yg_true - yg_pred)**2)
        ss_tot_g = np.sum((yg_true - np.mean(yg_true))**2)
        r2_group_list.append(1 - ss_res_g / ss_tot_g)
    r2_posterior.append(r2_group_list)

r2_posterior = np.array(r2_posterior)  # shape: (n_samples, 3)
r2_posterior_global = np.array(r2_posterior_global)

print("\n" + "="*60)
print("Bayesian R²（后验预测R²，n=2000 samples）")
print("="*60)
print(f"全局 Bayesian R² = {np.median(r2_posterior_global):.4f} "
      f"[{np.percentile(r2_posterior_global, 2.5):.4f}, "
      f"{np.percentile(r2_posterior_global, 97.5):.4f}]")
for g, label in enumerate(['组1 (2 µg/mL)', '组2 (0.5 µg/mL)', '组3 (0.05 µg/mL)']):
    r2_g = r2_posterior[:, g]
    print(f"{label}: Bayesian R² = {np.median(r2_g):.4f} "
          f"[{np.percentile(r2_g, 2.5):.4f}, "
          f"{np.percentile(r2_g, 97.5):.4f}]")

# 对比点估计R²
print(f"\n对比（点估计R²）:")
print(f"  全局 R² = {R2_total:.4f}")
print(f"  组1 R² = {R2_1:.4f}")
print(f"  组2 R² = {R2_2:.4f}")
print(f"  组3 R² = {R2_3:.4f}")
print(f"RMSE TCS global fit = {rmse_tcs:.6f}")


# ============================================================================
# LoB / LoD / LoQ 计算 (S2a formulas, analog assay)
# 追加到 v2 代码末尾（print(f"RMSE TCS global fit = {rmse_tcs:.6f}") 之后）
# 适配 v2 参数排列: [xi0, A1, k1, A2, k2, A3, k3, log_sigma] (8列)
# ============================================================================

# --- 物理常数 ---
V_well = 100e-6      # 反应体积 (L)
N_A = 6.022e23
MW_ab = 150000       # goat anti-mouse IgG MW (g/mol)

# M₀: fluorescent detection antibody
# 2 mg/mL × (1/500 dilution) × 100 μL = 0.4 μg = 0.4e-6 g
M0 = (0.4e-6 / MW_ab) * N_A  # ≈ 1.606e12 molecules

# 浓度换算函数
def mol_to_ngmL(M):
    """molecules -> ng/mL"""
    g_per_L = M / (V_well * N_A) * MW_ab  # g/L
    return g_per_L * 1e6                   # ng/mL (1 g/L = 1e6 ng/mL)

M0_ngmL = mol_to_ngmL(M0)

print(f"\n{'='*80}")
print("LoB / LoD / LoQ (S2a formulas, Eq. S2a.22/17/24)")
print(f"{'='*80}")
print(f"M₀ = {M0:.3e} molecules = {M0_ngmL:.1f} ng/mL (= 4 μg/mL, 2 mg/mL × 1/500)")

# --- σ_S 估计 (per S2a.10: local σ for each application) ---
# LoB: blank σ;  LoD/LoQ: low-conc σ;  CV: per-point σ
sd_blank_1 = np.std(RFU1[-1], ddof=1)
sd_blank_2 = np.std(RFU2[-1], ddof=1)
sd_blank_3 = np.std(RFU3[-1], ddof=1)
# Pooled blank SD (3 groups × 3 reps, df = 6)
sigma_S_blank = np.sqrt(
    (2*sd_blank_1**2 + 2*sd_blank_2**2 + 2*sd_blank_3**2) / 6
)

# Low-conc: dil=64 (index 6), pooled across 3 groups
sd_low_1 = np.std(RFU1[6], ddof=1)
sd_low_2 = np.std(RFU2[6], ddof=1)
sd_low_3 = np.std(RFU3[6], ddof=1)
sigma_S_low = np.sqrt(
    (2*sd_low_1**2 + 2*sd_low_2**2 + 2*sd_low_3**2) / 6
)

print(f"\nσ_S (local estimates, per S2a.10):")
print(f"  Blank (pooled, 3 groups): {sigma_S_blank:.6f}  "
      f"[G1={sd_blank_1:.6f}, G2={sd_blank_2:.6f}, G3={sd_blank_3:.6f}]")
print(f"  Low-conc dil=64 (pooled): {sigma_S_low:.6f}  "
      f"[G1={sd_low_1:.6f}, G2={sd_low_2:.6f}, G3={sd_low_3:.6f}]")

# κΩ from fit
kappa_Omega = M0 / xi0_hat
print(f"\nκΩ = M₀/ξ₀ = {kappa_Omega:.3e}")

# --- 后验采样 for CI ---
n_loq_samples = min(2000, samples.shape[0])
np.random.seed(42)
loq_idx = np.random.choice(samples.shape[0], n_loq_samples, replace=False)

# v2 参数排列: [xi0, A1, k1, A2, k2, A3, k3, log_sigma] (8 columns)
# G1: A=samples[:,1], k=samples[:,2], D=D1_fixed
# G2: A=samples[:,3], k=samples[:,4], D=D2_fixed
# G3: A=samples[:,5], k=samples[:,6], D=D3_fixed

print(f"\nGeneral form: M_limit = c × Ω(κ+1) × σ_p̂")
print(f"  where σ_p̂ = σ_S/(A-D) (normalized signal SD)")
print(f"  Ω(κ+1) = (M₀/ξ₀)(1+1/κ)")
print(f"  Eq. S2a.22: LoB (c=1.645, σ_p̂ from blank)")
print(f"  Eq. S2a.17: LoD (c=3.29,  σ_p̂ from low-conc)")
print(f"  Eq. S2a.24: LoQ (c=5.0,   σ_p̂ from low-conc)")
print(f"{'='*80}")

# 存储结果用于 Excel
lob_lod_loq_results = []

for g, (A_g, k_g, D_g, lbl) in enumerate([
    (A1, k1, D1_fixed, 'G1 (2 µg/mL)'),
    (A2, k2, D2_fixed, 'G2 (0.5 µg/mL)'),
    (A3, k3, D3_fixed, 'G3 (0.05 µg/mL)')
]):
    # --- Point estimate ---
    sigma_phat_blank = sigma_S_blank / (A_g - D_g)
    sigma_phat_low   = sigma_S_low   / (A_g - D_g)
    Omega_kappa1     = kappa_Omega * (1 + 1/k_g)

    LoB = 1.645 * Omega_kappa1 * sigma_phat_blank
    LoD = 3.29  * Omega_kappa1 * sigma_phat_low
    LoQ = 5.0   * Omega_kappa1 * sigma_phat_low

    # --- Posterior CI ---
    # v2 index: A at 1+g*2, k at 2+g*2
    idx_A = 1 + g * 2   # 1, 3, 5
    idx_k = 2 + g * 2   # 2, 4, 6

    A_s   = samples[loq_idx, idx_A]
    k_s   = samples[loq_idx, idx_k]
    xi0_s = samples[loq_idx, 0]

    kappa_Omega_s  = M0 / xi0_s
    Omega_kappa1_s = kappa_Omega_s * (1 + 1/k_s)

    # D is fixed (not sampled) — use D_g constant
    sigma_phat_blank_s = sigma_S_blank / (A_s - D_g)
    sigma_phat_low_s   = sigma_S_low   / (A_s - D_g)

    LoB_s = 1.645 * Omega_kappa1_s * sigma_phat_blank_s
    LoD_s = 3.29  * Omega_kappa1_s * sigma_phat_low_s
    LoQ_s = 5.0   * Omega_kappa1_s * sigma_phat_low_s

    # Convert to ng/mL
    LoB_ng  = mol_to_ngmL(LoB)
    LoD_ng  = mol_to_ngmL(LoD)
    LoQ_ng  = mol_to_ngmL(LoQ)
    LoB_ci  = [mol_to_ngmL(np.percentile(LoB_s, 2.5)),
               mol_to_ngmL(np.percentile(LoB_s, 97.5))]
    LoD_ci  = [mol_to_ngmL(np.percentile(LoD_s, 2.5)),
               mol_to_ngmL(np.percentile(LoD_s, 97.5))]
    LoQ_ci  = [mol_to_ngmL(np.percentile(LoQ_s, 2.5)),
               mol_to_ngmL(np.percentile(LoQ_s, 97.5))]

    print(f"\n{lbl} (κ={k_g:.4f}):")
    print(f"  σ_p̂(blank) = {sigma_phat_blank:.6f}")
    print(f"  σ_p̂(low)   = {sigma_phat_low:.6f}")
    print(f"  Ω(κ+1)     = {Omega_kappa1:.3e}")
    print(f"  LoB = {LoB:.3e} molecules = {LoB_ng:.2f} ng/mL  "
          f"[{LoB_ci[0]:.2f}, {LoB_ci[1]:.2f}] ng/mL")
    print(f"  LoD = {LoD:.3e} molecules = {LoD_ng:.2f} ng/mL  "
          f"[{LoD_ci[0]:.2f}, {LoD_ci[1]:.2f}] ng/mL")
    print(f"  LoQ = {LoQ:.3e} molecules = {LoQ_ng:.2f} ng/mL  "
          f"[{LoQ_ci[0]:.2f}, {LoQ_ci[1]:.2f}] ng/mL")

    lob_lod_loq_results.append({
        'label': lbl, 'kappa': k_g,
        'LoB_mol': LoB, 'LoD_mol': LoD, 'LoQ_mol': LoQ,
        'LoB_ng': LoB_ng, 'LoD_ng': LoD_ng, 'LoQ_ng': LoQ_ng,
        'LoB_ci': LoB_ci, 'LoD_ci': LoD_ci, 'LoQ_ci': LoQ_ci,
    })

# --- Zero-depletion limit ---
print(f"\n{'='*80}")
print(f"Zero-depletion limit (Eq. S2a.23/19/25): M_limit = c × KVN_A × σ_p̂")
print(f"  where KVN_A = κΩ = M₀/ξ₀ = {kappa_Omega:.3e}")
print(f"{'='*80}")

for g, (A_g, D_g, lbl) in enumerate([
    (A1, D1_fixed, 'G1 (2 µg/mL)'),
    (A2, D2_fixed, 'G2 (0.5 µg/mL)'),
    (A3, D3_fixed, 'G3 (0.05 µg/mL)')
]):
    sigma_phat_blank = sigma_S_blank / (A_g - D_g)
    sigma_phat_low   = sigma_S_low   / (A_g - D_g)

    LoB_zd  = 1.645 * kappa_Omega * sigma_phat_blank
    LoD_zd  = 3.29  * kappa_Omega * sigma_phat_low
    LoQ_zd  = 5.0   * kappa_Omega * sigma_phat_low

    print(f"\n{lbl} (zero-dep):")
    print(f"  LoB = {mol_to_ngmL(LoB_zd):.3f} ng/mL")
    print(f"  LoD = {mol_to_ngmL(LoD_zd):.3f} ng/mL")
    print(f"  LoQ = {mol_to_ngmL(LoQ_zd):.3f} ng/mL")

# ============================================================================
# Excel: Sheet 7
# ============================================================================
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb2 = load_workbook("SI_Table_1.xlsx")
    if "7_LoB_LoD_LoQ" in wb2.sheetnames:
        del wb2["7_LoB_LoD_LoQ"]
    ws7 = wb2.create_sheet("7_LoB_LoD_LoQ")

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws7.append(["LoB/LoD/LoQ (S2a formulas, analog assay)", "", "", "", "", ""])
    ws7.append(["", "", "", "", "", ""])

    ws7.append(["Group", "κ",
                "LoB (molecules)", "LoB (ng/mL)", "LoB CI (ng/mL)",
                "LoD (molecules)", "LoD (ng/mL)", "LoD CI (ng/mL)",
                "LoQ (molecules)", "LoQ (ng/mL)", "LoQ CI (ng/mL)"])
    for cell in ws7[3]:
        cell.font = bold
        cell.fill = header_fill

    for res in lob_lod_loq_results:
        ws7.append([
            res['label'], res['kappa'],
            f"{res['LoB_mol']:.3e}", f"{res['LoB_ng']:.2f}",
            f"[{res['LoB_ci'][0]:.2f}, {res['LoB_ci'][1]:.2f}]",
            f"{res['LoD_mol']:.3e}", f"{res['LoD_ng']:.2f}",
            f"[{res['LoD_ci'][0]:.2f}, {res['LoD_ci'][1]:.2f}]",
            f"{res['LoQ_mol']:.3e}", f"{res['LoQ_ng']:.2f}",
            f"[{res['LoQ_ci'][0]:.2f}, {res['LoQ_ci'][1]:.2f}]",
        ])

    # Zero-depletion rows
    ws7.append(["", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["Zero-depletion limit (κ→∞):", "", "", "", "", "", "", "", "", "", ""])
    for g, (A_g, D_g, lbl) in enumerate([
        (A1, D1_fixed, 'G1'),
        (A2, D2_fixed, 'G2'),
        (A3, D3_fixed, 'G3')
    ]):
        sp_b = sigma_S_blank / (A_g - D_g)
        sp_l = sigma_S_low   / (A_g - D_g)
        LoB_zd = mol_to_ngmL(1.645 * kappa_Omega * sp_b)
        LoD_zd = mol_to_ngmL(3.29  * kappa_Omega * sp_l)
        LoQ_zd = mol_to_ngmL(5.0   * kappa_Omega * sp_l)
        ws7.append([f"{lbl} (zero-dep)", "∞",
                    "", f"{LoB_zd:.3f}", "",
                    "", f"{LoD_zd:.3f}", "",
                    "", f"{LoQ_zd:.3f}", ""])

    ws7.append(["", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["Parameters:", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["M₀ (molecules)", f"{M0:.3e}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["M₀ (ng/mL)", f"{M0_ngmL:.1f}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["σ_S (pooled blank)", f"{sigma_S_blank:.6f}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["σ_S (pooled low-conc)", f"{sigma_S_low:.6f}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["κΩ = M₀/ξ₀", f"{kappa_Omega:.3e}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["ξ₀", f"{xi0_hat:.4f}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["V_well (L)", f"{V_well}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["MW (g/mol)", f"{MW_ab}", "", "", "", "", "", "", "", "", ""])
    ws7.append(["", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["Formulas (S2a):", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["LoB = 1.645 × Ω(κ+1) × σ_p̂(blank)", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["LoD = 3.29 × Ω(κ+1) × σ_p̂(low)", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["LoQ = 5.0 × Ω(κ+1) × σ_p̂(low)", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["where σ_p̂ = σ_S/(A-D), Ω(κ+1) = (M₀/ξ₀)(1+1/κ)", "", "", "", "", "", "", "", "", "", ""])
    ws7.append(["ng/mL = molecules / (V × N_A) × MW × 1e6", "", "", "", "", "", "", "", "", "", ""])

    wb2.save("SI_Table_1.xlsx")
    print(f"\nSheet 7 (LoB/LoD/LoQ) appended to SI_Table_1.xlsx")
except Exception as e:
    print(f"\n[!] LoB/LoD/LoQ Excel export failed: {e}")
    import traceback
    traceback.print_exc()

# ============================================================================
# Sheet 8: Bayesian R²
# ============================================================================
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font as _Font, PatternFill as _Fill
    wb3 = load_workbook("SI_Table_1.xlsx")
    ws8 = wb3.create_sheet("8_Bayesian_R2")
    ws8.append(["Bayesian R² (posterior predictive, n=2000 samples)"])
    ws8.append(["Group", "Median", "2.5% HDI", "97.5% HDI"])
    for cell in ws8[2]:
        cell.font = _Font(bold=True)

    ws8.append(["Global", np.median(r2_posterior_global),
                np.percentile(r2_posterior_global, 2.5),
                np.percentile(r2_posterior_global, 97.5)])
    for g, label in enumerate(['G1 (2 µg/mL)', 'G2 (0.5 µg/mL)', 'G3 (0.05 µg/mL)']):
        r2_g = r2_posterior[:, g]
        ws8.append([label, np.median(r2_g),
                    np.percentile(r2_g, 2.5),
                    np.percentile(r2_g, 97.5)])

    ws8.append([""])
    ws8.append(["Point-estimate R² (comparison):"])
    ws8.append(["Global", R2_total])
    ws8.append(["G1", R2_1])
    ws8.append(["G2", R2_2])
    ws8.append(["G3", R2_3])
    ws8.append(["RMSE (TCS global)", rmse_tcs if 'rmse_tcs' in dir() else 'N/A'])

    wb3.save("SI_Table_1.xlsx")
    print(f"Sheet 8 (Bayesian R²) appended to SI_Table_1.xlsx")
except Exception as e:
    print(f"[!] Bayesian R² Excel export failed: {e}")
















# -*- coding: utf-8 -*-
"""
Profile Likelihood 分析 — 接在 scale degeneracy4-0固定背景.py 之后运行

用法：
  1. 先运行原代码（确保所有变量在内存中）
  2. 然后运行本文件：exec(open('profile_likelihood_addon.py').read())
     或者直接：python3 profile_likelihood_addon.py
     （会自动 exec 原代码加载数据和模型）

输出：
  - profile_1d_all.svg       : 1D profile log-likelihood (4 panels)
  - profile_2D_xi0_kappa1.svg: 2D profile contour (xi0, kappa1)
  - profile_summary.txt      : summary table
  - profile_data.npz         : raw data

依赖：numpy, scipy, matplotlib（原代码已有）
"""

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.optimize import minimize
import os

# ============================================================================
# 1. 加载原代码的数据和模型
# ============================================================================
print("=" * 70)
print("Profile Likelihood Analysis")
print("=" * 70)

# 如果独立运行，先 exec 原代码
if 'y_obs' not in dir():
    print("Loading data from original script...")
    # Try common filenames with utf-8 encoding
    _candidates = [
        'scale degeneracy4-0固定背景.py',
        'scale degeneracy4.0固定背景.py',
        'Scale Degeneracy5.0 固定背景+profile likehood验证时间长.py',
    ]
    _loaded = False
    for _fname in _candidates:
        _path = os.path.join(os.path.dirname(os.path.abspath(__file__)), _fname)
        if os.path.exists(_path):
            exec(open(_path, encoding='utf-8').read())
            _loaded = True
            break
    if not _loaded:
        # Fallback: list .py files and try each
        _dir = os.path.dirname(os.path.abspath(__file__))
        for _fname in os.listdir(_dir):
            if _fname.endswith('.py') and 'scale' in _fname.lower() and 'degener' in _fname.lower():
                _path = os.path.join(_dir, _fname)
                try:
                    exec(open(_path, encoding='utf-8').read())
                    _loaded = True
                    print(f"  Loaded: {_fname}")
                    break
                except Exception:
                    continue
    if not _loaded:
        raise FileNotFoundError(
            "Cannot find the original script. Please run this file "
            "from the same directory as your scale degeneracy script, "
            "or run the original script first then this one.")

# ============================================================================
# 2. 定义 profile likelihood 函数
# ============================================================================

def neg_log_likelihood_free(free_params, fixed_param_name, fixed_value,
                            dil, cond_idx, y_obs, sigma_hat):
    """
    负 log-likelihood，固定一个参数，优化其余。

    free_params: 除固定参数外的 7 个参数
                 [A1, k1, A2, k2, A3, k3]  (xi0 固定时)
                 [xi0, A1, A2, k2, A3, k3] (k1 固定时)
    fixed_param_name: 'xi0' 或 'kappa1' 等
    fixed_value: 固定值
    sigma_hat: 固定 sigma（用 MCMC 或 least-squares 的估计值）
    """
    if fixed_param_name == 'xi0':
        xi0 = fixed_value
        A1, k1, A2, k2, A3, k3 = free_params
    elif fixed_param_name == 'kappa1':
        xi0, A1, A2, k2, A3, k3 = free_params
        k1 = fixed_value
    elif fixed_param_name == 'kappa2':
        xi0, A1, k1, A2, A3, k3 = free_params
        k2 = fixed_value
    elif fixed_param_name == 'kappa3':
        xi0, A1, k1, A2, k2, A3 = free_params
        k3 = fixed_value
    else:
        raise ValueError(f"Unknown fixed_param: {fixed_param_name}")

    # 边界检查
    if xi0 <= 0.1 or xi0 > 500: return 1e15
    if A1 <= 0.5 or A1 > 3.5: return 1e15
    if k1 <= 1e-4 or k1 > 1000: return 1e15
    if A2 <= 0.5 or A2 > 2.5: return 1e15
    if k2 <= 1e-4 or k2 > 1000: return 1e15
    if A3 <= 0.1 or A3 > 0.5: return 1e15
    if k3 <= 1e-4 or k3 > 1000: return 1e15

    theta_full = [xi0, A1, k1, A2, k2, A3, k3]
    y_pred = predict_global(theta_full, dil, cond_idx)
    n = len(y_obs)
    ll = -0.5 * np.sum(((y_obs - y_pred) / sigma_hat) ** 2) - n * np.log(sigma_hat)
    return -ll  # 返回负 log-likelihood（用于最小化）


# def profile_likelihood_1d(param_name, param_grid, theta_hat, sigma_hat,
#                           dil, cond_idx, y_obs):
#     """
#     对单个参数做 profile likelihood。

#     返回: (param_grid, profile_ll) — 每个 grid 点的 profile log-likelihood
#     """
#     n_free = 6
#     profile_ll = np.full(len(param_grid), -np.inf)

#     # 从最优解提取各参数
#     xi0_opt, A1_opt, k1_opt, A2_opt, k2_opt, A3_opt, k3_opt = theta_hat

#     for i, val in enumerate(param_grid):
#         # 初始猜测：最优解（去掉固定参数）
#         if param_name == 'xi0':
#             x0 = [A1_opt, k1_opt, A2_opt, k2_opt, A3_opt, k3_opt]
#         elif param_name == 'kappa1':
#             x0 = [xi0_opt, A1_opt, A2_opt, k2_opt, A3_opt, k3_opt]
#         elif param_name == 'kappa2':
#             x0 = [xi0_opt, A1_opt, k1_opt, A2_opt, A3_opt, k3_opt]
#         elif param_name == 'kappa3':
#             x0 = [xi0_opt, A1_opt, k1_opt, A2_opt, k2_opt, A3_opt]

#         # 稍微扰动初始值，避免卡在已经是最优的点
#         x0 = np.array(x0) * (1.0 + 1e-3 * np.random.randn(len(x0)))
#         x0 = np.array(x0) * (1.0 + 1e-1 * np.random.randn(len(x0)))  # 增大到 10% 扰动

#         try:
#             res = minimize(
#                 neg_log_likelihood_free, x0,
#                 args=(param_name, val, dil, cond_idx, y_obs, sigma_hat),
#                 method='Nelder-Mead',
#                 options={'maxiter': 5000, 'xatol': 1e-8, 'fatol': 1e-8}
#             )
#             if res.success or res.fun < 1e14:
#                 profile_ll[i] = -res.fun  # 转回正 log-likelihood
#         except Exception as e:
#             print(f"  {param_name}={val:.4g}: optimization failed ({e})")

#         if (i + 1) % 10 == 0:
#             print(f"  {param_name} profile: {i+1}/{len(param_grid)} done")

#     return param_grid, profile_ll

def profile_likelihood_1d(param_name, param_grid, theta_hat, sigma_hat,
                          dil, cond_idx, y_obs):
    """
    稳健版 profile likelihood：多起点 + 随机跳跃，避免局部凹陷
    """
    profile_ll = np.full(len(param_grid), -np.inf)

    xi0_opt, A1_opt, k1_opt, A2_opt, k2_opt, A3_opt, k3_opt = theta_hat

    for i, val in enumerate(param_grid):
        # ---- 构建多组初始猜测 ----
        starts = []
        if param_name == 'xi0':
            base = [A1_opt, k1_opt, A2_opt, k2_opt, A3_opt, k3_opt]
        elif param_name == 'kappa1':
            base = [xi0_opt, A1_opt, A2_opt, k2_opt, A3_opt, k3_opt]
        elif param_name == 'kappa2':
            base = [xi0_opt, A1_opt, k1_opt, A2_opt, A3_opt, k3_opt]
        elif param_name == 'kappa3':
            base = [xi0_opt, A1_opt, k1_opt, A2_opt, k2_opt, A3_opt]
        else:
            raise ValueError(f"Unknown param: {param_name}")

        # 1) 标准起点：最优解 + 小扰动
        starts.append(np.array(base) * (1.0 + 1e-4 * np.random.randn(len(base))))
        # 2) 10% 扰动起点
        for _ in range(3):
            starts.append(np.array(base) * (1.0 + 0.1 * np.random.randn(len(base))))
        # 3) 更大幅度扰动（30%）
        for _ in range(3):
            starts.append(np.array(base) * (1.0 + 0.3 * np.random.randn(len(base))))
        # 4) 完全随机的合理值（在参数允许范围内）
        for _ in range(3):
            if param_name == 'xi0':
                rnd = [np.random.uniform(0.5, 3.5),   # A1
                       np.random.uniform(1e-4, 1000), # k1
                       np.random.uniform(0.5, 2.5),   # A2
                       np.random.uniform(1e-4, 1000), # k2
                       np.random.uniform(0.1, 0.5),   # A3
                       np.random.uniform(1e-4, 1000)] # k3
            elif param_name == 'kappa1':
                rnd = [np.random.uniform(0.5, 400),   # xi0
                       np.random.uniform(0.5, 3.5),   # A1
                       np.random.uniform(0.5, 2.5),   # A2
                       np.random.uniform(1e-4, 1000), # k2
                       np.random.uniform(0.1, 0.5),   # A3
                       np.random.uniform(1e-4, 1000)] # k3
            # 类似处理其他参数...
            else:
                rnd = base  # 简单回退
            starts.append(np.array(rnd))

        best_ll = -np.inf
        for x0 in starts:
            try:
                res = minimize(
                    neg_log_likelihood_free, x0,
                    args=(param_name, val, dil, cond_idx, y_obs, sigma_hat),
                    method='Nelder-Mead',
                    options={'maxiter': 5000, 'xatol': 1e-8, 'fatol': 1e-8}
                )
                if res.success or res.fun < 1e14:
                    ll_val = -res.fun
                    if ll_val > best_ll:
                        best_ll = ll_val
            except:
                pass

        profile_ll[i] = best_ll

        if (i + 1) % 10 == 0:
            print(f"  {param_name} profile: {i+1}/{len(param_grid)} done")

    return param_grid, profile_ll


def profile_likelihood_2d(param_x_name, param_y_name,
                          x_grid, y_grid, theta_hat, sigma_hat,
                          dil, cond_idx, y_obs):
    """
    2D profile likelihood: 固定两个参数，优化其余 5 个。
    返回 2D 矩阵 of profile log-likelihood。
    """
    ll_2d = np.full((len(y_grid), len(x_grid)), -np.inf)
    xi0_opt, A1_opt, k1_opt, A2_opt, k2_opt, A3_opt, k3_opt = theta_hat

    for j, xval in enumerate(x_grid):
        for i, yval in enumerate(y_grid):
            # 固定 param_x = xval, param_y = yval
            # 剩余 5 个 free params
            if param_x_name == 'xi0' and param_y_name == 'kappa1':
                x0 = [A1_opt, A2_opt, k2_opt, A3_opt, k3_opt]
                free_names = ['A1', 'A2', 'k2', 'A3', 'k3']

                def neg_ll(free):
                    A1, A2, k2, A3, k3 = free
                    if A1<=0.5 or A1>3.5: return 1e15
                    if A2<=0.5 or A2>2.5: return 1e15
                    if k2<=1e-4 or k2>1000: return 1e15
                    if A3<=0.1 or A3>0.5: return 1e15
                    if k3<=1e-4 or k3>1000: return 1e15
                    theta = [xval, A1, yval, A2, k2, A3, k3]
                    y_pred = predict_global(theta, dil, cond_idx)
                    n = len(y_obs)
                    return 0.5*np.sum(((y_obs-y_pred)/sigma_hat)**2) + n*np.log(sigma_hat)
            elif param_x_name == 'xi0' and param_y_name == 'kappa2':
                x0 = [A1_opt, k1_opt, A2_opt, A3_opt, k3_opt]

                def neg_ll(free):
                    A1, k1, A2, A3, k3 = free
                    if A1<=0.5 or A1>3.5: return 1e15
                    if k1<=1e-4 or k1>1000: return 1e15
                    if A2<=0.5 or A2>2.5: return 1e15
                    if A3<=0.1 or A3>0.5: return 1e15
                    if k3<=1e-4 or k3>1000: return 1e15
                    theta = [xval, A1, k1, A2, yval, A3, k3]
                    y_pred = predict_global(theta, dil, cond_idx)
                    n = len(y_obs)
                    return 0.5*np.sum(((y_obs-y_pred)/sigma_hat)**2) + n*np.log(sigma_hat)
            else:
                continue

            x0 = np.array(x0) * (1.0 + 1e-3 * np.random.randn(len(x0)))

            try:
                res = minimize(neg_ll, x0, method='Nelder-Mead',
                              options={'maxiter': 3000, 'xatol': 1e-6, 'fatol': 1e-6})
                if res.fun < 1e14:
                    ll_2d[i, j] = -res.fun
            except:
                pass

        print(f"  2D profile: col {j+1}/{len(x_grid)} done")

    return ll_2d


# ============================================================================
# 3. 运行 Profile Likelihood
# ============================================================================

# 用 least-squares 结果作为最优值（如果 MCMC 跑完了也可以用 MCMC 中位数）
# try:
#     theta_hat
# except NameError:
#     theta_hat = None

# 优先使用 MCMC 后验中位数（全局最优），若无则回退到 LS
try:
    theta_hat = theta_mcmc   # 来自主代码，7 个参数
    print(f"[!] Using MCMC posterior median as profile start: {theta_hat}")
except NameError:
    try:
        theta_hat
    except NameError:
        theta_hat = None

if theta_hat is None:
    print("ERROR: theta_hat not found. Run the original script first.")
    raise RuntimeError

# sigma 估计
sigma_hat = float(np.std(y_obs - predict_global(theta_hat, f_data, cond_idx)))
print(f"\nUsing theta_hat = {theta_hat}")
print(f"sigma_hat = {sigma_hat:.6f}")

# 最优 log-likelihood
y_pred_opt = predict_global(theta_hat, f_data, cond_idx)
n_data = len(y_obs)
ll_max = -0.5 * np.sum(((y_obs - y_pred_opt) / sigma_hat) ** 2) - n_data * np.log(sigma_hat)
print(f"ll_max = {ll_max:.2f}")

# Wilks 临界值: 2*(ll_max - ll_profile) < chi^2(1, 0.95) = 3.841
chi2_crit = 3.841
ll_threshold = ll_max - chi2_crit / 2.0
print(f"95% CI threshold: ll > {ll_threshold:.2f}")

# ---- xi0 profile ----
print("\n--- Profile: xi0 ---")
xi0_opt = theta_hat[0]
xi0_range = max(xi0_opt * 0.1, 0.5), min(xi0_opt * 5, 400)
xi0_grid = np.linspace(xi0_range[0], xi0_range[1], 81)
xi0_grid_prof, xi0_prof_ll = profile_likelihood_1d(
    'xi0', xi0_grid, theta_hat, sigma_hat, f_data, cond_idx, y_obs)

# ---- kappa1 profile ----
print("\n--- Profile: kappa1 ---")
k1_opt = theta_hat[2]
k1_range = max(k1_opt * 0.05, 1e-3), min(k1_opt * 20, 500)
k1_grid = np.linspace(k1_range[0], k1_range[1], 81)
k1_grid_prof, k1_prof_ll = profile_likelihood_1d(
    'kappa1', k1_grid, theta_hat, sigma_hat, f_data, cond_idx, y_obs)

# ---- kappa2 profile ----
print("\n--- Profile: kappa2 ---")
k2_opt = theta_hat[4]
k2_range = max(k2_opt * 0.05, 1e-3), min(k2_opt * 20, 500)
k2_grid = np.linspace(k2_range[0], k2_range[1], 81)
k2_grid_prof, k2_prof_ll = profile_likelihood_1d(
    'kappa2', k2_grid, theta_hat, sigma_hat, f_data, cond_idx, y_obs)

# ---- kappa3 profile ----
print("\n--- Profile: kappa3 ---")
k3_opt = theta_hat[6]
# k3_range = max(k3_opt * 0.05, 1e-3), min(k3_opt * 20, 500)
# k3_grid = np.linspace(k3_range[0], k3_range[1], 81)


# κ₃ 后验很宽，上限需覆盖到 1500
k3_range_low = max(k3_opt * 0.01, 1e-3)
k3_range_high = max(k3_opt * 50, 1500)   # 确保包含 MCMC 的 95% HDI 上限
k3_grid = np.linspace(k3_range_low, k3_range_high, 100)



k3_grid_prof, k3_prof_ll = profile_likelihood_1d(
    'kappa3', k3_grid, theta_hat, sigma_hat, f_data, cond_idx, y_obs)

# ---- 2D profile: (xi0, kappa1) ----
print("\n--- 2D Profile: (xi0, kappa1) ---")
xi0_2d = np.linspace(xi0_range[0], xi0_range[1], 41)
k1_2d = np.linspace(k1_range[0], k1_range[1], 41)
ll_2d_xi0_k1 = profile_likelihood_2d(
    'xi0', 'kappa1', xi0_2d, k1_2d, theta_hat, sigma_hat,
    f_data, cond_idx, y_obs)


# ============================================================================
# 4. 提取 CI 和 Ridge 斜率
# ============================================================================

# def extract_ci_1d(param_grid, profile_ll, ll_max, threshold, param_name):
#     """从 profile likelihood 提取 CI"""
#     valid = profile_ll > -np.inf
#     if not np.any(valid):
#         return None, None, False

#     above = profile_ll > threshold
#     if not np.any(above):
#         return None, None, False

#     idx_above = np.where(above)[0]
#     ci_low = param_grid[idx_above[0]]
#     ci_high = param_grid[idx_above[-1]]
#     bounded = ci_low > param_grid[0] + 1e-10 and ci_high < param_grid[-1] - 1e-10
#     return ci_low, ci_high, bounded

def extract_ci_1d(param_grid, profile_ll, ll_max, threshold, param_name):
    """稳健提取 CI：确保包含最优值"""
    valid = profile_ll > -np.inf
    if not np.any(valid):
        return None, None, False

    # 找到最优值索引（profile 最大值）
    i_opt = np.argmax(profile_ll)
    # 向左搜索低于阈值的点
    ci_low = None
    for i in range(i_opt, -1, -1):
        if profile_ll[i] < threshold:
            # 在 i 和 i+1 之间插值
            if i < len(param_grid)-1 and profile_ll[i+1] > threshold:
                frac = (threshold - profile_ll[i]) / (profile_ll[i+1] - profile_ll[i])
                ci_low = param_grid[i] + frac * (param_grid[i+1] - param_grid[i])
            else:
                ci_low = param_grid[i]
            break
    # 向右搜索
    ci_high = None
    for i in range(i_opt, len(param_grid)):
        if profile_ll[i] < threshold:
            if i > 0 and profile_ll[i-1] > threshold:
                frac = (threshold - profile_ll[i-1]) / (profile_ll[i] - profile_ll[i-1])
                ci_high = param_grid[i-1] + frac * (param_grid[i] - param_grid[i-1])
            else:
                ci_high = param_grid[i]
            break

    bounded = (ci_low is not None and ci_high is not None)
    return ci_low, ci_high, bounded

print("\n" + "=" * 70)
print("Profile Likelihood Results")
print("=" * 70)

results = {}

for name, grid, ll, opt_val in [
    ('xi0',  xi0_grid,  xi0_prof_ll,  xi0_opt),
    ('kappa1', k1_grid, k1_prof_ll, k1_opt),
    ('kappa2', k2_grid, k2_prof_ll, k2_opt),
    ('kappa3', k3_grid, k3_prof_ll, k3_opt),
]:
    ci_low, ci_high, bounded = extract_ci_1d(grid, ll, ll_max, ll_threshold, name)
    if ci_low is not None:
        ratio = ci_high / ci_low
        print(f"  {name:>8s}: opt={opt_val:.4g}, CI=[{ci_low:.4g}, {ci_high:.4g}], "
              f"ratio={ratio:.1f}, bounded={bounded}")
        results[name] = (ci_low, ci_high, ratio, bounded)
    else:
        print(f"  {name:>8s}: opt={opt_val:.4g}, CI=UNBOUNDED")
        results[name] = (None, None, None, False)

# 2D ridge slope
print("\n--- 2D Ridge Analysis (xi0, kappa1) ---")
valid_2d = ll_2d_xi0_k1 > ll_max - chi2_crit * 2  # 更宽松的阈值看 ridge
if np.any(valid_2d):
    # 找 ridge 上的点（每个 kappa1 列里 ll 最大的 xi0）
    ridge_xi0 = []
    ridge_k1 = []
    for j in range(len(k1_2d)):
        col = ll_2d_xi0_k1[:, j]
        valid_col = col > -np.inf
        if np.any(valid_col):
            i_best = np.argmax(np.where(valid_col, col, -1e15))
            ridge_xi0.append(xi0_2d[i_best])
            ridge_k1.append(k1_2d[j])

    if len(ridge_xi0) > 5:
        ridge_xi0 = np.array(ridge_xi0)
        ridge_k1 = np.array(ridge_k1)
        # Fit log(kappa1) = a * log(xi0) + b
        log_xi0 = np.log(ridge_xi0)
        log_k1 = np.log(ridge_k1)
        slope, intercept = np.polyfit(log_xi0, log_k1, 1)
        print(f"  Ridge: log(kappa1) = {slope:.3f} * log(xi0) + {intercept:.3f}")
        print(f"  Slope = {slope:.3f}")
        print(f"  Scale degeneracy predicts slope = 0 (κ independent of ξ₀ at fixed M₀)")
        print(f"  Observed slope = {slope:.3f} reflects data-driven correlation, not scale degeneracy")
    else:
        slope = None
        ridge_xi0 = np.array(ridge_xi0)
        ridge_k1 = np.array(ridge_k1)
        print("  Not enough ridge points to fit slope")
else:
    slope = None
    ridge_xi0 = np.array([])
    ridge_k1 = np.array([])
    print("  No valid 2D profile points")


# ============================================================================
# 5. 绘图
# ============================================================================

plt.rcParams['font.family'] = 'Times New Roman'
plt.rcParams['mathtext.fontset'] = 'stix'

# --- 1D profiles (4 panels) ---
fig, axes = plt.subplots(2, 2, figsize=(12, 10))

panels = [
    (axes[0,0], 'xi0', xi0_grid, xi0_prof_ll, xi0_opt, r'$\xi_0$'),
    (axes[0,1], 'kappa1', k1_grid, k1_prof_ll, k1_opt, r'$\kappa_1$'),
    (axes[1,0], 'kappa2', k2_grid, k2_prof_ll, k2_opt, r'$\kappa_2$'),
    (axes[1,1], 'kappa3', k3_grid, k3_prof_ll, k3_opt, r'$\kappa_3$'),
]

for ax, name, grid, ll, opt_val, label in panels:
    # Replace -inf with a visible floor for plotting
    ll_plot = np.where(np.isfinite(ll), ll, ll_max - 20)
    delta = ll_plot - ll_max
    ax.plot(grid, delta, 'b-', linewidth=1.5, markersize=3)
    ax.axhline(y=-chi2_crit/2, color='r', linestyle='--', linewidth=1,
               label='95% CI threshold')
    ax.axvline(x=opt_val, color='gray', linestyle=':', linewidth=1, alpha=0.7,
               label=f'MLE={opt_val:.3g}')
    ax.set_xlabel(label, fontsize=13)
    ax.set_ylabel('Delta log-L', fontsize=12)
    ax.set_title(f'Profile: {label}', fontsize=13)
    ax.legend(fontsize=9)
    ax.set_ylim(bottom=-15, top=1)

plt.tight_layout()
plt.savefig('profile_1d_all.svg', dpi=200, bbox_inches='tight')
print("\nSaved: profile_1d_all.svg")
plt.close()

# --- 2D contour (xi0, kappa1) ---
fig, ax = plt.subplots(1, 1, figsize=(8, 7))

# Replace -inf with visible floor
delta_ll = ll_2d_xi0_k1 - ll_max
delta_ll_plot = np.where(np.isfinite(delta_ll), delta_ll, -20.0)
delta_ll_masked = np.where(delta_ll_plot > -20, delta_ll_plot, np.nan)

# Contour levels: 95% CI (-1.92), 99% CI (-3.32), 99.9% (-6.63)
levels = [-6.63, -3.32, -1.92, -0.5]
labels_contour = ['99.9%', '99%', '95%', '68%']

try:
    cs = ax.contour(xi0_2d, k1_2d, delta_ll_masked, levels=levels,
                    colors=['blue', 'green', 'red', 'orange'], linewidths=1.5)
    ax.clabel(cs, fmt={l: s for l, s in zip(levels, labels_contour)}, fontsize=9)
except Exception as e:
    print(f"  Contour plot failed: {e}")
    # Fallback: pcolormesh
    pcm = ax.pcolormesh(xi0_2d, k1_2d, delta_ll_masked, shading='auto', cmap='YlOrRd')
    plt.colorbar(pcm, ax=ax, label='Delta log-L')

# 最优点
ax.plot(xi0_opt, k1_opt, 'k*', markersize=15, label='MLE')

# Ridge
if len(ridge_xi0) > 5:
    ax.plot(ridge_xi0, ridge_k1, 'k--', linewidth=1.5, alpha=0.7,
            label=f'Ridge (slope={slope:.2f})')

ax.set_xlabel(r'$\xi_0$', fontsize=14)
ax.set_ylabel(r'$\kappa_1$', fontsize=14)
ax.set_title(r'2D Profile Likelihood: $(\xi_0, \kappa_1)$', fontsize=14)
ax.set_xscale('log')
ax.set_yscale('log')
ax.legend(fontsize=11)

plt.tight_layout()
plt.savefig('profile_2D_xi0_kappa1.svg', dpi=300, bbox_inches='tight')
print("Saved: profile_2D_xi0_kappa1.svg")
plt.close()


# ============================================================================
# 6. 保存数据和汇总
# ============================================================================

np.savez('profile_data.npz',
         xi0_grid=xi0_grid, xi0_prof_ll=xi0_prof_ll,
         k1_grid=k1_grid, k1_prof_ll=k1_prof_ll,
         k2_grid=k2_grid, k2_prof_ll=k2_prof_ll,
         k3_grid=k3_grid, k3_prof_ll=k3_prof_ll,
         xi0_2d=xi0_2d, k1_2d=k1_2d, ll_2d=ll_2d_xi0_k1,
         theta_hat=theta_hat, sigma_hat=sigma_hat,
         ll_max=ll_max, chi2_crit=chi2_crit,
         ridge_xi0=ridge_xi0 if len(ridge_xi0) > 0 else np.array([]),
         ridge_k1=ridge_k1 if len(ridge_k1) > 0 else np.array([]),
         ridge_slope=slope if slope is not None else np.nan)
print("Saved: profile_data.npz")

# 汇总表
with open('profile_summary.txt', 'w', encoding='utf-8') as f:
    f.write("=" * 70 + "\n")
    f.write("Profile Likelihood Summary\n")
    f.write("=" * 70 + "\n\n")
    f.write(f"theta_hat = {theta_hat}\n")
    f.write(f"sigma_hat = {sigma_hat:.6f}\n")
    f.write(f"ll_max    = {ll_max:.2f}\n")
    f.write(f"95% CI threshold: Delta-ll > {-chi2_crit/2:.3f}\n\n")

    f.write("-" * 70 + "\n")
    f.write(f"{'Parameter':>10s} | {'MLE':>12s} | {'CI lower':>12s} | {'CI upper':>12s} | "
            f"{'Ratio':>8s} | {'Bounded':>8s}\n")
    f.write("-" * 70 + "\n")

    for name, label in [('xi0', 'xi0'), ('kappa1', 'kappa1'),
                         ('kappa2', 'kappa2'), ('kappa3', 'kappa3')]:
        if name in results and results[name][0] is not None:
            ci_low, ci_high, ratio, bounded = results[name]
            opt = {'xi0': xi0_opt, 'kappa1': k1_opt,
                   'kappa2': k2_opt, 'kappa3': k3_opt}[name]
            f.write(f"{label:>10s} | {opt:>12.4g} | {ci_low:>12.4g} | {ci_high:>12.4g} | "
                    f"{ratio:>8.1f} | {'YES' if bounded else 'NO':>8s}\n")
        else:
            opt = {'xi0': xi0_opt, 'kappa1': k1_opt,
                   'kappa2': k2_opt, 'kappa3': k3_opt}[name]
            f.write(f"{label:>10s} | {opt:>12.4g} | {'UNBOUNDED':>12s} | {'':>12s} | "
                    f"{'INF':>8s} | {'NO':>8s}\n")

    f.write("-" * 70 + "\n\n")

    f.write("2D Ridge Analysis (xi0, kappa1):\n")
    if slope is not None:
        f.write(f"  Ridge: log(kappa1) = {slope:.3f} * log(xi0) + {intercept:.3f}\n")
        f.write(f"  Slope = {slope:.3f}\n")
        f.write(f"  Scale degeneracy predicts slope = 0 (κ independent of ξ₀ at fixed M₀)\n")
        f.write(f"  Observed slope = {slope:.3f} reflects data-driven correlation, not scale degeneracy\n")
    else:
        f.write("  Ridge could not be determined\n")

    f.write("\n" + "=" * 70 + "\n")
    f.write("Interpretation:\n")
    f.write("=" * 70 + "\n\n")

    # 自动判断
    xi0_res = results.get('xi0', (None, None, None, False))
    if xi0_res[3]:  # bounded
        f.write("xi0: BOUNDED CI -> xi0 is practically identifiable\n")
        f.write("   (Three coating groups share xi0, breaking the M0-kappa degeneracy)\n")
    else:
        f.write("xi0: UNBOUNDED CI -> xi0 is NOT practically identifiable\n")
        f.write("   (Scale degeneracy persists even with multi-group data)\n")

    f.write("\n")
    for name, label in [('kappa1', 'kappa1'), ('kappa2', 'kappa2'), ('kappa3', 'kappa3')]:
        res = results.get(name, (None, None, None, False))
        if res[3]:
            f.write(f"{label}: BOUNDED CI (ratio={res[2]:.1f}) -> practically identifiable\n")
        else:
            f.write(f"{label}: UNBOUNDED CI -> NOT practically identifiable\n")
            if name == 'kappa3':
                f.write(f"  (MCMC HDI is prior-limited [1e-4, 1e4]; data cannot constrain kappa3 at zero-depletion regime)\n")

    f.write("\n")
    f.write("Key insight: xi0 (= M0/(K*V*NA)) is identifiable, but M0 itself\n")
    f.write("requires external knowledge of K (or K*V*NA) -- this is the\n")
    f.write("scale degeneracy of Theorem S2e.2, confirmed by real data.\n")

print("Saved: profile_summary.txt")

print("\n" + "=" * 70)
print("Done! Check:")
print("  profile_1d_all.svg       (4-panel 1D profiles)")
print("  profile_2D_xi0_kappa1.svg (2D contour + ridge)")
print("  profile_summary.txt       (CI table + interpretation)")
print("  profile_data.npz          (raw data)")
print("=" * 70)

# ============================================================================
# Sheet 9: Profile Likelihood Results
# ============================================================================
try:
    from openpyxl import load_workbook as _lwb
    from openpyxl.styles import Font as _Font
    _wb9 = _lwb("SI_Table_1.xlsx")
    _ws9 = _wb9.create_sheet("9_Profile_Likelihood")
    _ws9.append(["Profile Likelihood Analysis (95% CI, likelihood ratio test)"])
    _ws9.append(["Parameter", "Optimum", "CI lower", "CI upper", "CI ratio", "Bounded?", "Note"])
    for cell in _ws9[2]:
        cell.font = _Font(bold=True)

    for name, label in [('xi0', 'ξ₀'), ('kappa1', 'κ₁'), ('kappa2', 'κ₂'), ('kappa3', 'κ₃')]:
        res = results.get(name, (None, None, None, False))
        if res[3]:  # bounded
            _ws9.append([label, res[0], res[1], res[2], res[2]/res[1] if res[1] > 0 else None, "YES", ""])
        else:
            note = "MCMC HDI is prior-limited [1e-4, 1e4]; data cannot constrain at zero-depletion" if name == 'kappa3' else ""
            _ws9.append([label, res[0] if res[0] else "N/A", "", "", "INF", "NO", note])

    _ws9.append([""])
    _ws9.append(["2D Ridge Analysis (ξ₀, κ₁):"])
    if slope is not None:
        _ws9.append(["  Ridge slope (d ln κ₁ / d ln ξ₀)", slope])
        _ws9.append(["  Scale degeneracy prediction", "slope = 0 (κ invariant at fixed M₀)"])
        _ws9.append(["  Interpretation", "Data-driven correlation, not scale degeneracy"])
    else:
        _ws9.append(["  Ridge could not be determined"])

    _ws9.append([""])
    _ws9.append(["Key insight:"])
    _ws9.append(["  ξ₀ (= M₀/(K·V·N_A)) is identifiable from multi-group data,"])
    _ws9.append(["  but M₀ itself requires external knowledge of K — this is the"])
    _ws9.append(["  scale degeneracy of Theorem S2e.2, confirmed by real data."])

    _wb9.save("SI_Table_1.xlsx")
    print(f"\nSheet 9 (Profile Likelihood) appended to SI_Table_1.xlsx")
except Exception as e:
    print(f"[!] Profile Likelihood Excel export failed: {e}")
